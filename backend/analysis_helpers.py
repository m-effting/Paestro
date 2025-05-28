"""
PAESTRO - Funções auxiliares para processamento de arquivos de análise

Este módulo contém funções específicas para processamento e
extração de dados de arquivos de análise em formato Excel.
"""

import io
import re
import logging
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from copy import copy
from datetime import datetime
from typing import Dict, Any, List, Optional, Tuple

# Configuração de logging
logger = logging.getLogger(__name__)

def detect_analysis_file(file_name: str, file_content: bytes) -> bool:
    """
    Detecta se um arquivo é um arquivo de análise com base no nome e conteúdo.
    
    Args:
        file_name: Nome do arquivo
        file_content: Conteúdo binário do arquivo
        
    Returns:
        bool: True se for um arquivo de análise, False caso contrário
    """
    # Verifica pelo nome do arquivo
    if "analise" in file_name.lower() or "análise" in file_name.lower():
        return True
        
    # Verifica pelo conteúdo (procura por palavras-chave típicas de arquivos de análise)
    try:
        with io.BytesIO(file_content) as buffer:
            df = pd.read_excel(buffer, nrows=10)
            content_str = ' '.join([str(val) for val in df.values.flatten() if pd.notna(val)])
            keywords = ["análise de frequência", "classificação", "monitorar faltas", 
                        "infrequente", "total de faltas", "%"]
            
            for keyword in keywords:
                if keyword.lower() in content_str.lower():
                    return True
    except:
        pass
        
    return False

def extract_date_from_analysis_file(file_name: str, file_content: bytes) -> Optional[str]:
    """
    Extrai a data de um arquivo de análise.
    
    Args:
        file_name: Nome do arquivo
        file_content: Conteúdo binário do arquivo
        
    Returns:
        str: Data no formato "DD-MM" ou None se não encontrada
    """
    # Tenta extrair do nome do arquivo
    date_match = re.search(r'(\d{2})[_-](\d{2})[_-](\d{4})', file_name)
    if date_match:
        dia, mes, ano = date_match.groups()
        return f"{dia}-{mes}"
    
    # Tenta extrair do conteúdo do arquivo
    try:
        with io.BytesIO(file_content) as buffer:
            df = pd.read_excel(buffer, nrows=10)
            for _, row in df.iterrows():
                row_str = ' '.join([str(val) for val in row.values if pd.notna(val)])
                if "data" in row_str.lower() and ":" in row_str:
                    data_str = row_str.split(":", 1)[1].strip()
                    data_match = re.search(r'(\d{2})/(\d{2})/\d{4}', data_str)
                    if data_match:
                        dia, mes = data_match.groups()
                        return f"{dia}-{mes}"
    except:
        pass
        
    return None

def process_analysis_file(ws, file_info: Dict[str, Any]) -> None:
    """
    Processa um arquivo de análise e adiciona seu conteúdo a uma planilha.
    
    Args:
        ws: Planilha do openpyxl onde o conteúdo será adicionado
        file_info: Informações do arquivo (nome, conteúdo)
    """
    try:
        # Obtém informações do arquivo
        content = file_info.get('content')
        file_name = file_info.get('filename', 'Arquivo desconhecido')
        
        logger.info(f"Processando arquivo de análise: {file_name}")
        
        if not content:
            ws.cell(row=1, column=1).value = "Erro: Conteúdo do arquivo não disponível"
            return
            
        # Tenta usar pandas para processar o arquivo
        try:
            with io.BytesIO(content) as buffer:
                df = pd.read_excel(buffer)
                
                # Cabeçalho
                row = 1
                ws.cell(row=row, column=1).value = "Análise de Frequência Escolar"
                ws.cell(row=row, column=1).font = Font(bold=True, size=14)
                ws.merge_cells(f'A{row}:F{row}')
                row += 1
                
                # Data de geração
                data_hora_atual = datetime.now().strftime('%d/%m/%Y %H:%M')
                ws.cell(row=row, column=1).value = f"Relatório gerado em: {data_hora_atual}"
                ws.cell(row=row, column=1).font = Font(italic=True)
                ws.merge_cells(f'A{row}:F{row}')
                row += 2
                
                # Se o DataFrame não está vazio
                if not df.empty:
                    # Adiciona cabeçalhos das colunas
                    for col_idx, col_name in enumerate(df.columns, 1):
                        ws.cell(row=row, column=col_idx).value = str(col_name)
                        ws.cell(row=row, column=col_idx).font = Font(bold=True)
                        ws.cell(row=row, column=col_idx).fill = PatternFill(
                            start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                    row += 1
                    
                    # Adiciona dados
                    for _, data_row in df.iterrows():
                        for col_idx, value in enumerate(data_row.values, 1):
                            ws.cell(row=row, column=col_idx).value = value
                        row += 1
                
                # Ajusta largura das colunas
                for col in range(1, len(df.columns) + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 15
                    
                logger.info(f"Arquivo de análise processado com sucesso: {file_name}")
                return
                
        except Exception as pandas_error:
            logger.warning(f"Erro ao processar com pandas: {pandas_error}. Tentando método alternativo...")
        
        # Se pandas falhar, tenta usar openpyxl diretamente
        try:
            with io.BytesIO(content) as buffer:
                analysis_wb = openpyxl.load_workbook(buffer)
                analysis_ws = analysis_wb.active
                
                # Título da seção (se não existir)
                if not analysis_ws.cell(row=1, column=1).value or "análise" not in str(analysis_ws.cell(row=1, column=1).value).lower():
                    ws.cell(row=1, column=1).value = "Análise de Frequência Escolar"
                    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
                    ws.merge_cells('A1:F1')
                    row_offset = 2
                else:
                    row_offset = 0
                
                # Copia o conteúdo
                for row in analysis_ws.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            target_row = cell.row + row_offset
                            ws.cell(row=target_row, column=cell.column).value = cell.value
                            
                            # Tenta copiar a formatação
                            try:
                                if cell.font:
                                    ws.cell(row=target_row, column=cell.column).font = copy(cell.font)
                                if cell.fill:
                                    ws.cell(row=target_row, column=cell.column).fill = copy(cell.fill)
                                if cell.alignment:
                                    ws.cell(row=target_row, column=cell.column).alignment = copy(cell.alignment)
                                if cell.border:
                                    ws.cell(row=target_row, column=cell.column).border = copy(cell.border)
                            except:
                                pass
                
                # Ajusta largura das colunas
                for col_idx in range(1, analysis_ws.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    if hasattr(analysis_ws.column_dimensions, col_letter) and analysis_ws.column_dimensions[col_letter].width:
                        ws.column_dimensions[col_letter].width = analysis_ws.column_dimensions[col_letter].width
                
                logger.info(f"Arquivo de análise processado com sucesso via openpyxl: {file_name}")
                return
                
        except Exception as openpyxl_error:
            logger.error(f"Erro ao processar com openpyxl: {openpyxl_error}")
        
        # Se todas as abordagens falharem
        ws.cell(row=1, column=1).value = "Erro: Não foi possível processar o arquivo de análise"
        ws.cell(row=2, column=1).value = f"Arquivo: {file_name}"
    
    except Exception as e:
        logger.error(f"Erro geral ao processar arquivo de análise: {e}")
        ws.cell(row=1, column=1).value = f"Erro ao processar arquivo de análise: {e}"