#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
PAESTRO - Sistema de Gestão de Chamadas Escolares

Este é o script principal do aplicativo PAESTRO,
que fornece ferramentas para importação, análise e exportação
de dados de chamadas escolares.
"""

from flask import Flask, request, jsonify, send_file, render_template
from flask import session, redirect, url_for
from datetime import datetime, timedelta
import os
import io
import sys
import re
import time
import tempfile
from werkzeug.utils import secure_filename

# Adiciona o diretório atual ao path para execução direta do script
sys.path.insert(0, os.path.abspath(os.path.dirname(os.path.dirname(__file__))))

# Importações locais do PAESTRO, com suporte a execução direta ou importada
try:
    # Tenta importar como módulo (run_flask.py importando)
    from backend.chamada_parser import parse_html_content
    from backend.excel_exporter import export_to_excel, get_excel_filename
    from backend.drive_exporter import get_drive_folders, export_attendance_drive 
    from backend.data import app_data, normalize_school_name
except ImportError:
    # Tenta importar para execução direta (python backend/app.py)
    from chamada_parser import parse_html_content
    from excel_exporter import export_to_excel, get_excel_filename
    from drive_exporter import get_drive_folders, export_attendance_drive 
    from data import app_data, normalize_school_name

# Importações do módulo de análise de chamadas
try:
    # Tenta importar como módulo (run_flask.py importando)
    from backend.analysis.analise_parser import analyze_attendance_html as process_html_file
    from backend.analysis.analise_parser import analyze_elementary_file
    from backend.analysis.rules_engine import apply_classification_rules
    from backend.analysis.utils import setup_new_logger
    
    # Configura o logger para análise de faltas
    logger = setup_new_logger()
    print("Módulo de análise carregado com sucesso.")
    
except (ImportError, ModuleNotFoundError) as e:
    try:
        # Tenta importar para execução direta (python backend/app.py)
        from analysis.analise_parser import analyze_attendance_html as process_html_file
        from analysis.analise_parser import analyze_elementary_file
        from analysis.rules_engine import apply_classification_rules
        from analysis.utils import setup_new_logger
        
        # Configura o logger para análise de faltas
        logger = setup_new_logger()
        print("Módulo de análise carregado com sucesso.")
    except (ImportError, ModuleNotFoundError) as e2:
        # Se falhar, tenta importar módulo local com caminhos alternativos
        print(f"Erro ao importar módulo de análise: {e2}")
        try:
            # Tenta importar com caminho relativo (para execução local direta)
            from attendance_analyzer import process_files, get_logs, export_to_file
        except ImportError as e3:
            try:
                # Tenta importar com caminho absoluto (para execução via VS Code/outro IDE)
                from backend.attendance_analyzer import process_files, get_logs, export_to_file
            except ImportError as e4:
                print(f"Não foi possível importar attendance_analyzer: {e4}")
                # Implementação mínima inline para evitar erro fatal
                def process_files(files):
                    return {"data": [], "errors": ["Módulo attendance_analyzer não disponível"]}
                def get_logs():
                    return ["Módulo attendance_analyzer não disponível"]
                def export_to_file(data, file_format):
                    return None, "text/plain", "erro.txt"
        
        # Funções para compatibilidade quando usa o módulo local
        def process_html_file(html_content):
            return process_files([{"content": html_content, "filename": "arquivo.html"}])["data"]
            
        def apply_classification_rules(data):
            return data  # Sem regras no módulo local
            
        def analyze_elementary_file(html_content):
            return {"school_info": {}, "students": []}

app = Flask(__name__,
            template_folder=os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'templates'),
            static_folder=os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'static'))

app.config['TEMPLATES_AUTO_RELOAD'] = True
app.secret_key = 'senha_ultramente_secreta'

# Configurações de sessão: 2 horas de expiração (7200 segundos)
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=2)

# Configura sessão baseada em sistema de arquivos para lidar com cookies grandes
from flask_session import Session
session_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'session_data')

# Cria o diretório de sessão se não existir
os.makedirs(session_dir, exist_ok=True)

app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_FILE_DIR'] = session_dir
app.config['SESSION_PERMANENT'] = True  # Sessões permanentes por padrão
app.config['SESSION_USE_SIGNER'] = True  # Assina os cookies para segurança
app.config['SESSION_FILE_THRESHOLD'] = 500  # Número máximo de arquivos antes de limpeza

# Inicializa o armazenamento de sessão
Session(app)

SENHA_CORRETA = "ProjetoPaestro@2025"


# Endpoint para obter anotações de uma escola
@app.route('/api/get_annotations', methods=['GET'])
def get_annotations():
    escola = request.args.get('escola')
    if not escola:
        return jsonify({'success': False, 'error': 'Escola não especificada'})

    normalized_escola = normalize_school_name(escola)
    annotations = app_data['unit_annotations'].get(normalized_escola, [])
    return jsonify({'success': True, 'annotations': annotations})

# Endpoint para adicionar uma nova anotação
@app.route('/api/add_annotation', methods=['POST'])
def add_annotation():
    data = request.get_json()
    escola = data.get('escola')
    anotacao = data.get('anotacao')

    if not escola or not anotacao:
        return jsonify({'success': False, 'error': 'Dados incompletos'})

    normalized_escola = normalize_school_name(escola)
    if normalized_escola not in app_data['unit_annotations']:
        app_data['unit_annotations'][normalized_escola] = []

    app_data['unit_annotations'][normalized_escola].append(anotacao)
    return jsonify({'success': True})

# Endpoint para excluir uma anotação
@app.route('/api/delete_annotation', methods=['POST'])
def delete_annotation():
    data = request.get_json()
    escola = data.get('escola')
    anotacao = data.get('anotacao')

    if not escola or not anotacao:
        return jsonify({'success': False, 'error': 'Dados incompletos'})

    normalized_escola = normalize_school_name(escola)
    if normalized_escola in app_data['unit_annotations']:
        if anotacao in app_data['unit_annotations'][normalized_escola]:
            app_data['unit_annotations'][normalized_escola].remove(anotacao)
            if not app_data['unit_annotations'][normalized_escola]:
                del app_data['unit_annotations'][normalized_escola]  # Remove chave se lista vazia
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': 'Anotação não encontrada'})
    return jsonify({'success': False, 'error': 'Escola não encontrada'})

@app.route('/')
def home():
    return render_template('index.html', now=datetime.now())

@app.route('/importar')
def import_page():
    """
    Página de Importação.
    Só permite acesso se o usuário estiver autenticado.
    """
    if not session.get("autenticado"):
        return redirect(url_for('home'))
    return render_template('importar.html')

@app.route('/analise')
def analysis_page():
    """
    Página de Análise de Chamadas.
    Só permite acesso se o usuário estiver autenticado.
    Inicializa o armazenamento de arquivos analisados na sessão do usuário.
    """
    # Temporariamente desabilitado para permitir testes
    # if not session.get("autenticado"):
    #     return redirect(url_for('home'))
    
    # Inicializa a lista de arquivos analisados na sessão se não existir
    if 'analyzed_files' not in session:
        session['analyzed_files'] = []
        
    return render_template('analise.html')

@app.route('/chamada')
def attendance_page():
    """
    Página de Chamadas.
    Só permite acesso se o usuário estiver autenticado.
    """
    if not session.get("autenticado"):
        return redirect(url_for('home'))
    return render_template('chamada.html',
                           current_user=app_data['current_user'],
                           current_date=datetime.now().strftime('%d/%m/%Y'))

@app.route('/api/get_saved_classes', methods=['GET'])
def get_saved_classes():
    try:
        escola = request.args.get('escola')
        if escola:
            # Retorna apenas as turmas salvas da escola especificada
            if escola in app_data['saved_classes']:
                return jsonify({
                    'success': True,
                    'saved_classes': list(app_data['saved_classes'][escola])
                })
            else:
                return jsonify({
                    'success': True,
                    'saved_classes': []
                })
        else:
            # Retorna todas as turmas salvas de todas as escolas como uma lista plana
            all_saved = [turma for escola_turmas in app_data['saved_classes'].values() for turma in escola_turmas]
            return jsonify({
                'success': True,
                'saved_classes': list(all_saved)
            })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 1000


@app.route('/api/get_saved_classes_status', methods=['GET'])
def get_saved_classes_status():
    try:
        # Retorna todas as turmas salvas de todas as escolas
        all_saved = [turma for escola_turmas in app_data['saved_classes'].values() for turma in escola_turmas]
        return jsonify({
            'success': True,
            'saved_classes': all_saved
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/login', methods=['POST'])
def login():
    """
    Rota para autenticação do usuário.
    Agora espera os campos:
     - username: Nome da dupla
     - periodo: Período selecionado
     - senha: Chave de acesso
    Se a senha estiver correta, o usuário é autenticado na sessão.
    """
    data = request.get_json()
    username = data.get('username')
    periodo = data.get('periodo')
    senha = data.get('senha')

    if senha != SENHA_CORRETA:
        return jsonify({'success': False, 'error': 'Senha incorreta!'}), 401

    # Se a senha estiver correta, registra o usuário
    session.permanent = True  # Torna a sessão permanente com duração definida nas configurações
    session['autenticado'] = True
    session['current_user'] = username
    session['periodo'] = periodo
    app_data['current_user'] = username
    app_data['periodo'] = periodo

    return jsonify({
        'success': True, 
        'username': app_data['current_user'],
        'periodo': app_data['periodo']
    })

@app.route('/api/upload', methods=['POST'])
def handle_file_upload():
    if 'files' not in request.files:
        return jsonify({'success': False, 'error': 'Nenhum arquivo enviado'})

    files = request.files.getlist('files')
    if not files:
        return jsonify({'success': False, 'error': 'Nenhum arquivo selecionado'})

    try:
        for file in files:
            if file.filename == '':
                continue

            try:
                html_content = file.read().decode('utf-8')
            except UnicodeDecodeError:
                return jsonify({'success': False, 'error': 'Erro de codificação no arquivo'})

            # Passa o nome do arquivo para o parser como fallback
            classes_dict, unidade_name = parse_html_content(html_content, file.filename)

            # Remove caracteres problemáticos mas mantém acentos
            school_name = re.sub(r'[\\/*?:"<>|]', '', unidade_name).strip()
            school_name = ' '.join(school_name.split())

            if not school_name:
                school_name = os.path.splitext(file.filename)[0]

            if school_name not in app_data['schools']:
                app_data['schools'][school_name] = {}

            app_data['schools'][school_name].update(classes_dict)
            app_data['html_content'][school_name] = html_content
            if not school_name:
                school_name = os.path.splitext(file.filename)[0]

            for turma, alunos in classes_dict.items():
                if turma not in app_data['attendance_status']:
                    app_data['attendance_status'][turma] = {}
                if turma not in app_data['observations']:
                    app_data['observations'][turma] = {}

                for aluno in alunos:
                    if aluno not in app_data['attendance_status'][turma]:
                        app_data['attendance_status'][turma][aluno] = 'P'
                    if aluno not in app_data['observations'][turma]:
                        app_data['observations'][turma][aluno] = ''

        app_data['file_uploaded'] = True
        return jsonify({
            'success': True,
            'schools': list(app_data['schools'].keys()),
            'message': 'Arquivos processados com sucesso'
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_current_user', methods=['GET'])
def get_current_user():
    # Prioriza informações da sessão, com fallback para app_data se não presente
    username = session.get('current_user') or app_data.get('current_user', '')
    periodo = session.get('periodo') or app_data.get('periodo', '')
    
    return jsonify({
        'success': True,
        'username': username,
        'periodo': periodo
    })

@app.route('/api/get_imported_files', methods=['GET'])
def get_imported_files():
    try:
        files = [{'name': escola} for escola in app_data['schools'].keys()]
        return jsonify({'success': True, 'files': files})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/delete_file', methods=['POST'])
def delete_file():
    try:
        data = request.json
        filename = data.get('filename')

        if not filename:
            return jsonify({'success': False, 'error': 'Nome do arquivo não fornecido'})

        if filename in app_data['schools']:
            # Captura as turmas associadas antes de deletar
            turmas_da_escola = set(app_data['schools'][filename].keys())

            # Remove a entrada da escola em saved_classes, se existir
            if filename in app_data['saved_classes']:
                del app_data['saved_classes'][filename]

            # Remove a escola e seu conteúdo
            del app_data['schools'][filename]
            del app_data['html_content'][filename]

            # Remove turmas associadas de attendance_status e observations
            for turma in turmas_da_escola:
                app_data['attendance_status'].pop(turma, None)
                app_data['observations'].pop(turma, None)

            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'Arquivo não encontrado'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_schools', methods=['GET'])
def get_schools():
    return jsonify({
        'success': True,
        'schools': list(app_data['schools'].keys())
    })

@app.route('/api/get_school_classes', methods=['POST'])
def get_school_classes():
    data = request.get_json()
    school = data.get('school')

    if not school or school not in app_data['schools']:
        return jsonify({'success': False, 'error': 'Escola não encontrada'})

    return jsonify({
        'success': True,
        'classes': list(app_data['schools'][school].keys()),
        'saved_classes': list(app_data['saved_classes'])
    })

@app.route('/api/get_class', methods=['POST'])
def get_class_data():
    data = request.get_json()
    school = data.get('school')
    turma = data.get('class')

    if not school or school not in app_data['schools']:
        return jsonify({'success': False, 'error': 'Escola não encontrada'})

    if turma not in app_data['schools'][school]:
        return jsonify({'success': False, 'error': 'Turma não encontrada'})

    alunos_originais = app_data['schools'][school][turma]
    alunos_data = []

    for aluno in alunos_originais:
        alunos_data.append({
            'nome': aluno,
            'presenca': app_data['attendance_status'].get(turma, {}).get(aluno, 'P'),
            'observacao': app_data['observations'].get(turma, {}).get(aluno, '')
        })

    return jsonify({
        'success': True,
        'alunos': alunos_data,
        'turma': turma,
        'total_alunos': len(alunos_data)
    })

@app.route('/api/get_turmas', methods=['GET'])
def get_turmas():
    return jsonify({
        'success': True,
        'turmas': list(app_data.get('classes', {}).keys())
    })

# Endpoints para a análise de frequência

@app.route('/api/analyze', methods=['POST'])
def process_analysis_files():
    """
    Processa arquivos HTML para análise de faltas e classifica os alunos.
    Integração com módulo analise_chamadas.
    Armazena os resultados na sessão do usuário, não compartilhando entre usuários.
    """
    # Temporariamente desabilitado para permitir testes
    # if not session.get("autenticado"):
    #     return jsonify({'success': False, 'error': 'Usuário não autenticado'}), 401
    
    # Inicializar storage de arquivos analisados na sessão se não existir
    if 'analyzed_files' not in session:
        session['analyzed_files'] = []
        
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Nenhum arquivo enviado'})
        
    files = request.files.getlist('file')
    if not files or files[0].filename == '':
        return jsonify({'success': False, 'error': 'Nenhum arquivo selecionado'})
        
    try:
        # Lista para armazenar os resultados de cada arquivo
        all_results = []
        processed_files = []
        error_files = []
        
        for file in files:
            try:
                html_content = file.read().decode('utf-8')
                
                # Usamos o arquivo analise_parser.py para análise, passando o nome do arquivo
                try:
                    # Passamos o nome do arquivo para que o parser possa extrair informações da turma
                    result = analyze_elementary_file(html_content, file.filename)
                    logger.info(f"Processou o arquivo {file.filename} com {len(result['students'])} alunos usando analyze_elementary_file")
                except Exception as parser_error:
                    logger.error(f"Erro ao analisar o arquivo {file.filename}: {parser_error}")
                    raise
                
                # Extrair os alunos da estrutura de resultado
                students_data = result.get('students', [])
                
                # Adiciona o nome do arquivo como origem para cada aluno
                for student in students_data:
                    student['origem_arquivo'] = file.filename
                
                # Aplica as regras de classificação aos dados dos alunos
                classified_data = apply_classification_rules({'students': students_data})
                
                all_results.extend(classified_data)
                processed_files.append(file.filename)
                
                # Registrar informações detalhadas sobre faltas por mês para os primeiros alunos (log)
                for idx, student in enumerate(students_data[:3]):  # Mostra apenas os primeiros 3 alunos para log
                    if isinstance(student, dict):
                        logger.info(f"Exemplo de dados de aluno {idx+1}: {student.get('aluno', 'N/A')}")
                        
                        # Garantir que valores P, F, FJ sejam sempre exibidos como números, nunca como 'N/A'
                        p_value = student.get('P')
                        if p_value is None or (isinstance(p_value, str) and p_value == 'N/A'):
                            p_value = 0
                            
                        f_value = student.get('F')
                        if f_value is None or (isinstance(f_value, str) and f_value == 'N/A'):
                            f_value = 0
                            
                        fj_value = student.get('FJ')
                        if fj_value is None or (isinstance(fj_value, str) and fj_value == 'N/A'):
                            fj_value = 0
                            
                        logger.info(f"  - P={p_value}, F={f_value}, FJ={fj_value}")
                        
                        # Exibir faltas por mês
                        faltas_texto = student.get('faltas_por_mes_texto')
                        if not faltas_texto or faltas_texto == 'N/A':
                            faltas_texto = "Sem faltas"
                        logger.info(f"  - Faltas por mês: {faltas_texto}")
                        
                        # Exibir maior falta mensal
                        if 'maior_falta_mensal' in student:
                            maior_falta = student.get('maior_falta_mensal')
                            if maior_falta is None or (isinstance(maior_falta, str) and maior_falta == 'N/A'):
                                maior_falta = 0
                            logger.info(f"  - Maior falta mensal: {maior_falta}")
                            
                        # Exibir percentual de presença
                        perc_presenca = student.get('percentual_presenca')
                        if perc_presenca is None or (isinstance(perc_presenca, str) and perc_presenca == 'N/A'):
                            perc_presenca = 0
                        logger.info(f"  - Percentual de presença: {perc_presenca}%")
                
            except Exception as file_error:
                logger.error(f"Erro ao processar arquivo {file.filename}: {file_error}")
                error_files.append({
                    'name': file.filename,
                    'error': str(file_error)
                })
        
        # Criar resumo para o front-end
        summary = {
            'total_students': len(all_results),
            'total_schools': len(set(item.get('escola', 'Desconhecida') for item in all_results)),
            'total_classes': len(set(item.get('turma', 'Desconhecida') for item in all_results)),
            'total_absentees': len([item for item in all_results if 'Faltoso' in item.get('status', []) or 'Faltoso' in item.get('situacao', [])]),
            'total_monitors': len([item for item in all_results if 
                any(status in ['Monitorar Faltas', 'Monitorar FJs'] for status in (item.get('status', []) if isinstance(item.get('status', []), list) else [item.get('status', '')]))
            ])
        }
        
        # Armazenar arquivos processados na sessão do usuário
        for idx, filename in enumerate(processed_files):
            # Verifica se o arquivo já existe na lista da sessão
            file_exists = False
            file_id = None
            for existing_file in session['analyzed_files']:
                if existing_file['filename'] == filename:
                    file_exists = True
                    file_id = existing_file['id']
                    break
            
            # Filtrar estudantes deste arquivo
            file_students = [student for student in all_results if student.get('origem_arquivo', '') == filename]
            
            # Se o arquivo já existe, atualiza; senão, cria novo
            if file_exists:
                for i, file in enumerate(session['analyzed_files']):
                    if file['id'] == file_id:
                        # Extrair o nome da escola (primeiro aluno ou escola mais comum)
                        escolas = {}
                        escola_nome = "Desconhecida"
                        for student in file_students:
                            nome_escola = student.get('escola') or student.get('unidade') or student.get('school_name')
                            if nome_escola:
                                escolas[nome_escola] = escolas.get(nome_escola, 0) + 1
                        
                        if escolas:
                            # Encontra a escola mais comum no arquivo
                            escola_nome = max(escolas.items(), key=lambda x: x[1])[0]
                        
                        import json
                        import gzip
                        import base64
                        
                        # Função para comprimir dados grandes
                        def compress_data(data):
                            # Converte para JSON e depois comprime
                            json_str = json.dumps(data)
                            compressed = gzip.compress(json_str.encode('utf-8'))
                            # Converte para base64 para armazenamento seguro na sessão
                            return base64.b64encode(compressed).decode('utf-8')
                        
                        # Comprimir resultados para não exceder limite de tamanho do cookie
                        compressed_results = compress_data(file_students)
                        
                        session['analyzed_files'][i] = {
                            'id': file_id,
                            'filename': filename,
                            'school_name': escola_nome,
                            'date': datetime.now().isoformat(),
                            'student_count': len(file_students),
                            'compressed_results': compressed_results,  # Armazena em formato comprimido
                            'summary': {
                                'total_students': len(file_students),
                                'total_schools': len(escolas),
                                'total_classes': len(set(item.get('turma', 'Desconhecida') for item in file_students)),
                                'total_absentees': len([item for item in file_students if 
                                                       (isinstance(item.get('status', []), list) and 'Faltoso' in item.get('status', [])) or
                                                       (isinstance(item.get('status', ''), str) and 'Faltoso' in item.get('status', '')) or
                                                       (isinstance(item.get('situacao', []), list) and 'Faltoso' in item.get('situacao', [])) or
                                                       (isinstance(item.get('situacao', ''), str) and 'Faltoso' in item.get('situacao', ''))
                                                      ]),
                                'total_monitors': len([item for item in file_students if 
                                                       (isinstance(item.get('status', []), list) and 
                                                        any(s in ['Monitorar Faltas', 'Monitorar FJs'] for s in item.get('status', []))) or
                                                       (isinstance(item.get('status', ''), str) and 
                                                        any(s in item.get('status', '') for s in ['Monitorar Faltas', 'Monitorar FJs'])) or
                                                       (isinstance(item.get('situacao', []), list) and 
                                                        any(s in ['Monitorar Faltas', 'Monitorar FJs'] for s in item.get('situacao', []))) or
                                                       (isinstance(item.get('situacao', ''), str) and 
                                                        any(s in item.get('situacao', '') for s in ['Monitorar Faltas', 'Monitorar FJs']))
                                                      ])
                            }
                        }
                        # Marcar sessão como modificada
                        session.modified = True
                        break
            else:
                # Gera um ID único para o arquivo
                new_id = str(int(time.time())) + str(len(session['analyzed_files']))
                
                # Extrair o nome da escola e turma (mais comuns no arquivo)
                escolas = {}
                turmas = {}
                escola_nome = "Desconhecida"
                turma_nome = "Desconhecida"
                
                for student in file_students:
                    nome_escola = student.get('escola') or student.get('unidade') or student.get('school_name')
                    nome_turma = student.get('turma') or student.get('class_name') or ""
                    
                    if nome_escola:
                        escolas[nome_escola] = escolas.get(nome_escola, 0) + 1
                    if nome_turma:
                        turmas[nome_turma] = turmas.get(nome_turma, 0) + 1
                
                if escolas:
                    # Encontra a escola mais comum no arquivo
                    escola_nome = max(escolas.items(), key=lambda x: x[1])[0]
                    
                if turmas:
                    # Encontra a turma mais comum no arquivo
                    turma_nome = max(turmas.items(), key=lambda x: x[1])[0]
                
                import json
                import gzip
                import base64
                
                # Função para comprimir dados grandes
                def compress_data(data):
                    # Converte para JSON e depois comprime
                    json_str = json.dumps(data)
                    compressed = gzip.compress(json_str.encode('utf-8'))
                    # Converte para base64 para armazenamento seguro na sessão
                    return base64.b64encode(compressed).decode('utf-8')
                
                # Comprimir resultados para não exceder limite de tamanho do cookie
                compressed_results = compress_data(file_students)
                
                # Adiciona arquivo à lista de arquivos analisados na sessão
                session['analyzed_files'].append({
                    'id': new_id,
                    'filename': filename,
                    'school_name': escola_nome,
                    'class_name': turma_nome,
                    'date': datetime.now().isoformat(),
                    'student_count': len(file_students),
                    'compressed_results': compressed_results,  # Armazena em formato comprimido
                    'summary': {
                        'total_students': len(file_students),
                        'total_schools': len(escolas),
                        'total_classes': len(set(item.get('turma', 'Desconhecida') for item in file_students)),
                        'total_absentees': len([item for item in file_students if 
                                               (isinstance(item.get('status', []), list) and 'Faltoso' in item.get('status', [])) or
                                               (isinstance(item.get('status', ''), str) and 'Faltoso' in item.get('status', '')) or
                                               (isinstance(item.get('situacao', []), list) and 'Faltoso' in item.get('situacao', [])) or
                                               (isinstance(item.get('situacao', ''), str) and 'Faltoso' in item.get('situacao', ''))
                                              ]),
                        'total_monitors': len([item for item in file_students if 
                                               (isinstance(item.get('status', []), list) and 
                                                any(s in ['Monitorar Faltas', 'Monitorar FJs'] for s in item.get('status', []))) or
                                               (isinstance(item.get('status', ''), str) and 
                                                any(s in item.get('status', '') for s in ['Monitorar Faltas', 'Monitorar FJs'])) or
                                               (isinstance(item.get('situacao', []), list) and 
                                                any(s in ['Monitorar Faltas', 'Monitorar FJs'] for s in item.get('situacao', []))) or
                                               (isinstance(item.get('situacao', ''), str) and 
                                                any(s in item.get('situacao', '') for s in ['Monitorar Faltas', 'Monitorar FJs']))
                                              ])
                    }
                })
                
                # Marcar sessão como modificada
                session.modified = True
        
        return jsonify({
            'success': True,
            'results': all_results,  # Mantendo a nomenclatura esperada pelo front-end
            'summary': summary,
            'processed_files': processed_files,
            'error_files': error_files
        })
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return jsonify({'success': False, 'error': str(e), 'details': error_details})

@app.route('/api/get_analyzed_files', methods=['GET'])
def get_analyzed_files():
    """
    Obtém a lista de arquivos já analisados da sessão do usuário.
    Implementa proteção contra expiração da sessão.
    """
    try:
        # Verifica se o storage de arquivos existe na sessão
        if 'analyzed_files' not in session:
            session['analyzed_files'] = []
            # Renova sessão ao acessar esta rota
            session.permanent = True
            session.modified = True
        
        # Sempre que obter arquivos analisados, atualiza o timestamp da sessão
        # para evitar que ela expire durante a análise
        session.modified = True
        
        return jsonify({
            'success': True,
            'files': [{
                'id': file['id'],
                'filename': file['filename'],
                'school_name': file.get('school_name', 'Desconhecida'),
                'class_name': file.get('class_name', 'Desconhecida'),
                'date': file['date'],
                'student_count': file.get('student_count', len(file.get('results', []))),
                'summary': file.get('summary', {})
            } for file in session['analyzed_files']]
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/delete_analyzed_file/<file_id>', methods=['DELETE'])
def delete_analyzed_file(file_id):
    """
    Exclui um arquivo analisado da lista de arquivos da sessão do usuário.
    Implementa proteção contra expiração da sessão.
    """
    if not file_id:
        return jsonify({'success': False, 'error': 'ID do arquivo não fornecido'})
    
    try:
        # Verifica se o storage de arquivos existe na sessão
        if 'analyzed_files' not in session:
            session['analyzed_files'] = []
            # Renova sessão ao acessar esta rota
            session.permanent = True
            session.modified = True
            return jsonify({'success': False, 'error': 'Arquivo não encontrado'})
        
        # Busca e remove o arquivo pelo ID na sessão
        for idx, file in enumerate(session['analyzed_files']):
            if file['id'] == file_id:
                session['analyzed_files'].pop(idx)
                # Sempre que modificar a sessão, atualiza seu timestamp
                session.permanent = True
                session.modified = True
                return jsonify({'success': True, 'message': 'Arquivo removido com sucesso'})
        
        return jsonify({'success': False, 'error': 'Arquivo não encontrado'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_analyzed_file_content/<file_id>', methods=['GET'])
def get_analyzed_file_content(file_id):
    """
    Obtém o conteúdo de um arquivo analisado pelo ID da sessão do usuário.
    Implementa proteção contra expiração da sessão.
    """
    if not file_id:
        return jsonify({'success': False, 'error': 'ID do arquivo não fornecido'})
    
    try:
        # Verifica se o storage de arquivos existe na sessão
        if 'analyzed_files' not in session:
            session['analyzed_files'] = []
            # Renova sessão ao acessar esta rota
            session.permanent = True
            session.modified = True
            return jsonify({'success': False, 'error': 'Arquivo não encontrado'})
        
        # Funções para compressão e descompressão
        import json
        import gzip
        import base64
        
        def decompress_data(compressed_str):
            if not compressed_str:
                return []
            # Decodifica base64, descomprime e carrega JSON
            decoded = base64.b64decode(compressed_str.encode('utf-8'))
            decompressed = gzip.decompress(decoded).decode('utf-8')
            return json.loads(decompressed)
        
        # Busca o conteúdo do arquivo pelo ID na sessão
        for file in session['analyzed_files']:
            if file['id'] == file_id:
                # Sempre que buscar um arquivo, atualiza o timestamp da sessão
                # para evitar que ela expire durante a análise
                session.modified = True
                
                # Verifica se há dados comprimidos e faz a descompressão
                results = []
                if 'compressed_results' in file:
                    results = decompress_data(file['compressed_results'])
                elif 'results' in file:
                    results = file.get('results', [])
                    
                return jsonify({
                    'success': True, 
                    'results': results,
                    'summary': file.get('summary', {})
                })
        
        return jsonify({'success': False, 'error': 'Arquivo não encontrado'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/logs', methods=['GET'])
def get_analysis_logs():
    """
    Obtém os logs recentes do analisador de frequência.
    """
    # Temporariamente desabilitado para permitir testes
    # if not session.get("autenticado"):
    #     return jsonify({'success': False, 'error': 'Usuário não autenticado'}), 401
    
    try:
        # Lê os logs do arquivo
        with open('attendance_parser.log', 'r') as log_file:
            logs = log_file.readlines()
            # Limita para as últimas 100 linhas para não sobrecarregar
            logs = logs[-100:]
            return jsonify({'success': True, 'logs': logs})
    except Exception as e:
        return jsonify({'success': False, 'error': f'Erro ao ler logs: {str(e)}'})

@app.route('/api/download', methods=['POST'])
def download_analysis_file():
    """
    Exporta os dados de frequência para um arquivo CSV ou Excel.
    """
    if not session.get("autenticado"):
        return jsonify({'success': False, 'error': 'Usuário não autenticado'}), 401
        
    try:
        data = request.get_json()
        format_type = data.get('format', 'excel')
        
        if not data.get('data'):
            return jsonify({'success': False, 'error': 'Nenhum dado para exportar'})
            
        # Construímos um buffer de memória para o arquivo
        import io
        import pandas as pd
        
        # Preparamos os dados por turma para o formato especificado
        alunos_por_turma = {}
        
        # Filtrar apenas alunos com status diferente de "Regular"
        alunos_nao_regulares = []
        for aluno in data['data']:
            status = aluno.get('status') or aluno.get('situacao') or 'Regular'
            if isinstance(status, list):
                status_normalizado = status
            else:
                status_normalizado = [str(status)]
            
            # Verifica se não é Regular (status deve conter alguma classificação como Faltoso ou Monitorar)
            if any(s != 'Regular' for s in status_normalizado) and not all(s == 'Regular' for s in status_normalizado):
                alunos_nao_regulares.append(aluno)
        
        # Agrupar alunos por turma
        for aluno in alunos_nao_regulares:
            turma = aluno.get('turma') or aluno.get('class_name') or 'Sem Turma'
            if turma not in alunos_por_turma:
                alunos_por_turma[turma] = []
            alunos_por_turma[turma].append(aluno)
        
        buffer = io.BytesIO()
        
        if format_type == 'csv':
            # Criar DataFrame e exportar para CSV
            df = pd.DataFrame(data['data'])
            df.to_csv(buffer, index=False, encoding='utf-8-sig')
            mime_type = 'text/csv'
            filename = f"analise_frequencia_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        else:
            # Uso do openpyxl para criar um Excel formatado conforme especificação
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Análise de Faltas"
            
            # Configuração de estilos
            title_font = Font(name='Arial', size=12, bold=True)
            header_font = Font(name='Arial', size=11, bold=True)
            data_font = Font(name='Arial', size=10)
            center_alignment = Alignment(horizontal='center', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center')
            thin_border = Side(style='thin')
            border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
            
            # Cabeçalho com nome da escola e data
            escolas = set()
            for aluno in data['data']:
                escola = aluno.get('escola') or aluno.get('school_name') or aluno.get('unidade') or 'Desconhecida'
                escolas.add(escola)
            
            # Configuração do fuso horário de Brasília
            import pytz
            tz_brasil = pytz.timezone('America/Sao_Paulo')
            data_hora_brasil = datetime.now(tz_brasil).strftime('%d/%m/%Y %H:%M')
            
            escola_str = ", ".join(escolas) if len(escolas) <= 3 else f"{len(escolas)} escolas"
            data_hora = data_hora_brasil  # Usa o horário de Brasília
            
            ws['A1'] = f"Escola(s): {escola_str}"
            ws['A1'].font = title_font
            ws.merge_cells('A1:I1')
            
            ws['A2'] = f"Data/Hora: {data_hora}"
            ws['A2'].font = header_font
            ws.merge_cells('A2:I2')
            
            # Linha em branco após o cabeçalho
            current_row = 4
            
            # Colunas padrão para a tabela de alunos
            columns = [
                "Aluno", "Status", "% Presença", "Total P", "Total F", "Total FJ", 
                "F por mês", "Nºs de contato", "Data do contato", "Motivo das faltas"
            ]
            
            # Processar cada turma separadamente
            for turma, alunos in alunos_por_turma.items():
                # Título da turma (centralizado)
                ws.cell(row=current_row, column=1, value=f"Turma: {turma}")
                ws.cell(row=current_row, column=1).font = title_font
                ws.cell(row=current_row, column=1).alignment = center_alignment
                ws.merge_cells(f'A{current_row}:J{current_row}')
                current_row += 1
                
                # Cabeçalho das colunas
                for col_idx, col_name in enumerate(columns, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=col_name)
                    cell.font = header_font
                    cell.alignment = center_alignment
                    cell.border = border
                current_row += 1
                
                # Dados dos alunos da turma
                for aluno in alunos:
                    # Identificando os campos do aluno
                    nome = aluno.get('aluno') or aluno.get('student_name') or 'Sem nome'
                    
                    # Formatar o status
                    status = aluno.get('status') or aluno.get('situacao') or 'Regular'
                    if isinstance(status, list):
                        status = ", ".join(status)
                    
                    percentual = aluno.get('percentual_presenca') or aluno.get('attendance_percentage') or 0
                    total_p = aluno.get('P') or aluno.get('presence_total') or 0
                    total_f = aluno.get('F') or aluno.get('absence_total') or 0
                    total_fj = aluno.get('FJ') or aluno.get('justified_total') or 0
                    
                    # Formatando F por mês
                    faltas_mensais = ""
                    if 'faltas_por_mes_texto' in aluno and aluno['faltas_por_mes_texto']:
                        faltas_mensais = aluno['faltas_por_mes_texto']
                    elif 'faltas_por_mes_formatado' in aluno and aluno['faltas_por_mes_formatado']:
                        faltas_por_mes = aluno['faltas_por_mes_formatado']
                        faltas_mensais = ", ".join([f"{mes}: {qtd}" for mes, qtd in faltas_por_mes.items()])
                    elif 'faltas_por_mes' in aluno and aluno['faltas_por_mes']:
                        faltas_por_mes = aluno['faltas_por_mes']
                        faltas_mensais = ", ".join([f"{mes}: {qtd}" for mes, qtd in faltas_por_mes.items()])
                    
                    # Inserir os dados na linha
                    values = [nome, status, f"{percentual}%", total_p, total_f, total_fj, faltas_mensais, "", "", ""]
                    for col_idx, value in enumerate(values, start=1):
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.font = data_font
                        cell.alignment = left_alignment if col_idx in [1, 2, 7, 10] else center_alignment
                        cell.border = border
                    
                    current_row += 1
                
                # Linha em branco entre turmas
                current_row += 1
            
            # Função para ajustar automaticamente as colunas ao tamanho do conteúdo
            from openpyxl.utils import get_column_letter
            
            # Larguras iniciais mínimas para cada coluna (para garantir um tamanho mínimo razoável)
            min_width = {
                'A': 25,  # Aluno
                'B': 15,  # Status
                'C': 12,  # % Presença
                'D': 10,  # Total P
                'E': 10,  # Total F
                'F': 10,  # Total FJ
                'G': 25,  # F por mês
                'H': 15,  # Nºs de contato
                'I': 15,  # Data do contato
                'J': 25,  # Motivo das faltas
            }
            
            # Define as larguras iniciais mínimas
            for col in range(1, 11):  # Colunas A a J
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = min_width.get(col_letter, 12)
                
            # Ajuste fino: adiciona um pouco de espaço extra para texto mais longo
            # Percorre todas as células de dados e amplia a coluna se necessário
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        col_letter = get_column_letter(cell.column)
                        # Cálculo aproximado considerando o tamanho do texto + um pouco de margem
                        # Usa o tamanho da fonte para ajustar a largura
                        font_size_factor = 0.9  # fator para ajuste de fonte
                        cell_length = len(str(cell.value)) * font_size_factor
                        
                        # Adiciona espaço para conteúdo mais longo nas colunas chave
                        if col_letter in ['A', 'B', 'G', 'J'] and cell_length > ws.column_dimensions[col_letter].width:
                            # Limita a largura máxima para não ficar muito extenso
                            max_width = min(cell_length, 50) 
                            ws.column_dimensions[col_letter].width = max_width
            
            wb.save(buffer)
            mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            # Usa o horário de Brasília também para o nome do arquivo
            filename = f"analise_frequencia_{datetime.now(tz_brasil).strftime('%Y%m%d_%H%M%S')}.xlsx"
            
        buffer.seek(0)
        return send_file(
            buffer,
            mimetype=mime_type,
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return jsonify({'success': False, 'error': str(e), 'details': error_details}), 500

@app.route('/api/save_attendance', methods=['POST'])
def save_attendance_data():
    data = request.json
    escola = data.get('escola')
    turma = data.get('turma')
    alunos = data.get('alunos')

    if not all([escola, turma, alunos]):
        return jsonify({'success': False, 'error': 'Dados incompletos'})

    try:
        if escola not in app_data['schools']:
            app_data['schools'][escola] = {}

        if turma not in app_data['schools'][escola]:
            app_data['schools'][escola][turma] = []

        if turma not in app_data['attendance_status']:
            app_data['attendance_status'][turma] = {}
        if turma not in app_data['observations']:
            app_data['observations'][turma] = {}

        for aluno in alunos:
            nome = aluno['nome']
            presenca = aluno['presenca']
            observacao = aluno['observacao']

            # Adiciona o aluno à lista da turma, se ainda não estiver presente
            if nome not in app_data['schools'][escola][turma]:
                app_data['schools'][escola][turma].append(nome)

            # Atualiza presença e observação no backend
            app_data['attendance_status'][turma][nome] = presenca
            app_data['observations'][turma][nome] = observacao

        # Adiciona a turma ao conjunto de turmas salvas da escola específica
        if escola not in app_data['saved_classes']:
            app_data['saved_classes'][escola] = set()
        app_data['saved_classes'][escola].add(turma)

        return jsonify({'success': True})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/clear_saved_classes', methods=['POST'])
def clear_saved_classes():
    try:
        print("Iniciando limpeza de turmas salvas...")
        print(f"Turmas salvas antes: {app_data['saved_classes']}")

        app_data['saved_classes'].clear()

        print(f"Turmas salvas depois: {app_data['saved_classes']}")
        print("Limpeza concluída com sucesso!")

        return jsonify({
            'success': True,
            'message': 'Todas as turmas salvas foram removidas',
            'saved_classes': list(app_data['saved_classes'])
        })
    except Exception as e:
        print(f"ERRO ao limpar turmas: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/exportar')
def export_page():
    """
    Página de Exportação.
    Só permite acesso se o usuário estiver autenticado.
    """
    if not session.get("autenticado"):
        return redirect(url_for('home'))
    escola = request.args.get('escola', '')
    return render_template('exportar.html',
                           escola=escola,
                           current_user=app_data['current_user'],
                           current_date=datetime.now().strftime('%d/%m/%Y'))

@app.route('/relatorio')
def relatorio_page():
    """
    Página de Relatório Consolidado.
    Só permite acesso se o usuário estiver autenticado.
    """
    if not session.get("autenticado"):
        return redirect(url_for('home'))
    return render_template('relatorio.html')

@app.route('/process_report_files', methods=['POST'])
def process_report_files():
    """
    Processa arquivos Excel para a geração de relatório consolidado.
    Extrai metadados e estrutura para apresentação na interface.
    """
    if not session.get("autenticado"):
        return jsonify({'error': 'Não autenticado'}), 401

    try:
        # Limpar arquivos anteriores se existirem
        if 'report_files' in session:
            # Remover arquivos temporários anteriores
            for file_info in session.get('report_files', []):
                if 'temp_path' in file_info and os.path.exists(file_info['temp_path']):
                    try:
                        os.unlink(file_info['temp_path'])
                    except Exception as e:
                        app.logger.warning(f"Não foi possível excluir o arquivo temporário: {e}")
            
        # Inicializar lista de arquivos na sessão
        session['report_files'] = []
        
        # Lista para armazenar os arquivos e seus conteúdos
        files_data = []
        
        # Verificar se há arquivos na requisição
        if 'files' not in request.files:
            app.logger.warning("Nenhum arquivo foi enviado na requisição")
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        # Processa cada arquivo enviado
        for file in request.files.getlist('files'):
            if file and file.filename.endswith(('.xlsx', '.xls')):
                try:
                    # Lê o conteúdo do arquivo
                    file_content = file.read()
                    
                    # Salva temporariamente o arquivo para processamento
                    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                    temp_file.write(file_content)
                    temp_file.close()
                    
                    # Adiciona à lista de arquivos processados
                    session['report_files'].append({
                        'filename': file.filename,
                        'temp_path': temp_file.name
                    })
                    
                    files_data.append({
                        'filename': file.filename,
                        'content': file_content
                    })
                    
                    app.logger.info(f"Arquivo processado e salvo: {file.filename}")
                except Exception as e:
                    app.logger.error(f"Erro ao processar arquivo {file.filename}: {str(e)}")
                    continue
        
        # Marcar a sessão como modificada para garantir que seja salva
        session.modified = True
        
        if not files_data:
            app.logger.warning("Nenhum arquivo Excel válido foi processado")
            return jsonify({'error': 'Nenhum arquivo Excel válido foi enviado'}), 400
        
        # Os arquivos já foram armazenados na sessão, não precisamos refazer
        for file_info in files_data:
            # Armazena apenas os metadados, não o conteúdo binário
            session['report_files'].append({
                'filename': file_info['filename'],
                # Armazena o conteúdo do arquivo em um arquivo temporário
                'temp_path': save_temp_file(file_info['content'])
            })
        
        # Extrai metadados básicos dos arquivos
        result = {
            'files': [],
            'total_alunos': 0,
            'total_turmas': 0,
            'datas_visita': [],
            'escola': None
        }
        
        for file_data in files_data:
            result['files'].append({
                'nome': file_data['filename'],
                'tipo': 'Análise' if 'analise' in file_data['filename'].lower() else 'Chamada',
                'data': '27/05/2025',  # Placeholder - será extraído do arquivo
                'alunos': 50,  # Placeholder - será calculado
                'turmas': 1,   # Placeholder - será calculado
                'escola': None
            })
        
        result['total_alunos'] = len(result['files']) * 50
        result['total_turmas'] = len(result['files'])
        result['datas_visita'] = ['27/05/2025']
        
        return jsonify(result)
    
    except Exception as e:
        logger.error(f"Erro ao processar arquivos para relatório: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/generate_consolidated_report', methods=['POST'])
def generate_report():
    """
    Gera um relatório consolidado a partir dos arquivos Excel processados.
    
    O relatório contém 3 abas:
    1. CHAMADAS - Todas as chamadas lado a lado para comparação
    2. ANÁLISE - Dados de análise de frequência 
    3. RELATÓRIO - Recomendações automáticas baseadas nos dados
    """
    if not session.get("autenticado"):
        return jsonify({'error': 'Não autenticado'}), 401
    
    try:
        app.logger.info("Iniciando geração de relatório consolidado")
        
        # Verifica se há arquivos processados na sessão
        if 'report_files' not in session or not session['report_files']:
            app.logger.warning("Nenhum arquivo encontrado na sessão para gerar relatório")
            return jsonify({'error': 'Nenhum arquivo foi processado. Por favor, faça o upload primeiro.'}), 400
        
        # Recupera os arquivos da sessão
        files_data = []
        for file_info in session.get('report_files', []):
            try:
                # Recupera o arquivo temporário
                temp_path = file_info.get('temp_path')
                
                if not temp_path or not os.path.exists(temp_path):
                    app.logger.warning(f"Arquivo temporário não encontrado: {temp_path}")
                    continue
                    
                with open(temp_path, 'rb') as f:
                    file_content = f.read()
                
                files_data.append({
                    'filename': file_info.get('filename', 'desconhecido'),
                    'content': file_content
                })
                app.logger.info(f"Arquivo recuperado com sucesso: {file_info.get('filename', 'desconhecido')}")
            except Exception as e:
                app.logger.error(f"Erro ao processar arquivo {file_info.get('filename', 'desconhecido')}: {str(e)}")
                continue
        
        if not files_data:
            app.logger.warning("Nenhum arquivo válido foi recuperado da sessão")
            return jsonify({'error': 'Nenhum arquivo válido foi encontrado para gerar o relatório.'}), 400
        
        # Importações necessárias
        import io
        
        try:
            # Importar funções do módulo de geração de relatórios
            # Processa os dados para identificar as escolas e datas
            app.logger.info(f"Processando {len(files_data)} arquivos para relatório consolidado")
            
            # Usa a função de geração de relatório consolidado
            try:
                # Tenta importar de diferentes locais (compatibilidade Replit/VS Code)
                try:
                    from backend.report_generator import generate_improved_report
                except ImportError:
                    try:
                        from report_generator import generate_improved_report
                    except ImportError:
                        # Se não conseguir importar, usa função interna simplificada
                        def generate_improved_report(files_data):
                            import openpyxl
                            from openpyxl.styles import Font, PatternFill
                            from openpyxl.utils import get_column_letter
                            import pandas as pd
                            import io
                            
                            # Cria novo workbook
                            wb = openpyxl.Workbook()
                            wb.remove(wb.active)  # Remove a planilha padrão
                            
                            # Aba CHAMADAS
                            ws_chamadas = wb.create_sheet("CHAMADAS")
                            ws_chamadas.append(["Relatório Consolidado de Chamadas"])
                            
                            # Processa arquivos de chamada
                            for file_info in files_data:
                                if 'analise' not in file_info['filename'].lower():
                                    try:
                                        with io.BytesIO(file_info['content']) as buffer:
                                            df = pd.read_excel(buffer)
                                            ws_chamadas.append([f"Arquivo: {file_info['filename']}"])
                                            for _, row in df.iterrows():
                                                ws_chamadas.append(row.tolist())
                                    except Exception as e:
                                        ws_chamadas.append([f"Erro ao processar {file_info['filename']}: {str(e)}"])
                            
                            # Aba ANÁLISE
                            analysis_files = [f for f in files_data if 'analise' in f['filename'].lower()]
                            if analysis_files:
                                ws_analise = wb.create_sheet("ANÁLISE")
                                for file_info in analysis_files:
                                    try:
                                        with io.BytesIO(file_info['content']) as buffer:
                                            df = pd.read_excel(buffer)
                                            ws_analise.append([f"Arquivo: {file_info['filename']}"])
                                            for _, row in df.iterrows():
                                                ws_analise.append(row.tolist())
                                    except Exception as e:
                                        ws_analise.append([f"Erro ao processar {file_info['filename']}: {str(e)}"])
                            
                            # Aba MONITORAR
                            ws_monitorar = wb.create_sheet("MONITORAR")
                            ws_monitorar.append(["Alunos para Monitoramento"])
                            ws_monitorar.append(["Nome", "Turma", "Percentual de Presença", "Status"])
                            
                            # Salva em buffer
                            buffer = io.BytesIO()
                            wb.save(buffer)
                            buffer.seek(0)
                            return buffer.getvalue()
                        
                        app.logger.info("Usando função interna de geração de relatório")
                
                # Gera o relatório com nossa implementação nova que resolve os problemas de:
                # 1. Status de presença (P/F/FJ) correto
                # 2. Datas em ordem cronológica
                # 3. Consolidação de observações por aluno
                # 4. Processamento adequado de arquivos de análise
                # 5. Nome do arquivo com período e escola extraídos dos cabeçalhos
                app.logger.info("Gerando relatório com datas ordenadas e status de presença corretos")
                report_result = generate_improved_report(files_data)
                
                # Verifica se o resultado contém as informações do cabeçalho
                if isinstance(report_result, dict) and 'excel_data' in report_result:
                    excel_output = report_result['excel_data']
                    school_name = report_result.get('school_name')
                    period = report_result.get('period')
                    app.logger.info(f"Informações extraídas - Escola: {school_name}, Período: {period}")
                else:
                    # Compatibilidade com versão antiga
                    excel_output = report_result
                    school_name = None
                    period = None
                    
                app.logger.info("Relatório consolidado gerado com sucesso")
            except Exception as e:
                # Se falhar, volta para o relatório padrão
                app.logger.warning(f"Erro na geração do relatório melhorado: {e}")
                app.logger.info("Gerando relatório consolidado com o método padrão")
                
                # Fallback: usa função básica de geração de relatório
                app.logger.error(f"Erro no gerador de relatório: {e}")
                return jsonify({'error': f'Erro ao gerar relatório: {str(e)}'}), 500
            
            # Nome do arquivo de saída - formato melhorado com informações dos cabeçalhos (sem data)
            if school_name and period:
                # Usa as informações extraídas dos cabeçalhos dos arquivos
                escola_formatada = re.sub(r'[^\w\s-]', '', school_name).strip().replace(' ', '_')
                periodo_formatado = period.replace(' ', '_')
                filename = f"Relatorio_Consolidado_de_Faltas_{periodo_formatado}_{escola_formatada}.xlsx"
                app.logger.info(f"Nome do arquivo gerado com informações dos cabeçalhos: {filename}")
            else:
                # Fallback para o formato anterior se não conseguir extrair informações
                escola = 'UnidadeEscolar'
                
                # Tenta extrair escola do primeiro arquivo de dados
                try:
                    if files_data:
                        first_file = files_data[0]['filename']
                        # Extrai nome da escola do nome do arquivo
                        match = re.search(r'^([^_]+)', first_file)
                        if match:
                            escola = match.group(1)
                except:
                    pass
                
                # Remove caracteres especiais do nome da escola
                escola = re.sub(r'[^\w\s-]', '', escola).strip().replace(' ', '_')
                
                filename = f"Relatorio_Consolidado_de_Faltas_Periodo_{escola}.xlsx"
                app.logger.warning(f"Usando formato padrão para nome do arquivo: {filename}")
            
            app.logger.info(f"Enviando relatório consolidado: {filename}")
            
            # Salva o arquivo temporariamente na sessão para download direto
            session['temp_report'] = {
                'data': excel_output,
                'filename': filename
            }
            
            # Retorna sucesso com URL de download
            return jsonify({
                'success': True,
                'filename': filename,
                'download_url': '/download_report'
            })
        except Exception as e:
            app.logger.error(f"Erro no gerador de relatório: {str(e)}")
            import traceback
            app.logger.error(traceback.format_exc())
            return jsonify({'error': f'Erro ao gerar o relatório: {str(e)}'}), 500
    except Exception as e:
        app.logger.error(f"Erro inesperado na geração de relatório: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': f'Erro ao gerar o relatório: {str(e)}'}), 500
        
        # Processar arquivos de análise
        row_analise = 5
        alunos_analise = {}
        
        try:
            # Tentar extrair dados dos arquivos Excel de análise
            for file_info in analise_files:
                try:
                    data_file = pd.read_excel(io.BytesIO(file_info.get('content')), engine='openpyxl')
                    
                    # Tentar identificar as colunas necessárias
                    turma_col = None
                    aluno_col = None
                    total_col = None
                    presenca_col = None
                    falta_col = None
                    freq_col = None
                    class_col = None
                    
                    for col in data_file.columns:
                        col_lower = str(col).lower()
                        if 'turma' in col_lower:
                            turma_col = col
                        elif 'aluno' in col_lower or 'nome' in col_lower:
                            aluno_col = col
                        elif 'total' in col_lower and 'aula' in col_lower:
                            total_col = col
                        elif 'presen' in col_lower:
                            presenca_col = col
                        elif 'falta' in col_lower:
                            falta_col = col
                        elif 'frequ' in col_lower or '%' in col_lower:
                            freq_col = col
                        elif 'class' in col_lower or 'situação' in col_lower:
                            class_col = col
                    
                    # Se encontramos as colunas mínimas necessárias
                    if turma_col and aluno_col:
                        for _, row in data_file.iterrows():
                            turma = str(row.get(turma_col, '')).strip()
                            aluno = str(row.get(aluno_col, '')).strip()
                            
                            if turma and aluno and turma != 'nan' and aluno != 'nan':
                                # Extrair os outros dados se disponíveis
                                total = str(row.get(total_col, '')) if total_col else ''
                                presenca = str(row.get(presenca_col, '')) if presenca_col else ''
                                falta = str(row.get(falta_col, '')) if falta_col else ''
                                freq = str(row.get(freq_col, '')) if freq_col else ''
                                classificacao = str(row.get(class_col, '')) if class_col else ''
                                
                                # Limpar valores nan
                                total = '' if total == 'nan' else total
                                presenca = '' if presenca == 'nan' else presenca
                                falta = '' if falta == 'nan' else falta
                                freq = '' if freq == 'nan' else freq
                                classificacao = '' if classificacao == 'nan' else classificacao
                                
                                # Armazenar para o aluno
                                key = f"{turma}|{aluno}"
                                alunos_analise[key] = {
                                    'turma': turma,
                                    'aluno': aluno,
                                    'total': total,
                                    'presenca': presenca,
                                    'falta': falta,
                                    'freq': freq,
                                    'classificacao': classificacao
                                }
                except Exception as e:
                    app.logger.warning(f"Erro ao processar arquivo de análise {file_info.get('filename')}: {str(e)}")
                    continue
        except Exception as e:
            app.logger.error(f"Erro geral ao processar arquivos de análise: {str(e)}")
        
        # Preencher a tabela de análise
        for key, dados in alunos_analise.items():
            ws_analise.cell(row=row_analise, column=1, value=dados['turma']).border = thin_border
            ws_analise.cell(row=row_analise, column=2, value=dados['aluno']).border = thin_border
            ws_analise.cell(row=row_analise, column=3, value=dados['total']).border = thin_border
            ws_analise.cell(row=row_analise, column=4, value=dados['presenca']).border = thin_border
            ws_analise.cell(row=row_analise, column=5, value=dados['falta']).border = thin_border
            ws_analise.cell(row=row_analise, column=6, value=dados['freq']).border = thin_border
            
            class_cell = ws_analise.cell(row=row_analise, column=7, value=dados['classificacao'])
            class_cell.border = thin_border
            
            # Colorir classificação
            if 'ausente' in dados['classificacao'].lower():
                class_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            elif 'monitoramento' in dados['classificacao'].lower():
                class_cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            elif 'regular' in dados['classificacao'].lower():
                class_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            
            row_analise += 1
        
        # CRIAR ABA DE RELATÓRIO (resumo)
        ws_relatorio = wb.create_sheet("RELATÓRIO")
        
        # Cabeçalho
        ws_relatorio.cell(row=1, column=1, value=f"Relatório de Recomendações - {escola_name}")
        ws_relatorio.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws_relatorio.merge_cells('A1:F1')
        
        # Data de criação
        ws_relatorio.cell(row=2, column=1, value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        ws_relatorio.cell(row=2, column=1).font = Font(italic=True)
        ws_relatorio.merge_cells('A2:F2')
        
        # Título dos arquivos processados
        ws_relatorio.cell(row=4, column=1, value="Arquivos Processados:")
        ws_relatorio.cell(row=4, column=1).font = Font(bold=True)
        
        # Listar arquivos processados
        row_rel = 5
        for file_info in files_data:
            file_name = file_info.get('filename', 'Desconhecido')
            ws_relatorio.cell(row=row_rel, column=1, value=file_name)
            row_rel += 1
        
        # Espaço
        row_rel += 1
        
        # Resumo das chamadas
        ws_relatorio.cell(row=row_rel, column=1, value="Resumo de Faltas por Data:")
        ws_relatorio.cell(row=row_rel, column=1).font = Font(bold=True)
        row_rel += 1
        
        # Cabeçalho do resumo de chamadas
        headers_resumo = ["Turma", "Data", "Total Alunos", "Presentes", "Ausentes", "% Presença"]
        for idx, header in enumerate(headers_resumo, 1):
            cell = ws_relatorio.cell(row=row_rel, column=idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = thin_border
        row_rel += 1
        
        # Gerar estatísticas por turma e data
        for turma in all_turmas:
            for data in datas:
                if data in chamada_data and turma in chamada_data[data]:
                    alunos_turma_data = chamada_data[data][turma]
                    total_alunos = len(alunos_turma_data)
                    
                    # Contar presentes e ausentes
                    presentes = 0
                    ausentes = 0
                    
                    for aluno, dados in alunos_turma_data.items():
                        status = dados.get('status', '').lower()
                        if 'presente' in status:
                            presentes += 1
                        elif 'ausente' in status or 'falta' in status:
                            ausentes += 1
                    
                    # Calcular percentual
                    perc_presenca = 0
                    if total_alunos > 0:
                        perc_presenca = round((presentes / total_alunos) * 100, 1)
                    
                    # Adicionar linha à tabela
                    ws_relatorio.cell(row=row_rel, column=1, value=turma).border = thin_border
                    ws_relatorio.cell(row=row_rel, column=2, value=data).border = thin_border
                    ws_relatorio.cell(row=row_rel, column=3, value=total_alunos).border = thin_border
                    ws_relatorio.cell(row=row_rel, column=4, value=presentes).border = thin_border
                    ws_relatorio.cell(row=row_rel, column=5, value=ausentes).border = thin_border
                    
                    freq_cell = ws_relatorio.cell(row=row_rel, column=6, value=f"{perc_presenca}%")
                    freq_cell.border = thin_border
                    
                    # Colorir célula conforme % de presença
                    if perc_presenca < 60:
                        freq_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                    elif perc_presenca < 75:
                        freq_cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                    else:
                        freq_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                    
                    row_rel += 1
        
        # Espaço
        row_rel += 2
        
        # Alunos com recomendações de intervenção
        ws_relatorio.cell(row=row_rel, column=1, value="Alunos Prioritários para Acompanhamento:")
        ws_relatorio.cell(row=row_rel, column=1).font = Font(bold=True)
        row_rel += 1
        
        # Cabeçalho da tabela de alunos prioritários
        priority_headers = ["Turma", "Aluno", "Tipo", "Justificativa"]
        for idx, header in enumerate(priority_headers, 1):
            cell = ws_relatorio.cell(row=row_rel, column=idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = thin_border
        row_rel += 1
        
        # Identificar alunos prioritários: aqueles com ausência em mais de uma data
        # ou classificados como "ausentes" na análise
        alunos_prioritarios = []
        alunos_ausencia_multipla = {}
        
        # Para cada aluno, contar suas ausências em diferentes datas
        for turma in all_turmas:
            for data in datas:
                if data in chamada_data and turma in chamada_data[data]:
                    for aluno, dados in chamada_data[data][turma].items():
                        status = dados.get('status', '').lower()
                        
                        if 'ausente' in status or 'falta' in status:
                            key = f"{turma}|{aluno}"
                            if key not in alunos_ausencia_multipla:
                                alunos_ausencia_multipla[key] = {
                                    'turma': turma,
                                    'aluno': aluno,
                                    'datas': []
                                }
                            
                            alunos_ausencia_multipla[key]['datas'].append(data)
        
        # Adicionar alunos com ausências múltiplas à lista de prioritários
        for key, dados in alunos_ausencia_multipla.items():
            if len(dados['datas']) > 1:  # Mais de uma ausência
                alunos_prioritarios.append({
                    'turma': dados['turma'],
                    'aluno': dados['aluno'],
                    'tipo': 'Alta Prioridade' if len(dados['datas']) >= len(datas) * 0.75 else 'Média Prioridade',
                    'justificativa': f"Faltou em {len(dados['datas'])} de {len(datas)} dias verificados."
                })
        
        # Adicionar alunos classificados como "ausentes" na análise
        for key, dados in alunos_analise.items():
            if 'ausente' in dados['classificacao'].lower():
                # Verificar se já está na lista
                ja_na_lista = False
                for ap in alunos_prioritarios:
                    if ap['turma'] == dados['turma'] and ap['aluno'] == dados['aluno']:
                        ja_na_lista = True
                        break
                
                if not ja_na_lista:
                    alunos_prioritarios.append({
                        'turma': dados['turma'],
                        'aluno': dados['aluno'],
                        'tipo': 'Alta Prioridade',
                        'justificativa': f"Classificado como {dados['classificacao']} na análise de frequência."
                    })
        
        # Ordenar por tipo (alta prioridade primeiro) e turma
        alunos_prioritarios.sort(key=lambda x: (0 if 'alta' in x['tipo'].lower() else 1, x['turma'], x['aluno']))
        
        # Adicionar alunos prioritários à tabela
        for ap in alunos_prioritarios:
            ws_relatorio.cell(row=row_rel, column=1, value=ap['turma']).border = thin_border
            ws_relatorio.cell(row=row_rel, column=2, value=ap['aluno']).border = thin_border
            
            tipo_cell = ws_relatorio.cell(row=row_rel, column=3, value=ap['tipo'])
            tipo_cell.border = thin_border
            
            # Colorir por tipo de prioridade
            if 'alta' in ap['tipo'].lower():
                tipo_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            else:
                tipo_cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            
            ws_relatorio.cell(row=row_rel, column=4, value=ap['justificativa']).border = thin_border
            row_rel += 1
        
        # Definir larguras das colunas na aba relatório
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws_relatorio.column_dimensions[col].width = 20
        
        # Salva o workbook em memória
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        excel_content = output.getvalue()
        
        # Nome do arquivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Relatorio_Consolidado_{timestamp}.xlsx"
        
        # Retorna o arquivo para download
        return send_file(
            io.BytesIO(excel_content),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Erro ao gerar relatório consolidado: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/download_report')
def download_report():
    """
    Rota para download direto do relatório consolidado gerado.
    """
    if 'temp_report' not in session:
        return "Nenhum relatório disponível para download", 404
    
    report_data = session['temp_report']
    
    # Remove da sessão após o download
    del session['temp_report']
    
    return send_file(
        io.BytesIO(report_data['data']),
        as_attachment=True,
        download_name=report_data['filename'],
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def save_temp_file(content):
    """
    Salva o conteúdo em um arquivo temporário e retorna o caminho.
    """
    import tempfile
    
    # Cria um arquivo temporário
    fd, path = tempfile.mkstemp(suffix='.xlsx')
    
    # Escreve o conteúdo no arquivo
    with os.fdopen(fd, 'wb') as f:
        f.write(content)
    
    return path

@app.route('/api/export_excel', methods=['GET'])
def export_attendance():
    try:
        # Obtém parâmetros da requisição ou valores padrão
        escola_selecionada = request.args.get('escola')
        periodo = request.args.get('periodo') or app_data.get('periodo', 'indefinido')
        current_user = app_data.get('current_user', 'indefinido')
        auto_clear = request.args.get('auto_clear', 'false').lower() == 'true'

        # Verifica se há turmas salvas em app_data
        if 'saved_classes' not in app_data or not app_data['saved_classes']:
            return jsonify({'success': False, 'error': 'Nenhuma turma salva para exportação'})

        # Define as turmas salvas a serem usadas
        if escola_selecionada:
            # Filtra turmas salvas da escola selecionada
            if escola_selecionada not in app_data['saved_classes'] or not app_data['saved_classes'][escola_selecionada]:
                return jsonify({'success': False, 'error': 'Nenhuma turma salva para a escola selecionada'})
            turmas_salvas = app_data['saved_classes'][escola_selecionada]
        else:
            # Usa todas as turmas salvas de todas as escolas
            turmas_salvas = set().union(*app_data['saved_classes'].values())
            if not turmas_salvas:
                return jsonify({'success': False, 'error': 'Nenhuma turma salva para exportação'})

        # Prepara os dados para exportação
        classes_to_export = {}
        attendance_to_export = {}
        observations_to_export = {}

        for turma in turmas_salvas:
            for escola, turmas in app_data['schools'].items():
                if turma in turmas:
                    # Filtra por escola, se especificada
                    if escola_selecionada and escola != escola_selecionada:
                        continue
                    classes_to_export[turma] = app_data['schools'][escola][turma]
                    attendance_to_export[turma] = app_data['attendance_status'].get(turma, {})
                    observations_to_export[turma] = app_data['observations'].get(turma, {})
                    break

        # Verifica se há turmas válidas para exportar
        if not classes_to_export:
            return jsonify({'success': False, 'error': 'Nenhuma turma válida encontrada para exportação'})

        # Gera o arquivo Excel
        output = export_to_excel(
            classes_to_export,
            attendance_to_export,
            observations_to_export,
            None,  # Não envia o HTML content
            current_user,
            periodo,
            escola_selecionada or "Todas as Escolas"
        )

        # Realiza limpeza, se solicitado
        if auto_clear:
            if escola_selecionada:
                app_data['saved_classes'][escola_selecionada].clear()
            else:
                app_data['saved_classes'].clear()
            for turma in classes_to_export:
                app_data['attendance_status'].pop(turma, None)
                app_data['observations'].pop(turma, None)

        # Retorna o arquivo Excel
        return send_file(
            output,
            as_attachment=True,
            download_name=get_excel_filename(escola_selecionada or "Todas as Escolas", periodo, current_user),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# Registrando os endpoints do Drive
app.route('/api/get_drive_folders', methods=['GET'])(get_drive_folders)
app.route('/api/export_excel_drive', methods=['POST'])(lambda: export_attendance_drive(app_data))

if __name__ == '__main__':
    # Adicionar o diretório atual ao path do sistema para encontrar os módulos corretamente
    sys.path.insert(0, os.path.abspath(os.path.dirname(os.path.dirname(__file__))))
    
    print(f"Template folder: {app.template_folder}")
    print(f"Static folder: {app.static_folder}")
    print(f"Templates existentes: {os.listdir(app.template_folder)}")
    
    # Iniciar o servidor na porta 5000, acessível por qualquer IP
    app.run(host='0.0.0.0', port=5000, debug=True)