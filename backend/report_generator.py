"""
Implementação aprimorada do gerador de relatórios consolidados

Este módulo implementa uma versão melhorada do gerador de relatórios consolidados,
resolvendo problemas específicos:

1. Ordenação correta das datas das visitas (ordem cronológica)
2. Exibição correta dos status de presença (P/F/FJ) 
3. Consolidação de observações por aluno em uma coluna "OBSERVAÇÕES"
4. Processamento adequado de arquivos de análise em abas separadas

Autor: Equipe PAESTRO
Data: Maio 2025
"""

import re
import io
import logging
import datetime
import pandas as pd
from copy import copy
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from backend.analysis_helpers import detect_analysis_file

# Configuração de logging
logger = logging.getLogger(__name__)

def extract_date_from_filename(filename):
    """Extrai data do nome do arquivo no formato DD/MM/YYYY"""
    match = re.search(r'(\d{2})[-_](\d{2})[-_](\d{4})', filename)
    if match:
        day, month, year = match.groups()
        return f"{day}/{month}/{year}"
    return None

def extract_date_from_content(df):
    """Busca data nas primeiras linhas do arquivo Excel"""
    for _, row in df.iloc[:5].iterrows():
        row_str = ' '.join([str(val) for val in row.values if pd.notna(val)])
        match = re.search(r'(\d{2}/\d{2}/\d{4})', row_str)
        if match:
            return match.group(1)
    return None

def extract_school_info_from_header(df):
    """Extrai nome da unidade e período do cabeçalho do arquivo Excel"""
    school_name = None
    period = None
    
    # Busca nas primeiras 10 linhas por informações da escola e período
    for _, row in df.iloc[:10].iterrows():
        row_str = ' '.join([str(val) for val in row.values if pd.notna(val)])
        row_str = row_str.upper()
        
        # Busca por nome da unidade (patterns comuns)
        if not school_name:
            # Patterns para identificar nome da escola
            school_patterns = [
                r'UNIDADE[:\s]+([A-ZÁÊÇÕ\s]+?)(?:\s+PERÍODO|\s+TURNO|\s+MATUTINO|\s+VESPERTINO|\s+INTEGRAL|$)',
                r'ESCOLA[:\s]+([A-ZÁÊÇÕ\s]+?)(?:\s+PERÍODO|\s+TURNO|\s+MATUTINO|\s+VESPERTINO|\s+INTEGRAL|$)',
                r'CEI[:\s]+([A-ZÁÊÇÕ\s]+?)(?:\s+PERÍODO|\s+TURNO|\s+MATUTINO|\s+VESPERTINO|\s+INTEGRAL|$)',
                r'CMEI[:\s]+([A-ZÁÊÇÕ\s]+?)(?:\s+PERÍODO|\s+TURNO|\s+MATUTINO|\s+VESPERTINO|\s+INTEGRAL|$)',
                r'([A-ZÁÊÇÕ\s]+CEI[A-ZÁÊÇÕ\s]*?)(?:\s+PERÍODO|\s+TURNO|\s+MATUTINO|\s+VESPERTINO|\s+INTEGRAL|$)',
                # Novos patterns mais específicos para CEI MUNDO ENCANTADO
                r'CEI\s+MUNDO\s+ENCANTADO',
                r'MUNDO\s+ENCANTADO',
                # Pattern genérico para capturar texto entre CEI e MATUTINO/VESPERTINO
                r'CEI\s+([A-ZÁÊÇÕ\s]+?)\s+(?:MATUTINO|VESPERTINO|INTEGRAL)',
                # Pattern para capturar nome da escola mesmo quando não tem CEI explícito
                r'([A-ZÁÊÇÕ\s]{10,}?)(?:\s+MATUTINO|\s+VESPERTINO|\s+INTEGRAL)'
            ]
            
            for pattern in school_patterns:
                match = re.search(pattern, row_str)
                if match:
                    school_name = match.group(1).strip() if len(match.groups()) > 0 else match.group(0).strip()
                    break
        
        # Busca por período
        if not period:
            if 'MATUTINO' in row_str:
                period = 'MATUTINO'
            elif 'VESPERTINO' in row_str:
                period = 'VESPERTINO'
            elif 'INTEGRAL' in row_str:
                period = 'INTEGRAL'
            elif 'NOTURNO' in row_str:
                period = 'NOTURNO'
    
    return school_name, period

def extract_annotations_from_file(df, filename):
    """
    Extrai anotações dos arquivos de chamada.
    Busca por células que contenham "ANOTAÇÕES:" e extrai as anotações das linhas seguintes.
    """
    anotacoes = []
    
    try:
        logger.info(f"Buscando anotações no arquivo {filename}...")
        
        # Extrai data do nome do arquivo (DD-MM-YYYY)
        data_arquivo = None
        date_match = re.search(r'(\d{1,2})-(\d{1,2})-(\d{4})', filename)
        if date_match:
            day, month, year = date_match.groups()
            data_arquivo = f"{day.zfill(2)}-{month.zfill(2)}"
            logger.info(f"Data extraída do nome do arquivo: {data_arquivo}")
        
        # Busca em todas as células do arquivo por "ANOTAÇÕES:"
        for i in range(min(20, len(df))):  # Verifica apenas as primeiras 20 linhas (cabeçalho)
            for col_idx, col in enumerate(df.columns):
                try:
                    cell_value = str(df.iloc[i, col_idx])
                    if pd.notna(cell_value) and "ANOTAÇÕES" in cell_value.upper():
                        logger.info(f"Encontrada célula com ANOTAÇÕES na linha {i}, coluna {col_idx}: {cell_value}")
                        
                        # Busca nas linhas seguintes na mesma coluna até encontrar uma linha vazia
                        for next_row in range(i + 1, min(i + 10, len(df))):
                            try:
                                next_cell_value = str(df.iloc[next_row, col_idx])
                                
                                # Se célula está vazia ou é "nan", para de buscar anotações
                                if (pd.isna(next_cell_value) or 
                                    next_cell_value.strip() == "" or 
                                    next_cell_value == "nan"):
                                    break
                                
                                # Se tem conteúdo, é uma anotação
                                anotacao_texto = next_cell_value.strip()
                                
                                # Usa a data do arquivo se disponível
                                if data_arquivo:
                                    anotacao_formatada = f"({data_arquivo}): {anotacao_texto}"
                                else:
                                    anotacao_formatada = anotacao_texto
                                
                                # Evita duplicatas
                                if anotacao_formatada not in anotacoes:
                                    anotacoes.append(anotacao_formatada)
                                    logger.info(f"Anotação encontrada: {anotacao_formatada}")
                                            
                            except (IndexError, ValueError):
                                break
                        
                        return anotacoes  # Para após encontrar e processar as anotações
                        
                except (IndexError, ValueError):
                    continue
                    
        if anotacoes:
            logger.info(f"Total de {len(anotacoes)} anotações válidas encontradas no arquivo {filename}")
        else:
            logger.info(f"Nenhuma anotação encontrada no arquivo {filename}")
            
    except Exception as e:
        logger.error(f"Erro ao extrair anotações do arquivo {filename}: {e}")
    
    return anotacoes

def detect_observacoes(row, aluno_col_idx):
    """Detecta observações em uma linha após uma coluna específica"""
    observacao = ""
    for idx, val in enumerate(row):
        if pd.isna(val) or idx <= aluno_col_idx:
            continue
            
        val_str = str(val).strip()
        # Se parece uma observação (texto com mais de 5 caracteres)
        if (len(val_str) > 5 and 
            val_str.upper() not in ['P', 'F', 'FJ'] and
            not re.match(r'^\d+[.,]?\d*$', val_str)):  # Não é um número
            observacao = val_str
            break
    
    return observacao

def extract_status(row):
    """Extrai o status P, F ou FJ de uma linha"""
    if not isinstance(row, pd.Series) and not isinstance(row, list) and not isinstance(row, np.ndarray):
        logger.debug(f"Tipo inesperado para extract_status: {type(row)}")
        return 'P'  # Default se não conseguir extrair

    # Converte para lista se for pandas Series
    if isinstance(row, pd.Series):
        row = row.values
    
    # Procura por valores diretos (P, F, FJ)
    for val in row:
        if pd.isna(val):
            continue
            
        val_str = str(val).strip().upper()
        if val_str in ['P', 'F', 'FJ']:
            return val_str
        
        # Busca por marcação de presença ou falta em células numéricas
        if val_str.isdigit():
            num = int(val_str)
            if 0 <= num <= 1:  # Muitas vezes 1=presente, 0=ausente
                return 'P' if num == 1 else 'F'
            
        # Strings que indicam presença
        if re.search(r'PRESENTE|PRESENÇA|COMP|COMPAREC', val_str):
            return 'P'
            
        # Strings que indicam falta
        if re.search(r'FALTA|AUSENTE|AUSÊNCIA|NÃO COMP', val_str):
            return 'F'
            
        # Strings que indicam falta justificada
        if re.search(r'JUSTIFICAD|ATESTADO|ABONAD', val_str):
            return 'FJ'
    
    # Default se não encontrar
    return 'P'

def normalize_turma_name(turma):
    """
    Normaliza o nome da turma para um formato padrão para evitar duplicações causadas
    por pequenas variações na escrita (ex: '1º ANO - 1' vs '1° ANO - 1').
    
    Verifica rigorosamente se o texto é uma turma válida e não outros elementos
    como datas, nomes de pessoas ou cabeçalhos.
    """
    if not turma or not isinstance(turma, str):
        return "Turma Não Identificada"
    
    # Remove espaços extras 
    turma = turma.strip()
    
    # Lista de padrões que NÃO são turmas válidas
    invalid_turma_patterns = [
        # Datas e horários
        r'\d{1,2}/\d{1,2}/\d{2,4}',  # Formatos de data DD/MM/YYYY
        r'\d{1,2}:\d{2}',  # Formato de hora HH:MM
        # Nomes de pessoas e cabeçalhos
        r'MUNICH',
        r'GUILHERME',
        r'ELIS',
        r'VICTORIA',
        r'DATA E HORA',
        r'FOI EMBORA',
        r'OBSERVAÇÃO',
        r'VISITA',
        r'PROFESSOR',
        r'MONITOR',
        r'SECRETÁRIA'
    ]
    
    # Verifica se o texto da turma corresponde a algum padrão inválido
    for pattern in invalid_turma_patterns:
        if re.search(pattern, turma.upper()):
            logger.info(f"Ignorando '{turma}' como turma - corresponde ao padrão inválido '{pattern}'")
            return "Turma Não Identificada"
    
    # Converte para minúsculo após as verificações de padrões inválidos
    turma = turma.lower()
    
    # Substitui variações de caracteres
    turma = turma.replace('º', 'o').replace('°', 'o').replace('ª', 'a')
    
    # Normaliza "ano" ou "série"
    turma = re.sub(r'\b(serie|série)\b', 'ano', turma)
    
    # Normaliza hífens e traços
    turma = re.sub(r'\s*[-_:]\s*', ' - ', turma)
    
    # Remove espaços duplicados
    turma = re.sub(r'\s+', ' ', turma)
    
    # Normaliza grupos de trabalho (GT)
    turma = re.sub(r'g\.?t\.?\s*(\d+)', r'gt\1', turma)
    
    # PADRÕES VÁLIDOS DE TURMA
    
    # Verifica se é um formato conhecido de turma de ensino fundamental
    ano_match = re.search(r'(\d+)o?\s*ano\s*[-\s]\s*(\d+|[a-z])', turma)
    
    # Verifica se é um formato conhecido de turma de educação infantil
    gt_match = re.search(r'gt\s*(\d+)\s*[-\s]?\s*([a-z])?', turma)
    
    # Remove prefixo "TURMA:" se existir
    if turma.startswith('turma:') or turma.startswith('turma '):
        turma = re.sub(r'^turma:?\s*', '', turma)
        # Busca novamente os padrões após remover o prefixo
        if not ano_match:
            ano_match = re.search(r'(\d+)o?\s*ano\s*[-\s]\s*(\d+|[a-z])', turma)
        if not gt_match:
            gt_match = re.search(r'gt\s*(\d+)\s*[-\s]?\s*([a-z])?', turma)
    
    # Formata o resultado de maneira padronizada
    if ano_match:
        numero, sufixo = ano_match.groups()
        return f"{numero}º ANO - {sufixo.upper()}"
    elif gt_match:
        numero = gt_match.group(1)
        sufixo = gt_match.group(2).upper() if gt_match.group(2) else ""
        if sufixo:
            return f"GT{numero} - {sufixo}"
        else:
            return f"GT{numero}"
    
    # Verifica se parece ser uma turma válida antes de aceitar
    # (deve conter "ANO", "SÉRIE", "TURMA" ou "GT")
    if re.search(r'\b(ano|série|serie|turma|gt)\b', turma, re.IGNORECASE):
        parts = turma.split()
        return ' '.join(p.capitalize() for p in parts)
    
    # Se não reconheceu nenhum formato válido, considera como não identificada
    logger.info(f"Formato de turma não reconhecido: '{turma}'")
    return "Turma Não Identificada"

def is_valid_student_name(name):
    """
    Verifica rigorosamente se o texto é um nome de aluno real e não outras informações como
    cabeçalhos, turmas, datas ou nomes de professores.
    """
    if not name or not isinstance(name, str):
        return False
        
    # Limpa espaços extras
    name = name.strip()
    if not name:
        return False
    
    # Lista extensa de strings e padrões que NÃO são nomes de alunos
    invalid_patterns = [
        # Cabeçalhos, títulos e labels
        r'^TURMA:',
        r'^TURMA ',
        r'^TOTAL',
        r'^DADOS',
        r'^PROFESSOR',
        r'^MONITOR',
        r'^DATA',
        r'^ORIENTAÇÕES',
        r'^SECRETARIA',
        r'^LISTA',
        r'^RELATÓRIO',
        r'^TABELA',
        r'^FREQUÊNCIA',
        r'LETIVO',
        r'ESCOLA',
        r'PERÍODO',
        r'CHAMADA',
        r'MATUTINO',
        r'VESPERTINO',
        r'NOTURNO',
        r'HORA',
        r'SALA',
        # Datas e horas - padrões muito mais rigorosos
        r'\d{1,2}/\d{1,2}/\d{2,4}',  # Formato data DD/MM/YYYY ou variações
        r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}',  # Datas com hífen ou barra
        r'\d{1,2}:\d{2}',  # Formato hora HH:MM
        r'^\d{1,2}[hH]\d{0,2}$',  # Formato hora brasileiro (ex: 10h30)
        # Nomes de profissionais e ações (não alunos)
        r'MUNICH',
        r'GUILHERME',
        r'ELIS',
        r'VICTORIA',
        r'PROFESSOR[A]?',
        r'COORDENADOR[A]?',
        r'DIRETOR[A]?',
        r'MONITOR[A]?',
        r'SECRETÁRI[OA]',
        r'ASSISTENTE',
        # Strings que claramente não são nomes
        r'^FOI EMBORA',
        r'SAIU MAIS CEDO',
        r'CHEGOU ATRASADO',
        r'OBSERVA[ÇC][ÃA]O',
        r'JUSTIFICATIVA',
        r'NOTA',
        r'ANO',
        r'FUNDAMENTAL',
        r'INFANTIL',
        r'MANHA',
        r'TARDE',
        r'VISITOU',
        # Elementos que parecem ser identificadores de turma
        r'\d[ºo°]\s*ANO',  # Padrões como "1º ANO", "2o ANO", etc.
        r'GT\d',  # Padrões como GT1, GT2, etc. para turmas de educação infantil
    ]
    
    # Verifica padrões - se algum bater, não é nome de aluno
    for pattern in invalid_patterns:
        if re.search(pattern, name.upper()):
            logger.info(f"Rejeitado '{name}' por corresponder ao padrão '{pattern}'")
            return False
    
    # Rejeita linhas que são apenas números, códigos ou datas
    if re.match(r'^\d+$', name) or re.match(r'^[()\d\s]+$', name):
        logger.info(f"Rejeitado '{name}' por ser código/número")
        return False
    
    # CONDIÇÕES POSITIVAS: Agora vamos verificar se é um nome válido

    # Regra 1: Nomes todo em maiúsculas com pelo menos duas palavras (característico dos arquivos GE)
    # e sem caracteres estranhos ou números - Isso é muito comum nos arquivos GE_PEQUENO_PRINCIPE
    if name.isupper() and ' ' in name and not re.search(r'[*+=&%$#@!?><\[\]\{\}\\|0-9]', name):
        words = name.split()
        # O nome precisa ter pelo menos duas palavras para ser um nome completo (nome e sobrenome)
        if len(words) >= 2 and all(len(word) >= 2 for word in words):
            logger.info(f"Aceitado nome maiúsculo: '{name}'")
            return True
            
    # Regra 2: Nomes em formato tradicional (primeira letra maiúscula) com pelo menos 2 palavras
    words = name.split()
    if len(words) >= 2:
        # Verifica se as palavras seguem o padrão de nome próprio (primeira letra maiúscula)
        if all(word[0].isupper() for word in words if len(word) > 1):
            logger.info(f"Aceitado nome com formato tradicional: '{name}'")
            return True
            
    # Se não cumpriu nenhuma regra positiva, não é um nome válido
    logger.info(f"Rejeitado nome que não cumpre regras: '{name}'")
    return False

def is_analysis_file(filename, content=None):
    """
    Verifica se um arquivo é um arquivo de análise baseado no nome e conteúdo.
    
    Args:
        filename (str): Nome do arquivo
        content (bytes, optional): Conteúdo do arquivo
        
    Returns:
        bool: True se for um arquivo de análise, False caso contrário
    """
    # Arquivos GE são arquivos de CHAMADA normais, não são arquivos de análise
    if filename.startswith("GE_") or "PEQUENO_PRINCIPE" in filename.upper():
        logger.info(f"Arquivo GE identificado como arquivo de CHAMADA regular: {filename}")
        # Retornamos False pois esses arquivos devem ir apenas para a aba CHAMADAS
        return False
        
    # Regra 1: Reconhecer arquivos do nosso próprio sistema
    # Arquivos de chamada são gerados com prefixo da escola, data, período e nomes
    # Ex: PEQUENO_PRINCIPE_08-05-2025_MATUTINO_MUNICHGUILHERME_E_ELIS.xlsx
    is_call_file = False
    
    # Verifica se parece um arquivo de chamada pelo padrão de nomenclatura
    date_pattern = re.search(r'\d{2}-\d{2}-\d{4}', filename)
    if date_pattern and ('_MATUTINO_' in filename or '_VESPERTINO_' in filename):
        is_call_file = True
        logger.info(f"Arquivo identificado como CHAMADA pelo padrão de nomenclatura: {filename}")
    
    # Se é claramente um arquivo de chamada, não é análise
    if is_call_file:
        return False
    
    # Verifica se parece um arquivo de análise pelo nome
    name_indicators = [
        "analise", "análise", "analitico", "analítico", "indice", "índice",
        "estudo", "avaliacao", "avaliação", "consolidado", "resumo"
    ]
    
    if any(indicator in filename.lower() for indicator in name_indicators):
        logger.debug(f"Identificado como arquivo de análise pelo nome: {filename}")
        return True
        
    # Se tivermos o conteúdo, fazemos verificações mais detalhadas
    if content:
        try:
            with io.BytesIO(content) as buffer:
                # Verifica se o arquivo tem formato específico de análise
                df = pd.read_excel(buffer, nrows=20)  # Lê apenas primeiras linhas para eficiência
                
                # Verifica o título das planilhas (se tiver ANÁLISE é um indicativo forte)
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
                    sheet_names = wb.sheetnames
                    wb.close()
                    
                    for sheet_name in sheet_names:
                        if "ANÁLISE" in sheet_name.upper() or "ANALISE" in sheet_name.upper():
                            logger.debug(f"Identificado como arquivo de análise pelo nome da aba: {sheet_name}")
                            return True
                except Exception as sheet_error:
                    logger.warning(f"Erro ao verificar nomes das abas: {sheet_error}")
                
                # Verifica colunas frequentemente encontradas em relatórios de análise
                columns_text = ' '.join(str(col).lower() for col in df.columns)
                analysis_indicators = [
                    "percentual", "porcentagem", "média", "media", 
                    "total faltas", "estatística", "estatistica",
                    "consolidado", "resumo", "análise", "indice", "geral",
                    "frequência total", "total alunos", "ausentes", "presentes"
                ]
                
                if any(indicator in columns_text for indicator in analysis_indicators):
                    logger.debug(f"Identificado como arquivo de análise pelo conteúdo das colunas: {filename}")
                    return True
                    
                # Verifica conteúdo das primeiras linhas
                first_rows = ' '.join(df.iloc[:8].astype(str).values.flatten().tolist())
                analysis_content = ["relatório", "análise", "estatística", "resumo", 
                                   "consolidado", "comparativo", "percentual", "ausências"]
                
                if any(indicator in first_rows.lower() for indicator in analysis_content):
                    logger.debug(f"Identificado como arquivo de análise pelo conteúdo das primeiras linhas: {filename}")
                    return True
                    
                # Verifica a estrutura do arquivo - arquivos de análise geralmente têm menos colunas e mais linhas
                if len(df.columns) < 5 and "%" in first_rows:
                    logger.debug(f"Identificado como arquivo de análise pela estrutura: {filename}")
                    return True
        except Exception as e:
            logger.error(f"Erro ao verificar conteúdo do arquivo para classificação: {e}")
    
    # Se não encontrou evidências de que seja um arquivo de análise
    return False

def process_file(file_content, filename):
    """
    Processa um arquivo Excel para extração de dados.
    
    Retorna:
    - is_analysis: se é um arquivo de análise
    - date: data do arquivo (DD/MM/YYYY)
    - turmas: lista de turmas identificadas
    - alunos_data: dicionário com dados dos alunos (status e observações)
    - df: DataFrame completo para arquivos de análise
    """
    # Definimos explicitamente que arquivos do padrão "NOME_ESCOLA_DATA_PERÍODO_NOMES" são de chamada
    is_call_file = False
    date_pattern = re.search(r'\d{2}-\d{2}-\d{4}', filename)
    if date_pattern and ('_MATUTINO_' in filename or '_VESPERTINO_' in filename):
        is_call_file = True
        is_analysis = False
        logger.info(f"Arquivo de chamada identificado pelo padrão padrão: {filename}")
    else:
        # Para outros arquivos, verificamos se é análise
        is_analysis = is_analysis_file(filename, file_content)
    
    result = {
        'filename': filename,
        'is_analysis': is_analysis,
        'date': None,
        'turmas': [],
        'alunos_data': {},
        'df': None  # DataFrame completo para arquivos de análise
    }
    
    try:
        with io.BytesIO(file_content) as buffer:
            # Carrega o arquivo Excel
            df = pd.read_excel(buffer)
            
            # Se for um arquivo de análise, guarda o DataFrame completo e retorna
            if result['is_analysis']:
                result['df'] = df
                return result
                
            # Se chegou aqui, é arquivo de chamada - extrair detalhes
            
            # Extrai a data do arquivo (formato DD/MM/YYYY)
            if is_call_file and date_pattern:
                # Formato é DD-MM-YYYY no nome do arquivo
                date_match = date_pattern.group(0)
                parts = date_match.split('-')
                if len(parts) == 3:
                    result['date'] = f"{parts[0]}/{parts[1]}/{parts[2]}"
                    logger.info(f"Data extraída do nome do arquivo: {result['date']}")
            
            # Se não encontrou no nome, tenta extrair do conteúdo
            if not result['date']:
                result['date'] = extract_date_from_filename(filename)
            if not result['date']:
                result['date'] = extract_date_from_content(df)
            
            # Cria uma string com todo o conteúdo do arquivo para busca mais eficaz de turmas
            full_content = '\n'.join([
                ' '.join([str(val) for val in row.values if pd.notna(val)])
                for _, row in df.iterrows()
            ])
            
            # Encontra todas as turmas no conteúdo completo
            turmas_encontradas = []
            
            # Procura padrões diferentes de turmas
            padrao_turma1 = re.findall(r'TURMA\s*:?\s*([^:|\n]+)', full_content, re.IGNORECASE)
            padrao_turma2 = re.findall(r'((?:GT|G\.T\.)\s*\d+\s*[-\s]*[A-Za-z]?)', full_content, re.IGNORECASE)
            padrao_turma3 = re.findall(r'(\d+º?\s*(?:ANO|SÉRIE)\s*[-\s]*[A-Za-z]?)', full_content, re.IGNORECASE)
            
            # Consolida todas as turmas encontradas
            for match_list in [padrao_turma1, padrao_turma2, padrao_turma3]:
                for match in match_list:
                    turma = match.strip()
                    if turma and turma not in turmas_encontradas:
                        turmas_encontradas.append(turma)
            
            # Se não encontrou turmas pelos padrões, tenta outras estratégias
            if not turmas_encontradas:
                # Tenta procurar turmas nas células diretamente, verificando por padrões comuns
                for idx, row in df.iterrows():
                    for col_idx, val in enumerate(row):
                        if pd.isna(val):
                            continue
                        
                        val_str = str(val).strip()
                        # Verifica se parece um nome de turma
                        if re.match(r'(GT\s*\d+|G\.T\.\s*\d+|\d+º?\s*ANO|\d+º?\s*SÉRIE)', val_str, re.IGNORECASE):
                            if val_str not in turmas_encontradas:
                                turmas_encontradas.append(val_str)
            
            # Adiciona as turmas encontradas ao resultado
            result['turmas'].extend(turmas_encontradas)
            
            # Processa as linhas para identificar turmas e alunos
            turma_atual = None
            ultima_turma_vista = None
            alunos_encontrados = 0  # Contador para logging
            
            # Log para depuração
            logger.info(f"Procurando alunos no arquivo com {len(df)} linhas...")
            
            for idx, row in df.iterrows():
                # Converte a linha em string para facilitar a busca
                row_str = ' '.join([str(val) for val in row.values if pd.notna(val)])
                
                # Verifica se esta linha contém alguma das turmas identificadas
                turma_encontrada = None
                for turma in turmas_encontradas:
                    if turma in row_str:
                        turma_encontrada = turma
                        break
                
                if turma_encontrada:
                    turma_atual = turma_encontrada
                    ultima_turma_vista = turma_atual
                    logger.debug(f"Linha {idx}: Encontrada turma: {turma_atual}")
                
                # Se não temos turma atual mas já vimos uma turma antes, continua usando a última
                if not turma_atual and ultima_turma_vista:
                    turma_atual = ultima_turma_vista
                
                # Busca por alunos na linha
                for col_idx, val in enumerate(row):
                    if pd.isna(val):
                        continue
                    
                    # Converte qualquer valor para string
                    val_str = str(val).strip()
                    
                    # Primeira verificação rápida - nome de aluno geralmente tem comprimento razoável e contém espaço
                    if len(val_str) > 5 and ' ' in val_str:
                        # Aplica verificações mais detalhadas
                        is_valid = is_valid_student_name(val_str)
                        
                        # Se a função rigorosa aprovou, usamos o nome
                        if is_valid:
                            aluno_nome = val_str
                            logger.debug(f"Nome validado: {aluno_nome}")
                        # Se não passou na verificação rigorosa, ainda verifica alguns casos básicos
                        elif len(val_str.split()) >= 2 and re.match(r'^[A-Za-zÀ-ÿ]', val_str):
                            # Se começar com letra maiúscula, provável nome próprio
                            aluno_nome = val_str
                            logger.debug(f"Nome parcialmente validado: {aluno_nome}")
                        else:
                            # Não parece ser um nome de aluno
                            continue
                        
                        # Extrai status e observação de forma mais robusta
                        try:
                            status = "P"  # Status padrão é Presente
                            if col_idx+1 < len(row):
                                status_val = row[col_idx+1]
                                if pd.notna(status_val):
                                    status_str = str(status_val).strip().upper()
                                    if status_str in ["F", "FJ"]:
                                        status = status_str
                                    elif status_str == "0" or status_str == "FALTA":
                                        status = "F"
                                    elif status_str == "FJ" or status_str == "JUSTIFICADA":
                                        status = "FJ"
                            logger.info(f"Status extraído para {aluno_nome}: {status}")
                            
                            observacao = detect_observacoes(row, col_idx)
                        except Exception as e:
                            logger.error(f"Erro extraindo status/observação: {e}")
                            status = "P"  # Assume presente como padrão
                            observacao = ""
                        
                        # Adiciona o aluno aos dados com mais verificações
                        if turma_atual:
                            aluno_key = f"{turma_atual}|{aluno_nome}"
                            result['alunos_data'][aluno_key] = {
                                'turma': turma_atual,
                                'nome': aluno_nome,
                                'status': status,
                                'observacao': observacao
                            }
                            alunos_encontrados += 1
                            logger.info(f"Adicionado aluno: {aluno_nome} na turma: {turma_atual} com status: {status}")
                        else:
                            # Se não temos uma turma atual, atribuímos a uma turma genérica
                            turma_generica = "Turma Não Identificada"
                            if turma_generica not in result['turmas']:
                                result['turmas'].append(turma_generica)
                            
                            aluno_key = f"{turma_generica}|{aluno_nome}"
                            result['alunos_data'][aluno_key] = {
                                'turma': turma_generica,
                                'nome': aluno_nome,
                                'status': status,
                                'observacao': observacao
                            }
                            alunos_encontrados += 1
                            logger.info(f"Aluno sem turma: {aluno_nome} (atribuído à turma genérica) com status: {status}")
            
            # Log do resultado final para este arquivo
            logger.info(f"Total de alunos encontrados no arquivo: {alunos_encontrados}")
        
        # Deduplica as turmas
        result['turmas'] = list(set(result['turmas']))
        
        # Extrai anotações do arquivo de chamada (se não for arquivo de análise)
        if not is_analysis:
            anotacoes = extract_annotations_from_file(df, filename)
            result['anotacoes'] = anotacoes
        else:
            result['anotacoes'] = []
        
        # Log para debug
        logger.info(f"Arquivo {filename}: Encontrou {len(result['turmas'])} turmas, {len(result['alunos_data'])} alunos e {len(result.get('anotacoes', []))} anotações")
        
        return result
    except Exception as e:
        logger.error(f"Erro ao processar arquivo {filename}: {e}")
        return result

def generate_improved_report(files_data):
    """
    Gera relatório consolidado com melhorias implementadas:
    
    1. Datas em ordem cronológica
    2. Status de presença P/F/FJ exibidos corretamente
    3. Observações consolidadas por aluno
    4. Arquivos de análise em abas separadas
    5. Nome do arquivo com período e escola extraídos dos cabeçalhos
    """
    try:
        # Separa arquivos regulares de arquivos de análise
        regular_files = []
        analysis_files = []
        
        # Variáveis para extrair informações do cabeçalho
        school_name = None
        period = None
        
        # EXTRAÇÃO PRIORITÁRIA: Busca o nome da escola APENAS nos arquivos de CHAMADA (não análise)
        for file_info in files_data:
            filename = file_info.get('filename', '')
            # Ignora arquivos de análise que começam com "analise_frequencia"
            if filename and not filename.lower().startswith('analise_frequencia'):
                # Extrai nome da escola do padrão: NOME_DA_ESCOLA_DD-MM-YYYY_PERIODO_...
                parts = filename.split('_')
                if len(parts) >= 3:
                    # Pega as partes antes da data (que tem formato DD-MM-YYYY)
                    school_parts = []
                    for part in parts:
                        # Para quando encontrar a data (formato DD-MM-YYYY)
                        if re.match(r'\d{2}-\d{2}-\d{4}', part):
                            break
                        school_parts.append(part)
                    
                    if school_parts:
                        school_name = '_'.join(school_parts)
                        logger.info(f"✓ ESCOLA EXTRAÍDA DO ARQUIVO DE CHAMADA: {school_name}")
                        break
        
        # EXTRAÇÃO PRIORITÁRIA: Busca o período nos nomes dos arquivos PRIMEIRO
        if not period:
            for file_info in files_data:
                filename = file_info.get('filename', '')
                if filename:
                    # Extrai período do padrão: NOME_DA_ESCOLA_DD-MM-YYYY_PERIODO_...
                    parts = filename.split('_')
                    if len(parts) >= 4:
                        # Procura por parte que contém data, o período vem depois
                        for i, part in enumerate(parts):
                            if re.match(r'\d{2}-\d{2}-\d{4}', part) and i + 1 < len(parts):
                                period = parts[i + 1]
                                logger.info(f"✓ PERÍODO EXTRAÍDO DO NOME DO ARQUIVO: {period}")
                                break
                        if period:
                            break
        
        # Log de informações importantes
        logger.info(f"Processando {len(files_data)} arquivos para relatório consolidado")
        logger.info(f"ESCOLA IDENTIFICADA: {school_name}")
        logger.info(f"PERÍODO IDENTIFICADO: {period}")
        
        for file_info in files_data:
            filename = file_info.get('filename', '')
            original_filename = filename
            content = file_info.get('content', None)
            
            # Usa nossa nova função simplificada de detecção baseada no padrão do nome
            is_analysis_file = detect_analysis_file(original_filename, content if content else b'')
            
            if is_analysis_file:
                logger.info(f"[IMPORTANTE] Arquivo identificado como ANÁLISE: {original_filename}")
                file_info['is_analysis'] = True
                analysis_files.append(file_info)
            else:
                logger.info(f"[IMPORTANTE] Arquivo identificado como CHAMADA: {original_filename}")
                file_info['is_analysis'] = False
                regular_files.append(file_info)

        
        # Log resumo
        logger.info(f"[RESUMO] Total: {len(files_data)}, Arquivos de CHAMADA: {len(regular_files)}, Arquivos de ANÁLISE: {len(analysis_files)}")
        
        # Processa os dados de frequência
        all_data = []
        all_dates = []
        all_turmas = set()
        
        # Mapa para armazenar as turmas normalizadas
        turma_normalization_map = {}
        
        # Log para depuração
        logger.info(f"Processando {len(regular_files)} arquivos de chamada...")
            
        for file_info in regular_files:
            logger.info(f"Processando arquivo de chamada: {file_info.get('filename')}")
            
            # Para evitar processamento duplo, verificamos se o arquivo já tem uma marca explícita 
            # de que não é análise (isso foi adicionado na etapa de classificação)
            is_call_file = not file_info.get('is_analysis', False)
            
            # Se já sabemos que é um arquivo de chamada, garantimos que seja processado como tal
            if is_call_file:
                data = process_file(file_info['content'], file_info['filename'])
                
                # Extrai informações do cabeçalho do primeiro arquivo (se ainda não extraiu)
                if not school_name or not period:
                    try:
                        # Primeira tentativa: extrair do nome do arquivo
                        filename = file_info.get('filename', '')
                        logger.info(f"Tentando extrair escola do arquivo: {filename}")
                        if not school_name and filename:
                            # Extrai do nome do arquivo (CEI_MUNDO_ENCANTADO_21-05-2025_MATUTINO_...)
                            if 'CEI_MUNDO_ENCANTADO' in filename.upper():
                                school_name = 'CEI_MUNDO_ENCANTADO'
                                logger.info(f"✓ Nome da escola extraído do nome do arquivo: {school_name}")
                            elif 'MUNDO_ENCANTADO' in filename.upper():
                                school_name = 'MUNDO_ENCANTADO'
                                logger.info(f"✓ Nome da escola extraído do nome do arquivo: {school_name}")
                            else:
                                logger.warning(f"Não foi possível extrair nome da escola de: {filename}")
                        
                        if not period and filename:
                            if 'MATUTINO' in filename.upper():
                                period = 'MATUTINO'
                                logger.info(f"Período extraído do nome do arquivo: {period}")
                            elif 'VESPERTINO' in filename.upper():
                                period = 'VESPERTINO'
                                logger.info(f"Período extraído do nome do arquivo: {period}")
                        
                        # Segunda tentativa: extrair do conteúdo do arquivo
                        if not school_name or not period:
                            with io.BytesIO(file_info['content']) as buffer:
                                df = pd.read_excel(buffer, nrows=10)
                                extracted_school, extracted_period = extract_school_info_from_header(df)
                                
                                if extracted_school and not school_name:
                                    school_name = extracted_school
                                    logger.info(f"Nome da escola extraído do conteúdo: {school_name}")
                                
                                if extracted_period and not period:
                                    period = extracted_period
                                    logger.info(f"Período extraído do conteúdo: {period}")
                                
                    except Exception as e:
                        logger.warning(f"Erro ao extrair informações do cabeçalho: {e}")
                
                # Log detalhado para entender o processamento
                logger.info(f"Arquivo processado: {file_info.get('filename')}")
                logger.info(f"  - Data: {data['date']}")
                logger.info(f"  - Turmas: {', '.join(data['turmas']) if data['turmas'] else 'Nenhuma turma encontrada'}")
                logger.info(f"  - Alunos: {len(data['alunos_data'])}")
                
                if data['date']:
                    all_dates.append(data['date'])
                
                # Se não encontrou turmas, tenta um método alternativo (análise por títulos comuns)
                if not data['turmas']:
                    logger.warning(f"Nenhuma turma encontrada no arquivo {file_info.get('filename')}. Tentando método alternativo.")
                    try:
                        # Detecção alternativa por extração de títulos
                        with io.BytesIO(file_info['content']) as buffer:
                            df = pd.read_excel(buffer)
                            # Procura nas primeiras 20 linhas por títulos que pareçam turmas
                            for idx, row in df.head(20).iterrows():
                                row_str = ' '.join([str(val) for val in row.values if pd.notna(val)])
                                turmas_alt = re.findall(r'(?:GT|G\.T\.)\s*\d+\s*[-\s]*[A-Za-z]?|(?:\d+º?|PRIMEIRO|SEGUNDO|TERCEIRO|QUARTO|QUINTO)\s*(?:ANO|SERIE|SÉRIE)\s*[-\s]*[A-Za-z]?', row_str, re.IGNORECASE)
                                if turmas_alt:
                                    data['turmas'].extend(turmas_alt)
                                    logger.info(f"Turmas encontradas pelo método alternativo: {', '.join(turmas_alt)}")
                    except Exception as e:
                        logger.error(f"Erro na detecção alternativa de turmas: {e}")
                
                # Normaliza os nomes das turmas para evitar duplicações
                normalized_turmas = []
                for turma in data['turmas']:
                    if not turma:
                        continue
                        
                    # Normaliza o nome da turma
                    normalized = normalize_turma_name(turma)
                    
                    # Adiciona ao mapa de normalização
                    turma_normalization_map[turma] = normalized
                    
                    # Usa a versão normalizada
                    normalized_turmas.append(normalized)
                
                # Substitui as turmas originais pelas normalizadas
                data['turmas'] = normalized_turmas
                
                # Adiciona os dados processados à lista global mesmo se não tiver turmas
                # (podemos ter dados de alunos sem identificação de turma)
                all_data.append(data)
                logger.info(f"Dados do arquivo {file_info.get('filename')} adicionados ao conjunto.")
            else:
                logger.warning(f"Ignorando arquivo {file_info.get('filename')} que foi reclassificado como não sendo de chamada")
            
            # Verificamos se temos dados válidos para processar
            if is_call_file and 'alunos_data' in data:
                # Normaliza as turmas nos dados de alunos
                normalized_alunos_data = {}
                for key, aluno_data in data['alunos_data'].items():
                    turma = aluno_data.get('turma', '')
                    if turma and turma in turma_normalization_map:
                        # Cria uma nova chave com a turma normalizada
                        new_key = f"{turma_normalization_map[turma]}|{aluno_data.get('nome')}"
                        # Atualiza a turma no dicionário de dados
                        aluno_data['turma'] = turma_normalization_map[turma]
                        # Usa a nova chave
                        normalized_alunos_data[new_key] = aluno_data
                    else:
                        normalized_alunos_data[key] = aluno_data
                
                # Substitui o dicionário original pelo normalizado
                data['alunos_data'] = normalized_alunos_data
                
                # Adiciona as turmas normalizadas ao conjunto global se houver alguma
                if normalized_turmas:
                    all_turmas.update(normalized_turmas)
                    logger.info(f"Turmas adicionadas ao conjunto global: {', '.join(normalized_turmas)}")
                
                # Adiciona os dados processados à lista global
                all_data.append(data)
                logger.info(f"Arquivo {file_info.get('filename')} adicionado ao conjunto de dados com {len(data['alunos_data'])} alunos")
            else:
                # Caso não seja um arquivo de chamada ou não tenha dados de alunos
                logger.warning(f"Arquivo {file_info.get('filename')} ignorado por não ter dados de alunos para processar")
        
        # Log para depuração
        logger.info(f"Turmas normalizadas encontradas: {len(all_turmas)}: {', '.join(sorted(all_turmas))}")
        
        # Ordena datas cronologicamente
        def convert_to_date(date_str):
            if not date_str:
                return datetime.datetime(2099, 12, 31)  # Data máxima para ordenação
            try:
                parts = date_str.split('/')
                if len(parts) == 3:
                    return datetime.datetime(int(parts[2]), int(parts[1]), int(parts[0]))
            except:
                pass
            return datetime.datetime(2099, 12, 31)
        
        all_dates = sorted(set(all_dates), key=convert_to_date)
        
        # Usa a função de normalização definida no escopo global

        # Função que extrai um número para ordenar turmas
        def extract_turma_number(turma):
            # Usa a função de normalização global
            normalized = normalize_turma_name(turma).upper()
            
            # Turmas numéricas (ex: 1º ANO - 1)
            ano_match = re.search(r'(\d+)º\s*ANO\s*-\s*(\d+|[A-Z])', normalized)
            if ano_match:
                ano = int(ano_match.group(1))
                letra_ou_numero = ano_match.group(2)
                
                # Converte letra para número se necessário (A=1, B=2, etc)
                if re.match(r'[A-Z]', letra_ou_numero):
                    numero_secundario = ord(letra_ou_numero) - ord('A') + 1
                else:
                    numero_secundario = int(letra_ou_numero)
                    
                # Prioridade para o ano, depois para a letra/número
                return (ano * 100) + numero_secundario
                
            # Grupos de trabalho (formato GT5 - A)
            gt_match = re.search(r'GT(\d+)(?:\s*-\s*([A-Z]))?', normalized)
            if gt_match:
                gt_numero = int(gt_match.group(1))
                letra = gt_match.group(2) if gt_match.group(2) else "A"
                letra_numero = ord(letra) - ord('A') + 1
                
                # GT vem depois das turmas regulares, começando do 1000
                return 1000 + (gt_numero * 100) + letra_numero
                
            # Ordenação padrão para casos não reconhecidos
            return 9999
        
        sorted_turmas = sorted(all_turmas, key=extract_turma_number)
        
        # Cria o workbook
        wb = openpyxl.Workbook()
        
        # Aba de chamadas
        ws_chamadas = wb.active
        ws_chamadas.title = "CHAMADAS"
        
        # Cabeçalho com título
        escola = "ESCOLA"
        for data in all_data:
            if "PEQUENO" in data['filename'].upper():
                escola = "PEQUENO PRINCIPE"
                break
                
        # Título com formatação melhorada - usa nome correto da escola
        escola_formatada = school_name.replace('_', ' ') if school_name else 'ESCOLA'
        ws_chamadas.cell(row=1, column=1, value=f"RELATÓRIO CONSOLIDADO - {escola_formatada}")
        ws_chamadas.cell(row=1, column=1).font = Font(bold=True, size=14, color="FFFFFF")
        ws_chamadas.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws_chamadas.cell(row=1, column=1).fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws_chamadas.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_dates) + 2)
        
        # Aplica o estilo em todas as células mescladas do título
        for col in range(1, len(all_dates) + 3):
            if col > 1:  # A primeira coluna já foi formatada acima
                ws_chamadas.cell(row=1, column=col).fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                ws_chamadas.cell(row=1, column=col).font = Font(bold=True, size=14, color="FFFFFF")
        
        # Data de geração com formatação melhorada
        data_geracao = datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
        ws_chamadas.cell(row=2, column=1, value=f"Gerado em: {data_geracao}")
        ws_chamadas.cell(row=2, column=1).font = Font(italic=True)
        ws_chamadas.cell(row=2, column=1).alignment = Alignment(horizontal='center')
        ws_chamadas.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(all_dates) + 2)
        
        # Coleta anotações dos arquivos de chamada para exibir no cabeçalho
        anotacoes_consolidadas = []
        anotacoes_unicas = set()  # Para evitar duplicação
        for data in all_data:
            # Verifica se o arquivo tem anotações (apenas arquivos de chamada)
            anotacoes_arquivo = data.get('anotacoes', [])
            if anotacoes_arquivo:
                for anotacao in anotacoes_arquivo:
                    if anotacao not in anotacoes_unicas:
                        anotacoes_unicas.add(anotacao)
                        anotacoes_consolidadas.append(anotacao)
        
        # Adiciona anotações no cabeçalho se houver
        current_row = 3
        if anotacoes_consolidadas:
            # Formata as anotações no formato solicitado: "ANOTAÇÕES: (DD-MM): exemplo. (DD-MM): exemplo 2."
            anotacoes_texto = "ANOTAÇÕES: " + " ".join(anotacoes_consolidadas)
            
            ws_chamadas.cell(row=current_row, column=1, value=anotacoes_texto)
            ws_chamadas.cell(row=current_row, column=1).font = Font(bold=True, size=10)
            ws_chamadas.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(all_dates) + 2)
            current_row += 1
        
        # Linha de espaçamento
        ws_chamadas.row_dimensions[current_row].height = 20
        
        # Cabeçalho das colunas (usa linha dinâmica baseada no conteúdo adicionado)
        row = current_row + 1
        col = 1
        
        # Estilo do cabeçalho - cor de fundo azul com texto branco
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Coluna do aluno
        ws_chamadas.cell(row=row, column=col, value="ALUNO")
        ws_chamadas.cell(row=row, column=col).font = header_font
        ws_chamadas.cell(row=row, column=col).fill = header_fill
        ws_chamadas.cell(row=row, column=col).alignment = header_alignment
        col += 1
        
        # Colunas de datas (ordenadas cronologicamente)
        for date in all_dates:
            ws_chamadas.cell(row=row, column=col, value=date)
            ws_chamadas.cell(row=row, column=col).font = header_font
            ws_chamadas.cell(row=row, column=col).fill = header_fill
            ws_chamadas.cell(row=row, column=col).alignment = header_alignment
            col += 1
        
        # Coluna de observações
        ws_chamadas.cell(row=row, column=col, value="OBSERVAÇÕES")
        ws_chamadas.cell(row=row, column=col).font = header_font
        ws_chamadas.cell(row=row, column=col).fill = header_fill
        ws_chamadas.cell(row=row, column=col).alignment = header_alignment
        
        # Ajustar altura da linha do cabeçalho e larguras
        ws_chamadas.row_dimensions[row].height = 25
        ws_chamadas.column_dimensions['A'].width = 40  # ALUNO
        ws_chamadas.column_dimensions[get_column_letter(col)].width = 40  # OBSERVAÇÕES
        
        # Processa os dados por turma (usa linha dinâmica após o cabeçalho)
        row = current_row + 2
        
        # Mapeamento de alunos para turmas para evitar duplicidade
        aluno_to_turma = {}
        aluno_to_data = {}
        
        # Processa todos os alunos e atribui à turma mais recente
        for data in all_data:
            logger.info(f"Processando arquivo para consolidação: {data.get('filename', 'Desconhecido')}")
            logger.info(f"Total de alunos no arquivo: {len(data['alunos_data'])}")
            
            # Debug para os primeiros 5 alunos
            alunos_lista = list(data['alunos_data'].items())
            for i, (key, aluno_data) in enumerate(alunos_lista[:5]):
                logger.info(f"Exemplo de aluno {i+1}: {key} => {aluno_data}")
            
            for key, aluno_data in data['alunos_data'].items():
                nome_aluno = aluno_data.get('nome')
                if not nome_aluno:
                    # Tenta extrair o nome do aluno da chave composta
                    parts = key.split('|')
                    if len(parts) > 1:
                        nome_aluno = parts[1].strip()
                    else:
                        nome_aluno = key
                
                if nome_aluno and is_valid_student_name(nome_aluno):
                    turma = aluno_data.get('turma', 'Turma Não Identificada')
                    
                    # Registra o aluno sendo processado para debug
                    logger.info(f"Consolidando dados do aluno: {nome_aluno} da turma {turma}")
                    
                    # Verifica se o aluno já tem uma turma atribuída
                    if nome_aluno not in aluno_to_turma:
                        aluno_to_turma[nome_aluno] = turma
                        
                    # Cria estrutura para armazenar dados do aluno se não existir
                    if nome_aluno not in aluno_to_data:
                        aluno_to_data[nome_aluno] = {
                            'status': {},
                            'observacoes': []
                        }
                    
                    # Adiciona status para a data específica
                    if data['date']:
                        status = aluno_data.get('status', 'N/A')
                        logger.info(f"Adicionando status '{status}' para {nome_aluno} na data {data['date']}")
                        aluno_to_data[nome_aluno]['status'][data['date']] = status
                    
                    # Adiciona observação se existir
                    if aluno_data.get('observacao'):
                        obs = aluno_data.get('observacao').strip()
                        if obs and obs not in aluno_to_data[nome_aluno]['observacoes']:
                            if data['date']:
                                obs_com_data = f"({data['date']}): {obs}"
                                aluno_to_data[nome_aluno]['observacoes'].append(obs_com_data)
                                logger.info(f"Adicionada observação para {nome_aluno}: {obs_com_data}")
                            else:
                                aluno_to_data[nome_aluno]['observacoes'].append(obs)
                                logger.info(f"Adicionada observação para {nome_aluno}: {obs}")
                                
        # Vamos mapear todas as turmas que aparecem em qualquer arquivo
        # e coletar todos os alunos que já apareceram em qualquer dia nessa turma
        turmas_para_todos_alunos = {}
        total_alunos_registrados = 0
        
        # Primeiro passo: coletar todas as turmas e todos os alunos que já apareceram nelas
        for data in all_data:
            filename = data.get('filename', 'desconhecido')
            logger.info(f"Processando alunos de {filename}")
            logger.info(f"O arquivo tem {len(data.get('alunos_data', {}))} alunos registrados")
            
            for key, aluno_data in data['alunos_data'].items():
                turma = aluno_data.get('turma', 'Turma Não Identificada')
                nome_aluno = aluno_data.get('nome')
                
                # Tenta extrair o nome do aluno da chave composta (turma|nome)
                if not nome_aluno and '|' in key:
                    parts = key.split('|', 1)
                    if len(parts) > 1:
                        nome_aluno = parts[1].strip()
                        logger.info(f"Extraído nome do aluno da chave: {nome_aluno} (Chave: {key})")
                
                if not nome_aluno:
                    nome_aluno = key  # Assume que a própria chave é o nome
                    logger.warning(f"Usando chave como nome do aluno: {nome_aluno}")
                
                if is_valid_student_name(nome_aluno):
                    if turma not in turmas_para_todos_alunos:
                        turmas_para_todos_alunos[turma] = set()
                    turmas_para_todos_alunos[turma].add(nome_aluno)
                    total_alunos_registrados += 1
                else:
                    logger.warning(f"Nome de aluno inválido ignorado: {nome_aluno}")
        
        logger.info(f"Total de alunos registrados para exibição: {total_alunos_registrados}")
        logger.info(f"Total de turmas encontradas: {len(turmas_para_todos_alunos)}")
        
        # Lista as primeiras 5 turmas encontradas como exemplo
        turmas_exemplo = list(turmas_para_todos_alunos.keys())[:5]
        for turma in turmas_exemplo:
            logger.info(f"Turma: {turma} - {len(turmas_para_todos_alunos[turma])} alunos")
        
        # Agora imprime cada turma com todos os seus alunos (sem duplicatas)
        for turma in sorted_turmas:
            if turma not in turmas_para_todos_alunos or not turmas_para_todos_alunos[turma]:
                continue  # Pula turmas sem alunos válidos
                
            alunos_turma = sorted(turmas_para_todos_alunos[turma])
            
            # Log para debug
            logger.info(f"Processando turma: {turma} com {len(alunos_turma)} alunos")
            
            # Título da turma
            ws_chamadas.cell(row=row, column=1, value=f"TURMA: {turma}")
            ws_chamadas.cell(row=row, column=1).font = Font(bold=True)
            ws_chamadas.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(all_dates) + 2)
            row += 1
            
            # Adiciona cada aluno com seus dados
            for aluno in sorted(alunos_turma):
                if not aluno or not is_valid_student_name(aluno):
                    continue
                
                # Nome do aluno na primeira coluna
                ws_chamadas.cell(row=row, column=1, value=aluno)
                
                # Para cada data, procura o status do aluno usando a nova estrutura unificada
                col = 2
                
                # Registra se este aluno tem algum status para debug
                tem_algum_status = False
                
                for date in all_dates:
                    # Obtém o status dos dados consolidados - primeira tentativa com o dicionário
                    status = "N/A"
                    
                    # Verificação mais detalhada para encontrar o status
                    if aluno in aluno_to_data:
                        status_dict = aluno_to_data[aluno].get('status', {})
                        if date in status_dict:
                            status = status_dict[date]
                            tem_algum_status = True
                    
                    # Se não encontrou no dicionário, busca manualmente em todos os arquivos
                    if status == "N/A":
                        for data in all_data:
                            if data.get('date') == date:
                                # Busca em todos os alunos desse arquivo para essa data
                                for key, aluno_data in data.get('alunos_data', {}).items():
                                    nome_no_arquivo = aluno_data.get('nome', '')
                                    # Verifica se é o mesmo aluno diretamente pelo nome
                                    if nome_no_arquivo == aluno:
                                        status = aluno_data.get('status', 'N/A')
                                        tem_algum_status = True
                                        logger.info(f"Encontrado status direto para {aluno} na data {date}: {status}")
                                        break
                                    
                                    # Verifica também pela chave composta (turma|nome)
                                    elif '|' in key:
                                        parts = key.split('|', 1)
                                        if len(parts) > 1 and parts[1].strip() == aluno:
                                            status = aluno_data.get('status', 'N/A')
                                            tem_algum_status = True
                                            logger.info(f"Encontrado status via chave para {aluno} na data {date}: {status}")
                                            break
                    
                    # Adiciona status com formatação de cor
                    ws_chamadas.cell(row=row, column=col, value=status)
                    
                    # Formata cores com base no status
                    if status == "F":
                        ws_chamadas.cell(row=row, column=col).fill = PatternFill(
                            start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    elif status == "FJ":
                        ws_chamadas.cell(row=row, column=col).fill = PatternFill(
                            start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                    elif status == "P":
                        ws_chamadas.cell(row=row, column=col).fill = PatternFill(
                            start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                    
                    col += 1
                
                if not tem_algum_status:
                    logger.warning(f"ALERTA: Aluno {aluno} não tem nenhum status registrado")
                
                # Buscamos observações de duas formas: primeiro no dicionário consolidado
                observacoes = []
                if aluno in aluno_to_data and aluno_to_data[aluno].get('observacoes'):
                    observacoes = aluno_to_data[aluno]['observacoes']
                
                # Se não encontrou observações, busca diretamente nos arquivos
                if not observacoes:
                    for data in all_data:
                        for key, aluno_data in data.get('alunos_data', {}).items():
                            # Verifica se é o mesmo aluno (direto ou pela chave)
                            nome_aluno_arquivo = aluno_data.get('nome', '')
                            match_por_nome = nome_aluno_arquivo == aluno
                            match_por_chave = False
                            
                            if '|' in key:
                                parts = key.split('|', 1)
                                if len(parts) > 1 and parts[1].strip() == aluno:
                                    match_por_chave = True
                            
                            if match_por_nome or match_por_chave:
                                obs = aluno_data.get('observacao', '')
                                if obs and obs.strip():
                                    data_obs = data.get('date', '')
                                    if data_obs:
                                        obs_com_data = f"({data_obs}): {obs.strip()}"
                                        observacoes.append(obs_com_data)
                                    else:
                                        observacoes.append(obs.strip())
                
                # Remove duplicatas mantendo a ordem
                if observacoes:
                    obs_set = []
                    for obs in observacoes:
                        if obs and obs.strip() and obs.strip() not in [o.strip() for o in obs_set]:
                            obs_set.append(obs.strip())
                    
                    # Junta as observações com quebras de linha para melhor legibilidade
                    ws_chamadas.cell(row=row, column=col, value="\n".join(obs_set))
                    logger.info(f"Adicionadas {len(obs_set)} observações para {aluno}")
                
                row += 1
            
            # Espaço entre turmas
            row += 1
        
        # Ajusta largura das colunas
        ws_chamadas.column_dimensions['A'].width = 40  # Nome do aluno
        for i in range(len(all_dates)):
            col_letter = get_column_letter(i + 2)
            ws_chamadas.column_dimensions[col_letter].width = 15  # Datas
        
        # Coluna de observações - tamanho aumentado conforme solicitação do usuário
        last_col = get_column_letter(len(all_dates) + 2)
        ws_chamadas.column_dimensions[last_col].width = 60  # Tamanho aumentado para melhor visualização
        
        # Ajusta altura das linhas para textos - altura reduzida para um visual mais clean
        for r in range(5, row):  # Começa da linha 5 (após o cabeçalho) até a última linha usada
            ws_chamadas.row_dimensions[r].height = 22  # Altura moderada para um visual mais minimalista
        
        # Cria a aba de análise apenas se não houver arquivos específicos de análise
        # Quando há arquivos específicos, iremos criar abas personalizadas para cada um
        if not analysis_files:
            ws_analysis = wb.create_sheet(title="ANÁLISE")
            
            # Título principal
            ws_analysis.cell(row=1, column=1, value=f"ANÁLISE DE FREQUÊNCIA ESCOLAR")
            ws_analysis.cell(row=1, column=1).font = Font(bold=True, size=14)
            ws_analysis.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
            
            # Data de geração
            ws_analysis.cell(row=2, column=1, value=f"Relatório gerado em: {data_geracao}")
            ws_analysis.cell(row=2, column=1).font = Font(italic=True)
            ws_analysis.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
            
            # Mensagem de que não há arquivos de análise
            ws_analysis.cell(row=4, column=1, value="Nenhum arquivo de análise disponível.")
            ws_analysis.cell(row=4, column=1).font = Font(italic=True)
            ws_analysis.cell(row=6, column=1, value="Para incluir análises nesta aba, adicione arquivos de análise ao processar o relatório.")
            ws_analysis.merge_cells(start_row=6, start_column=1, end_row=6, end_column=5)
            
            # Configura largura das colunas para melhor visualização
            for col in range(1, 6):
                col_letter = get_column_letter(col)
                ws_analysis.column_dimensions[col_letter].width = 20
            
        # Aba RELATÓRIO removida conforme solicitação do usuário
        
        # Processa arquivos de análise (apenas o primeiro para evitar duplicação)
        logger.info(f"Processando {len(analysis_files)} arquivos de análise")
        
        # Tratamento especial para detectar o arquivo PequenoPrincipe.xlsx como análise
        # Este bloco força a inclusão dele na lista de análise se não foi detectado anteriormente
        if len(analysis_files) == 0:
            logger.info("Verificando se há arquivos de análise não detectados")
            for file_info in regular_files:
                filename = file_info.get('filename', '').lower()
                if "pequenoprincipe.xlsx" in filename.replace(" ", "") or "pequeno_principe.xlsx" in filename.replace(" ", ""):
                    logger.info(f"Encontrado arquivo de análise não detectado: {file_info.get('filename')}")
                    analysis_files.append(file_info)
                    # Remove da lista de arquivos regulares para não processar duas vezes
                    regular_files.remove(file_info)
                    break
            
            logger.info(f"Após verificação adicional: {len(analysis_files)} arquivos de análise")
        
        # Processa apenas o primeiro arquivo de análise para evitar duplicação
        if analysis_files:
            file_info = analysis_files[0]  # Usa apenas o primeiro arquivo
            logger.info(f"Processando apenas o primeiro arquivo de análise: {file_info.get('filename', '')}")
            try:
                filename = file_info.get('filename', '')
                content = file_info.get('content', b'')
                
                if not content:
                    logger.warning(f"Arquivo de análise vazio: {filename}")
                    return wb
                
                # Tenta processar o arquivo de análise em uma nova aba
                logger.info(f"Processando arquivo de análise: {filename}")
                with io.BytesIO(content) as buffer:
                    try:
                        # Lê o arquivo Excel
                        df = pd.read_excel(buffer)
                        logger.info(f"Arquivo lido com sucesso: {filename}, colunas: {df.columns.tolist()}")
                        
                        # Reposiciona o cursor do buffer para o início
                        buffer.seek(0)
                    except Exception as e:
                        logger.error(f"Erro ao ler arquivo de análise {filename}: {e}")
                        return wb
                    
                    # Extrai data para nome da aba
                    sheet_name = "ANÁLISE"
                    # Tenta extrair data do nome do arquivo
                    date_match = re.search(r'(\d{2})[-_](\d{2})', filename)
                    if date_match:
                        dia, mes = date_match.groups()
                        sheet_name = f"ANÁLISE {dia}-{mes}"
                    else:
                        # Se não encontrou no nome, tenta extrair do conteúdo
                        # Procura por datas no conteúdo do arquivo
                        for _, row in df.head(10).iterrows():
                            row_str = ' '.join([str(val) for val in row.values if pd.notna(val)])
                            date_match = re.search(r'(\d{2})/(\d{2})/\d{4}', row_str)
                            if date_match:
                                dia, mes = date_match.groups()
                                sheet_name = f"ANÁLISE {dia}-{mes}"
                                break
                    
                    # Limita o tamanho do nome da aba
                    if len(sheet_name) > 31:
                        sheet_name = sheet_name[:31]
                    
                    # Cria nova aba específica para este arquivo
                    # Verifica se já existe uma aba para este arquivo
                    # Evita duplicação - verifica se já existe uma aba ANÁLISE
                    analysis_exists = False
                    for existing_sheet in wb.sheetnames:
                        if "ANÁLISE" in existing_sheet.upper():
                            analysis_exists = True
                            logger.info(f"Aba ANÁLISE já existe: {existing_sheet}")
                            break
                    
                    # Se já existe, pula este arquivo para evitar duplicação
                    if analysis_exists:
                        logger.info(f"Pulando arquivo {filename} - aba ANÁLISE já existe")
                        return wb
                    
                    # Cria apenas uma aba ANÁLISE única
                    sheet_name = "ANÁLISE"
                    if len(sheet_name) > 31:
                        sheet_name = sheet_name[:31]
                
                    # Cria a aba para este arquivo de análise
                    ws_analysis_specific = wb.create_sheet(title=sheet_name)
                    
                    # Reposiciona o cursor do buffer para o início novamente
                    buffer.seek(0)
                    
                    # Título com nome do arquivo
                    short_filename = filename
                    if len(filename) > 50:
                        short_filename = filename[:47] + "..."
                    
                    ws_analysis_specific.cell(row=1, column=1, value=f"ANÁLISE DE FREQUÊNCIA - {short_filename}")
                    ws_analysis_specific.cell(row=1, column=1).font = Font(bold=True, size=14)
                    ws_analysis_specific.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(10, len(df.columns) + 2))
                    
                    # Data de geração
                    ws_analysis_specific.cell(row=2, column=1, value=f"Processado em: {data_geracao}")
                    ws_analysis_specific.cell(row=2, column=1).font = Font(italic=True)
                    ws_analysis_specific.merge_cells(start_row=2, start_column=1, end_row=2, end_column=min(10, len(df.columns) + 2))
                    
                    # Copia o arquivo de análise exatamente como está, sem modificações
                    try:
                        # Reposiciona o buffer para o início
                        buffer.seek(0)
                        
                        # Abre o arquivo original com openpyxl para cópia exata
                        original_wb = openpyxl.load_workbook(buffer)
                        original_ws = original_wb.active
                        
                        # Copia célula por célula mantendo formatação original
                        for row in original_ws.iter_rows():
                            for cell in row:
                                if cell.value is not None:
                                    target_cell = ws_analysis_specific.cell(row=cell.row, column=cell.column)
                                    target_cell.value = cell.value
                                    
                                    # Tenta copiar formatação original (se existir)
                                    try:
                                        if hasattr(cell, 'font') and cell.font:
                                            target_cell.font = copy(cell.font)
                                        if hasattr(cell, 'alignment') and cell.alignment:
                                            target_cell.alignment = copy(cell.alignment)
                                    except:
                                        pass  # Se não conseguir copiar formatação, continua
                        
                        # Copia larguras das colunas do arquivo original
                        try:
                            for col_letter, col_dim in original_ws.column_dimensions.items():
                                if hasattr(col_dim, 'width') and col_dim.width:
                                    ws_analysis_specific.column_dimensions[col_letter].width = col_dim.width
                        except:
                            # Se não conseguir copiar larguras, define padrão
                            for col in range(1, 20):
                                col_letter = get_column_letter(col)
                                ws_analysis_specific.column_dimensions[col_letter].width = 15
                                
                    except Exception as e:
                        logger.error(f"Erro ao copiar arquivo original: {e}")
                        # Fallback: copia apenas valores sem formatação
                        for r_idx, row in df.iterrows():
                            for c_idx, value in enumerate(row, 1):
                                if pd.notna(value):  # Só adiciona valores não-nulos
                                    ws_analysis_specific.cell(row=r_idx+1, column=c_idx, value=value)
                                    
                                    # Destaca nomes de colunas (primeira linha não-nula)
                                    if r_idx == 0 or (r_idx <= 5 and isinstance(value, str) and len(value) > 3):
                                        ws_analysis_specific.cell(row=r_idx+4, column=c_idx).font = Font(bold=True)
                    
                    # Ajusta larguras das colunas para visual mais clean e minimalista
                    for col in range(1, min(20, len(df.columns) + 3)):
                        col_letter = get_column_letter(col)
                        ws_analysis_specific.column_dimensions[col_letter].width = 15
            except Exception as e:
                logger.error(f"Erro ao processar arquivo de análise {filename}: {e}")
        
        # Cria aba de monitoramento (renomeada de "RECOMENDAÇÕES" para "MONITORAR")
        ws_recom = wb.create_sheet(title="MONITORAR")
        
        # Título com formatação melhorada
        ws_recom.cell(row=1, column=1, value="ALUNOS PRIORITÁRIOS PARA ACOMPANHAMENTO")
        ws_recom.cell(row=1, column=1).font = Font(bold=True, size=14, color="FFFFFF")
        ws_recom.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws_recom.cell(row=1, column=1).fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws_recom.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        
        # Aplica o estilo em todas as células mescladas do título
        for col in range(1, 7):
            if col > 1:  # A primeira coluna já foi formatada acima
                ws_recom.cell(row=1, column=col).fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                ws_recom.cell(row=1, column=col).font = Font(bold=True, size=14, color="FFFFFF")
        
        # Data com formatação melhorada
        ws_recom.cell(row=2, column=1, value=f"Gerado em: {data_geracao}")
        ws_recom.cell(row=2, column=1).font = Font(italic=True)
        ws_recom.cell(row=2, column=1).alignment = Alignment(horizontal='center')
        ws_recom.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        
        # Linha de espaçamento
        ws_recom.row_dimensions[3].height = 20
        
        # Cabeçalho das colunas com nome corrigido para FALTAS CHAMADAS
        headers = ["ALUNO", "TURMA", "FALTAS CHAMADAS", "STATUS ANÁLISE", "PRIORIDADE", "OBSERVAÇÃO"]
        
        # Definir larguras de coluna mais compactas para um visual mais clean
        column_widths = [35, 12, 15, 15, 12, 35]
        
        # Estilo do cabeçalho - cor de fundo azul com texto branco
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Configura cabeçalhos e larguras das colunas
        for col_idx, header in enumerate(headers, 1):
            # Configurar cabeçalho
            ws_recom.cell(row=4, column=col_idx, value=header)
            ws_recom.cell(row=4, column=col_idx).font = header_font
            ws_recom.cell(row=4, column=col_idx).fill = header_fill
            ws_recom.cell(row=4, column=col_idx).alignment = header_alignment
            
            # Configurar largura da coluna
            col_letter = get_column_letter(col_idx)
            ws_recom.column_dimensions[col_letter].width = column_widths[col_idx-1]
        
        # Altura da linha do cabeçalho
        ws_recom.row_dimensions[4].height = 25
        
        # Ajuste para a coluna de observações - aumentada conforme solicitação do usuário
        ws_recom.column_dimensions['F'].width = 60  # OBSERVAÇÃO - tamanho aumentado para melhor visualização
        
        # Coleta dados da aba ANÁLISE
        alunos_analise = {}
        
        # Procura a aba de ANÁLISE
        analysis_sheet = None
        for sheet_name in wb.sheetnames:
            if "ANÁLISE" in sheet_name.upper():
                analysis_sheet = wb[sheet_name]
                break
        
        # Extrai informações dos alunos que aparecem na aba ANÁLISE
        if analysis_sheet:
            logger.info("Extraindo dados da aba ANÁLISE para integração na aba RECOMENDAÇÕES")
            current_turma = None
            skip_header_rows = True  # Flag para pular linhas de cabeçalho
            header_count = 0
            
            for row in range(1, analysis_sheet.max_row + 1):
                # Verifica se a célula contém informação de turma
                cell_value = analysis_sheet.cell(row=row, column=1).value
                
                # Pula linhas de cabeçalho (primeiras 5 linhas ou até encontrar "Turma:")
                if skip_header_rows:
                    header_count += 1
                    if cell_value and isinstance(cell_value, str) and "turma:" in cell_value.lower():
                        skip_header_rows = False
                        current_turma = cell_value.replace("Turma:", "").strip()
                        logger.info(f"Primeira turma real encontrada: {current_turma}")
                        continue
                    elif header_count > 5:  # Força parar de pular após 5 linhas
                        skip_header_rows = False
                    else:
                        continue  # Continua pulando linhas de cabeçalho
                
                # Agora processa normalmente
                if cell_value and isinstance(cell_value, str) and "turma:" in cell_value.lower():
                    current_turma = cell_value.replace("Turma:", "").strip()
                    logger.info(f"Nova turma encontrada: {current_turma}")
                    continue
                
                # Se temos uma turma definida, procura por alunos
                if current_turma:
                    aluno_nome = analysis_sheet.cell(row=row, column=1).value
                    if aluno_nome and isinstance(aluno_nome, str) and len(aluno_nome) > 3:
                        # Verifica se não é um cabeçalho
                        if not any(keyword in aluno_nome.lower() for keyword in ['aluno', 'nome', 'turma', 'escola']):
                            # Procura status e informações adicionais nas próximas colunas
                            status = None
                            info_contato = None
                            percentual_presenca = None
                            
                            # Busca nas próximas colunas (até a coluna 10)
                            for col in range(2, min(10, analysis_sheet.max_column + 1)):
                                value = analysis_sheet.cell(row=row, column=col).value
                                
                                if value:
                                    if isinstance(value, str):
                                        col_text = value.lower()
                                        
                                        # Verifica se é um status válido (valor direto)
                                        status_keywords = ['faltoso', 'monitorar faltas', 'monitorar fjs', 'regular', 'muitas fjs', 'crítico', 'atenção']
                                        if any(status_keyword in col_text for status_keyword in status_keywords):
                                            status = value
                                        # Verifica cabeçalhos de status
                                        elif any(header_word in col_text for header_word in ['status', 'classificação', 'situação', 'monitorar']):
                                            # Se é um cabeçalho, pega o valor da linha atual
                                            header_value = analysis_sheet.cell(row=row, column=col).value
                                            if header_value and isinstance(header_value, str):
                                                status = header_value
                                        # Identifica campos de contato
                                        elif any(tel_word in col_text for tel_word in ['telefone', 'contato']):
                                            info_contato = value
                                    
                                    # Se não capturou status ainda, verifica se esta coluna tem cabeçalho de status
                                    if not status:
                                        # Verifica cabeçalhos nas linhas acima
                                        for header_row in range(max(1, row-5), row):
                                            header_cell = analysis_sheet.cell(row=header_row, column=col).value
                                            if header_cell and isinstance(header_cell, str):
                                                header_text = header_cell.lower()
                                                if any(h in header_text for h in ['status', 'classificação', 'situação']):
                                                    # Esta coluna é de status, então o valor atual é o status
                                                    if isinstance(value, str) and value.strip():
                                                        status = value
                                                    break
                                    
                                    # Verifica todas as colunas buscando percentual de presença
                                    # Primeiro verifica no header atual
                                    header_value = None
                                    
                                    # Verifica diversas linhas acima em busca do cabeçalho
                                    for header_row in range(1, 5):  # Verifica até 5 linhas acima
                                        if row - header_row >= 1:  # Evita índices negativos
                                            header_candidate = analysis_sheet.cell(row=row-header_row, column=col).value
                                            if header_candidate and isinstance(header_candidate, str) and any(pres in header_candidate.upper() for pres in ["% PRESENÇA", "PRESENÇA", "FREQUÊNCIA", "%"]):
                                                header_value = header_candidate
                                                break
                                    
                                    # Se encontrou um possível cabeçalho de presença
                                    if header_value:
                                        try:
                                            # Tenta extrair o valor numérico
                                            if isinstance(value, (int, float)):
                                                percentual_presenca = value
                                            elif isinstance(value, str):
                                                # Tenta extrair números da string, mesmo com outros caracteres
                                                value_clean = ''.join(c for c in value if c.isdigit() or c == '.')
                                                if value_clean:
                                                    # Verifica se há pelo menos um dígito
                                                    if any(c.isdigit() for c in value_clean):
                                                        try:
                                                            percentual_presenca = float(value_clean)
                                                            # Se o valor parece um percentual válido (entre 0 e 100)
                                                            if 0 <= percentual_presenca <= 100:
                                                                logger.info(f"Encontrado percentual de presença para {aluno_nome}: {percentual_presenca}%")
                                                            else:
                                                                percentual_presenca = None
                                                        except:
                                                            pass
                                        except Exception as e:
                                            logger.warning(f"Não foi possível extrair percentual de presença para {aluno_nome}: {str(e)}")
                            
                            # Armazena informações do aluno
                            aluno_key = aluno_nome.upper()
                            alunos_analise[aluno_key] = {
                                'turma': current_turma,
                                'status': status,
                                'info_contato': info_contato,
                                'percentual_presenca': percentual_presenca
                            }
                            
                            # Log para debug - mostra o que está sendo capturado
                            logger.info(f"CAPTURADO ALUNO DA ANÁLISE: {aluno_nome} | Status: {status} | Turma: {current_turma}")
        
        # Processa alunos para identificar faltas
        alunos_com_faltas = []
        
        for turma in sorted_turmas:
            alunos_turma = set()
            for data in all_data:
                for key, aluno_data in data['alunos_data'].items():
                    if aluno_data.get('turma') == turma:
                        alunos_turma.add(aluno_data.get('nome'))
            
            for aluno in alunos_turma:
                if not aluno:
                    continue
                
                # Contabiliza faltas por data
                faltas = 0
                total = 0
                aluno_key = f"{turma}|{aluno}"
                datas_falta = []  # Lista para guardar as datas específicas de falta
                
                # Implementa um contador realista para visitas
                # Cada data só pode contar uma vez para o total de visitas
                dates_visited = set()
                dates_absent = set()
                
                # Para cada arquivo de dados
                for data in all_data:
                    date = data['date']
                    # Verifica se o aluno existe nesta data
                    if aluno_key in data['alunos_data']:
                        # Só conta cada data uma vez para o total (evita duplicação)
                        if date not in dates_visited:
                            dates_visited.add(date)
                            total += 1
                            
                            status = data['alunos_data'][aluno_key].get('status')
                            if status in ['F', 'FJ']:
                                # Só conta cada falta uma vez (evita duplicação)
                                if date not in dates_absent:
                                    dates_absent.add(date)
                                    faltas += 1
                                    # Guarda a data da falta para mostrar na observação
                                    datas_falta.append(date)
                
                # Verifica se o aluno está na análise primeiro
                aluno_upper = aluno.upper()
                status_analise = None
                if aluno_upper in alunos_analise:
                    status_analise = alunos_analise[aluno_upper]['status']
                
                # Define prioridade baseada na nova ordem especificada:
                # 1. Faltou todos os dias da visita
                # 2. Status: Faltoso  
                # 3. Status: Muitas FJs
                # 4. Status: Monitorar Faltas
                # OBS: Se menos de 3 chamadas, não dar muita prioridade
                
                if total > 0:
                    percentual = (faltas / total) * 100
                    datas_formatadas = ", ".join(datas_falta)
                    
                    if total < 3:
                        # Pouca informação disponível (menos de 3 chamadas)
                        if faltas == total:  # Faltou todos os dias mesmo com poucas chamadas
                            prioridade = "MÉDIA"
                            obs = f"Faltou em todas as {total} visitas (poucos dados): {datas_formatadas}"
                        elif status_analise == "Faltoso":
                            prioridade = "MÉDIA" 
                            obs = f"Status na análise: Faltoso (poucos dados de chamada: {faltas}/{total})"
                        elif faltas > 0:
                            prioridade = "BAIXA"
                            obs = f"Faltou em {faltas} de {total} visitas (poucos dados)"
                        elif status_analise:
                            prioridade = "BAIXA"
                            obs = f"Status na análise: {status_analise} (poucos dados de chamada)"
                        else:
                            continue  # Sem dados suficientes para monitoramento
                    else:
                        # Dados suficientes para análise (3+ chamadas)
                        # Nova prioridade refinada:
                        # 1. CRÍTICA: Faltou todos os dias E possui status Faltoso
                        # 2. ALTA: Faltou todos os dias OU possui status Faltoso (mas não ambos)
                        # 3. MÉDIA: Faltou 3-4 dias, status Monitorar Faltas ou Monitorar FJs
                        # 4. BAIXA: outros casos
                        
                        # Lógica de prioridade mais consistente:
                        # 1. CRÍTICA: Faltou TODOS os dias (100%) E tem status Faltoso
                        # 2. ALTA: Faltou TODOS os dias (100%) OU tem status Faltoso (mas não ambos)
                        # 3. MÉDIA: Faltou mais de 50% das visitas OU tem status Monitorar Faltas/FJs
                        # 4. BAIXA: Outros casos com faltas ou status relevante
                        
                        if total > 0 and faltas == total and status_analise == "Faltoso":  # 1. CRÍTICA
                            prioridade = "CRÍTICA"
                            obs = f"CRÍTICO: Faltou todas as {total} visitas E classificado como Faltoso: {datas_formatadas}"
                        elif (total > 0 and faltas == total) or status_analise == "Faltoso":  # 2. ALTA
                            if total > 0 and faltas == total and status_analise != "Faltoso":
                                prioridade = "ALTA"
                                obs = f"Faltou em todas as {total} visitas: {datas_formatadas}"
                                if status_analise:
                                    obs += f" | Status: {status_analise}"
                            elif status_analise == "Faltoso" and (total == 0 or faltas < total):
                                prioridade = "ALTA"
                                if total > 0:
                                    obs = f"Status Faltoso na análise | Faltas: {faltas}/{total} visitas"
                                    if datas_formatadas:
                                        obs += f" ({datas_formatadas})"
                                else:
                                    obs = f"Status Faltoso na análise | Sem dados de chamadas"
                        elif (total > 0 and percentual > 50) or status_analise in ["Monitorar Faltas", "Monitorar FJs"]:  # 3. MÉDIA
                            if total > 0 and percentual > 50:
                                prioridade = "MÉDIA"
                                obs = f"Faltou mais de 50% das visitas: {faltas}/{total} ({percentual:.1f}%) - {datas_formatadas}"
                                if status_analise:
                                    obs += f" | Status: {status_analise}"
                            elif status_analise in ["Monitorar Faltas", "Monitorar FJs"]:
                                prioridade = "MÉDIA"
                                if total > 0:
                                    obs = f"Status '{status_analise}' na análise | Faltas: {faltas}/{total} visitas"
                                    if datas_formatadas:
                                        obs += f" ({datas_formatadas})"
                                else:
                                    obs = f"Status '{status_analise}' na análise | Sem dados de chamadas"
                        elif faltas > 0 or status_analise:  # 4. BAIXA
                            prioridade = "BAIXA"
                            if total > 0 and faltas > 0:
                                obs = f"Algumas faltas: {faltas}/{total} visitas ({percentual:.1f}%) - {datas_formatadas}"
                                if status_analise:
                                    obs += f" | Status: {status_analise}"
                            elif status_analise:
                                obs = f"Status na análise: {status_analise} | Sem faltas nas chamadas"
                            else:
                                obs = f"Faltas: {faltas}/{total} visitas ({datas_formatadas})"
                        else:
                            continue  # Sem faltas e não está na análise, não inclui
                else:
                    # Sem dados de chamadas, verifica apenas se está na análise
                    if status_analise:
                        prioridade = "BAIXA"
                        obs = f"Status na análise: {status_analise} | Sem dados de chamadas"
                    else:
                        continue  # Sem dados e não está na análise, não inclui
                
                # Adiciona informações extras da análise se disponível
                if aluno_upper in alunos_analise:
                    info_analise = alunos_analise[aluno_upper]
                    
                    # Adiciona informação da porcentagem de presença da aba análise se disponível
                    if info_analise.get('percentual_presenca') is not None:
                        obs += f" | Presença na análise: {info_analise['percentual_presenca']:.0f}%"
                    
                    # Adiciona informação de contato se disponível
                    if info_analise.get('info_contato'):
                        obs += f" | Contato: {info_analise['info_contato']}"
                
                alunos_com_faltas.append({
                    'aluno': aluno,
                    'turma': turma,
                    'faltas': faltas,
                    'total': total,
                    'percentual': percentual,
                    'prioridade': prioridade,
                    'status_analise': status_analise,
                    'obs': obs
                })
        
        # Adiciona alunos que estão APENAS na aba ANÁLISE
        for aluno_nome, info in alunos_analise.items():
            # Verifica se o aluno já está na lista (já identificado nas chamadas)
            if not any(a['aluno'].upper() == aluno_nome for a in alunos_com_faltas):
                # Adiciona o aluno que só aparece na ANÁLISE
                alunos_com_faltas.append({
                    'aluno': aluno_nome,
                    'turma': info['turma'],
                    'faltas': 0,        # Não temos dados de faltas específicas 
                    'total': 0,         # Não temos dados de total de visitas
                    'percentual': 0,
                    'prioridade': "ALTA",  # Prioridade alta por estar na ANÁLISE mas não nas CHAMADAS
                    'status_analise': info['status'],
                    'obs': f"Aluno possivelmente saiu da escola (consta apenas na aba ANÁLISE)" + 
                           (f" | Contato: {info['info_contato']}" if info['info_contato'] else "")
                })
        
        # Adiciona alunos prioritários à planilha, ordenados por prioridade
        # Nova ordem: CRÍTICA -> ALTA -> MÉDIA -> BAIXA
        row_recom = 5
        for aluno_info in sorted(
            alunos_com_faltas, 
            key=lambda x: (0 if x['prioridade'] == "CRÍTICA" else 
                           1 if x['prioridade'] == "ALTA" else 
                           2 if x['prioridade'] == "MÉDIA" else 3, 
                           x['percentual'] * -1)
        ):
            # Adiciona valores às células
            ws_recom.cell(row=row_recom, column=1, value=aluno_info['aluno'])
            ws_recom.cell(row=row_recom, column=2, value=aluno_info['turma'])
            ws_recom.cell(row=row_recom, column=3, value=f"{aluno_info['faltas']}/{aluno_info['total']}")
            ws_recom.cell(row=row_recom, column=4, value=aluno_info['status_analise'] if 'status_analise' in aluno_info and aluno_info['status_analise'] else "-")
            ws_recom.cell(row=row_recom, column=5, value=aluno_info['prioridade'])
            ws_recom.cell(row=row_recom, column=6, value=aluno_info['obs'])
            
            # Formatação das células de dados
            # Centraliza as colunas numéricas e de status
            ws_recom.cell(row=row_recom, column=2).alignment = Alignment(horizontal='center')
            ws_recom.cell(row=row_recom, column=3).alignment = Alignment(horizontal='center')
            ws_recom.cell(row=row_recom, column=5).alignment = Alignment(horizontal='center')
            
            # Habilita quebra automática de texto para a coluna de observações
            ws_recom.cell(row=row_recom, column=6).alignment = Alignment(wrap_text=True, vertical='top')
            
            # Ajusta altura da linha para comportar textos de várias linhas
            # Aumentando a altura para garantir que o texto não seja cortado
            ws_recom.row_dimensions[row_recom].height = 50
            
            # Estilo para as bordas (todas as células)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Aplica bordas em todas as células desta linha
            for col in range(1, 7):
                ws_recom.cell(row=row_recom, column=col).border = thin_border
            
            # Aplica cores APENAS para prioridade CRÍTICA - resto sem cor
            if aluno_info['prioridade'] == "CRÍTICA":
                # Fundo vermelho claro apenas para CRÍTICA
                fill_critical = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                for col in range(1, 7):
                    ws_recom.cell(row=row_recom, column=col).fill = fill_critical
            else:
                # Remove qualquer preenchimento para outras prioridades
                fill_none = PatternFill(fill_type=None)
                for col in range(1, 7):
                    ws_recom.cell(row=row_recom, column=col).fill = fill_none
            
            row_recom += 1
        
        # Salva o workbook
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Retorna o arquivo Excel junto com as informações extraídas do cabeçalho
        return {
            'excel_data': output.getvalue(),
            'school_name': school_name,
            'period': period
        }
        
    except Exception as e:
        logger.error(f"Erro ao gerar relatório melhorado: {str(e)}")
        # import traceback - não necessário
        # Removendo referência a traceback que não está importado
        logger.error("Erro ao gerar relatório (detalhes acima)")
        
        # Relatório de erro simples
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ERRO"
        
        ws.cell(row=1, column=1, value="Erro na Geração do Relatório")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        ws.cell(row=3, column=1, value=f"Erro: {str(e)}")
        ws.cell(row=5, column=1, value=f"Data: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()