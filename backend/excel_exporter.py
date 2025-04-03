from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import io
from datetime import datetime
import pytz
import re
import unicodedata

def export_to_excel(classes, attendance_status, observations, html_content=None, current_user=None, periodo=None, escola_nome=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "LISTA DE PRESENÇA"  # Título em maiúsculas

    # NOVO: Define o fuso horário de Brasília e obtém a data/hora atual
    br_tz = pytz.timezone('America/Sao_Paulo')
    current_time = datetime.now(br_tz).strftime('%d/%m/%Y %H:%M')

    # Cabeçalho com informações gerais
    header_rows = [
        ("UNIDADE:", escola_nome.upper() if escola_nome else "NÃO INFORMADO"),
        ("RESPONSÁVEIS:", current_user.upper() if current_user else "NÃO INFORMADO"),
        ("PERÍODO:", periodo.upper() if periodo else "NÃO INFORMADO"),
        ("DATA E HORA:", current_time)
    ]

    for i, (label, value) in enumerate(header_rows, start=1):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'B{i}'] = value
        ws.merge_cells(f'B{i}:D{i}')

    current_row = len(header_rows) + 2

    if not classes:
        ws.merge_cells(f"A{current_row}:D{current_row}")
        ws[f"A{current_row}"] = "NENHUMA CHAMADA SALVA ENCONTRADA"
        ws[f"A{current_row}"].font = Font(italic=True)
    else:
        for turma, alunos in classes.items():
            # Cabeçalho da turma (agora sem o nome da unidade entre parênteses)
            turma_display = turma.split('(')[0].strip() if '(' in turma else turma
            ws.merge_cells(f"A{current_row}:D{current_row}")
            ws[f"A{current_row}"] = f"TURMA: {turma_display.upper()}"
            ws[f"A{current_row}"].font = Font(bold=True, size=12)
            current_row += 1

            # Cabeçalho da tabela
            headers = ["ALUNO", "PRESENÇA", "OBSERVAÇÃO"]
            for col, header in enumerate(headers, start=1):
                ws.cell(row=current_row, column=col).value = header
                ws.cell(row=current_row, column=col).font = Font(bold=True)
            current_row += 1

            # Dados dos alunos
            for aluno in alunos:
                ws.cell(row=current_row, column=1).value = aluno
                ws.cell(row=current_row, column=2).value = attendance_status.get(turma, {}).get(aluno, "P")
                ws.cell(row=current_row, column=3).value = observations.get(turma, {}).get(aluno, "")
                current_row += 1

            current_row += 1  # Espaço entre turmas

    # Ajuste de colunas
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 40

    # Centralizar verticalmente e alinhar à esquerda todas as células
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='center', horizontal='left')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def get_excel_filename(escola_nome=None, periodo=None, current_user=None):
    """
    Gera o nome do arquivo no formato: NOME_DA_UNIDADE_DIA-MÊS-ANO_PERIODO_NOME_DA_DUPLA.xlsx
    """
    
    def sanitize(text, remove_hyphens=False):
        if not text or not isinstance(text, str):
            return ""
        
        # Remove símbolos específicos
        symbols_to_remove = ['º', 'ª', '°', '¨', '´', '`', '^', '~']
        for symbol in symbols_to_remove:
            text = text.replace(symbol, '')
        
        # Normaliza caracteres unicode (remove acentos)
        text = unicodedata.normalize('NFKD', text)
        text = ''.join([c for c in text if not unicodedata.combining(c)])
        
        # Padrão de caracteres permitidos
        pattern = r'[^\w\s]' if remove_hyphens else r'[^\w\s-]'
        text = re.sub(pattern, '', text)
        
        # Substitui todos os separadores por _
        text = re.sub(r'[\s\-\.]+', '_', text.strip())
        text = re.sub(r'_+', '_', text)
        text = text.strip('_')
        
        return text.upper()

    components = [
        sanitize(escola_nome, remove_hyphens=True) or "UNIDADE_NAO_INFORMADA",
        datetime.now().strftime('%d-%m-%Y'),
        sanitize(periodo) or "PERIODO_NAO_INFORMADO",
        sanitize(current_user) or "DUPLA_NAO_INFORMADA"
    ]
    
    filename = "_".join(filter(None, components)) + ".xlsx"
    return filename[:100]