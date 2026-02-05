import pandas as pd
import re
import os
import logging
import unicodedata
from io import BytesIO
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# Imports para PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

logger = logging.getLogger(__name__)

def make_hash_key(text):
    """
    Cria uma chave única simplificada.
    Remove acentos, espaços, pontuação e caracteres especiais.
    Mantém apenas LETRAS e NÚMEROS.
    """
    if not text or pd.isna(text): return ""
    text_str = str(text)
    normalized = unicodedata.normalize('NFD', text_str)
    shaved = ''.join(c for c in normalized if unicodedata.category(c) != 'Mn')
    shaved = shaved.upper()
    return re.sub(r'[^A-Z0-9]', '', shaved)

def clean_display_text(text):
    """Limpa texto apenas para exibição bonita no Excel."""
    if not text or pd.isna(text): return ""
    return str(text).strip().upper().replace('\xa0', ' ').replace('  ', ' ')

def generate_pdf_bytes(school_name, monitor_rows):
    """
    Gera o PDF oficial de encaminhamento de alunos prioritários.
    Retorna um buffer de bytes com o PDF.
    """
    buffer = BytesIO()
    # Margens reduzidas para 1cm para aproveitar melhor a largura da folha
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1*cm, leftMargin=1*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm
    )

    styles = getSampleStyleSheet()
    story = []

    # --- Estilos Personalizados ---
    style_header_right = ParagraphStyle(
        'HeaderRight',
        parent=styles['Normal'],
        alignment=2, # 0=Left, 1=Center, 2=Right
        fontSize=10,
        leading=12,
        fontName='Helvetica-Bold'
    )
    
    style_title_left = ParagraphStyle(
        'TitleLeft',
        parent=styles['Normal'],
        alignment=0,
        fontSize=10,
        leading=12,
        fontName='Helvetica-Bold'
    )

    style_normal = ParagraphStyle(
        'MyNormal',
        parent=styles['Normal'],
        fontSize=10,
        leading=12,
        spaceAfter=6
    )

    style_list = ParagraphStyle(
        'MyList',
        parent=styles['Normal'],
        fontSize=10,
        leading=12,
        leftIndent=20,
        spaceAfter=3
    )
    
    # Estilo para texto dentro da tabela
    style_table_text = ParagraphStyle(
        'TableText',
        parent=styles['Normal'],
        fontSize=9,
        leading=10,
        alignment=0 # Left
    )

    style_table_center = ParagraphStyle(
        'TableCenter',
        parent=styles['Normal'],
        fontSize=9,
        leading=10,
        alignment=1 # Center
    )

    # --- Cabeçalho Direito ---
    header_text = """
    PREFEITURA DE PALHOÇA<br/>
    SECRETARIA MUNICIPAL DE EDUCAÇÃO<br/>
    CENTRAL DE MATRÍCULAS
    """
    story.append(Paragraph(header_text, style_header_right))
    story.append(Spacer(1, 1*cm))

    # --- Conteúdo Inicial ---
    current_date = datetime.now().strftime("%d/%m/%Y")
    
    content_block_1 = f"""
    Data: {current_date}.<br/>
    À Direção da Escola: {school_name}.<br/>
    Assunto: Encaminhamento de alunos prioritários para busca ativa.
    """
    story.append(Paragraph(content_block_1, style_title_left))
    story.append(Spacer(1, 0.5*cm))

    text_body = """
    Prezada Direção,<br/><br/>
    Com base nos dados mais recentes coletados pelo Projeto Paestro, identificamos alunos faltosos nas
    análises de presença. Solicitamos, por gentileza, que seja realizada busca ativa junto às famílias para
    verificar os motivos das ausências e promover o retorno às aulas ou as devidas ações.<br/><br/>
    Segue abaixo a lista de alunos prioritários:
    """
    story.append(Paragraph(text_body, style_normal))
    story.append(Spacer(1, 0.5*cm))

    # --- Tabela ---
    headers = [
        Paragraph('<b>Turma</b>', style_table_text),
        Paragraph('<b>Aluno</b>', style_table_text),
        Paragraph('<b>Visitas</b>', style_table_center),
        Paragraph('<b>% Freq. Geral</b>', style_table_center),
        Paragraph('<b>Motivo/Providência</b>', style_table_text)
    ]
    table_data = [headers]

    rows_added = 0
    if monitor_rows:
        for row in monitor_rows:
            # 1. Lógica de Filtragem para PDF (RÍGIDA)
            st_upper = str(row.get('Status', '')).upper()
            
            # Verifica Status Prioritário: FALTOSO, MUITAS FJS ou FALTOU VISITAS
            # NOTA: Monitorar Faltas NÃO entra aqui se não tiver um desses status juntos
            is_priority_status = "FALTOSO" in st_upper or "MUITAS" in st_upper or "FALTOU VISITAS" in st_upper or "ABANDONO" in st_upper
            
            # Verifica se faltou em todos os dias E se o total de dias é >= 4 (redundância de segurança)
            faltas_str = str(row.get('Faltas', '0/0'))
            is_full_absence = False
            try:
                if '/' in faltas_str:
                    num, den = faltas_str.split('/')
                    # AQUI: Garante que só considera falta total se for 4/4, 5/5 ou maior.
                    if num == den and int(den) >= 4:
                        is_full_absence = True
            except: pass

            # Se não for nem status prioritário nem falta total (>=4), pula
            if not (is_priority_status or is_full_absence):
                continue

            rows_added += 1

            # 2. Lógica da Coluna % Freq. Geral
            # Pega o percentual base que já vem com as datas (ex: "0% (01/02 - 05/02)" ou "-")
            perc_display = str(row.get('Percent', ''))
            
            # Adiciona o status DEPOIS dos parênteses (ou do hífen) APENAS se for status prioritário
            if is_priority_status:
                # Recupera o status original para exibição
                raw_status = row.get('Status', '')
                perc_display = f"{perc_display} <b>{raw_status}</b>"
            
            # Cria Paragraphs
            turma_p = Paragraph(str(row.get('Turma', '')), style_table_text)
            aluno_p = Paragraph(str(row.get('Aluno', '')), style_table_text)
            faltas_p = Paragraph(faltas_str, style_table_center)
            perc_p = Paragraph(perc_display, style_table_center)
            
            line = [
                turma_p,
                aluno_p,
                faltas_p,
                perc_p,
                '' # Coluna vazia para preenchimento manual
            ]
            table_data.append(line)

    if rows_added == 0:
        table_data.append(['-', 'Nenhum aluno prioritário identificado conforme critérios.', '-', '-', '-'])

    # Configuração da Tabela
    # Largura total disponível A4 (21cm) - Margens (2cm total) = 19cm útil
    col_widths = [2.0*cm, 5.5*cm, 2.0*cm, 4.5*cm, 5.0*cm]
    
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    
    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    # --- Solicitação Final ---
    solicitacao_title = "<b>Solicitação</b>"
    story.append(Paragraph(solicitacao_title, style_normal))
    
    intro_sol = "Pedimos que a escola:"
    story.append(Paragraph(intro_sol, style_normal))

    items = [
        "1. Realize contato com os responsáveis pelos alunos listados;",
        "2. Verifique os motivos das ausências;",
        "3. Informar os motivos identificados para as ausências e as providências que serão adotadas para evitar novas faltas por parte dos alunos.",
        "4. Informe este setor em caso de necessidade de apoio adicional.",
        "5. Entregue os motivos e providências citados no 3º item dentro de sete dias úteis."
    ]

    for item in items:
        story.append(Paragraph(item, style_list))

    story.append(Spacer(1, 1.5*cm))

    # --- Assinatura ---
    footer_text = """
    SECRETARIA MUNICIPAL DE EDUCAÇÃO<br/>
    Central de Matrículas<br/>
    PAESTRO.
    """
    story.append(Paragraph(footer_text, style_title_left))

    doc.build(story)
    buffer.seek(0)
    return buffer

def process_consolidated_report(file_paths):
    attendance_files = []
    analysis_file_path = None
    
    date_pattern = re.compile(r'(\d{2}-\d{2}-\d{2,4})')
    
    # 1. Classificação
    for path in file_paths:
        filename = os.path.basename(path)
        match = date_pattern.search(filename)
        is_analysis = any(x in filename.lower() for x in ["analise", "análise", "situacao", "situação"])
        
        if match and not is_analysis:
            date_str = match.group(1)
            parts = date_str.split('-')
            if len(parts[2]) == 2: date_str = f"{parts[0]}-{parts[1]}-20{parts[2]}"
            
            attendance_files.append({
                'path': path, 
                'date': date_str, 
                'filename': filename
            })
        elif is_analysis or (not match and not analysis_file_path):
            analysis_file_path = path

    # Estruturas de Dados
    consolidated_data = {} 
    school_map = {}
    class_map = {}
    student_map = {}
    unit_annotations = {}
    all_dates = set()
    
    primary_school_hash = None 
    primary_school_name = "Escola Não Identificada"

    # ==========================================================================
    # 2. PROCESSAMENTO DAS CHAMADAS (VISITAS)
    # ==========================================================================
    for att_file in attendance_files:
        try:
            df = pd.read_excel(att_file['path'], header=None, engine='openpyxl')
            file_school_hash = None
            
            # --- Identificação da Escola ---
            for index, row in df.iterrows():
                row_vals = [str(x).strip() for x in row.values if pd.notna(x)]
                row_str = " ".join(row_vals).upper()
                
                if "UNIDADE:" in row_str:
                    for val in row_vals:
                        if val.upper().replace(":", "") != "UNIDADE" and len(val) > 3:
                            display = clean_display_text(val)
                            h_school = make_hash_key(display)
                            file_school_hash = h_school
                            
                            if not primary_school_hash:
                                primary_school_hash = h_school
                                primary_school_name = display
                                school_map[h_school] = display
                                if h_school not in consolidated_data: consolidated_data[h_school] = {}
                                if h_school not in unit_annotations: unit_annotations[h_school] = {}
                            break
                if file_school_hash: break 
            
            if not primary_school_hash:
                primary_school_hash = "ESCOLA_GENERICA"
                primary_school_name = "Unidade Escolar"
                school_map[primary_school_hash] = primary_school_name
                consolidated_data[primary_school_hash] = {}
                unit_annotations[primary_school_hash] = {}
            
            curr_school_hash = primary_school_hash
            
            curr_class_hash = None
            idx_aluno = -1
            idx_presenca = -1
            idx_obs = -1
            
            # --- Leitura das Linhas ---
            for index, row in df.iterrows():
                row_vals = [str(x).strip() for x in row.values if pd.notna(x)]
                row_str = " ".join(row_vals).upper()
                
                # Anotações
                if "•" in row_str:
                    date_key = att_file['date']
                    if date_key not in unit_annotations[curr_school_hash]:
                        unit_annotations[curr_school_hash][date_key] = []
                    clean_note = clean_display_text(row_str)
                    if clean_note not in unit_annotations[curr_school_hash][date_key]:
                        unit_annotations[curr_school_hash][date_key].append(clean_note)

                # Cabeçalho da Tabela
                row_hashes = [make_hash_key(x) for x in row.values if pd.notna(x)]
                if "ALUNO" in row_hashes and ("PRESENCA" in row_hashes or "PRESENÇA" in row_hashes):
                    idx_aluno = -1; idx_presenca = -1; idx_obs = -1
                    for i, cell in enumerate(row.values):
                        h_val = make_hash_key(cell)
                        if h_val == "ALUNO": idx_aluno = i
                        elif h_val in ["PRESENCA", "PRESENÇA"]: idx_presenca = i
                        elif "OBSERVA" in h_val: idx_obs = i
                    continue 

                # Identificação de Turma
                if "TURMA:" in row_str:
                    raw_turma = ""
                    for val in row_vals:
                        if "TURMA:" in val.upper():
                            parts = val.upper().split("TURMA:")
                            if len(parts) > 1 and parts[1].strip():
                                raw_turma = parts[1].strip()
                                break
                    if not raw_turma:
                        try:
                            orig_row = list(row.values)
                            for i, c in enumerate(orig_row):
                                if isinstance(c, str) and "TURMA:" in c.upper():
                                    if i+1 < len(orig_row) and pd.notna(orig_row[i+1]):
                                        raw_turma = str(orig_row[i+1]).strip()
                                        break
                        except: pass

                    if raw_turma:
                        display_class = clean_display_text(raw_turma)
                        h_class = make_hash_key(display_class)
                        curr_class_hash = h_class
                        if h_class not in class_map: class_map[h_class] = display_class
                        if h_class not in consolidated_data[curr_school_hash]:
                            consolidated_data[curr_school_hash][h_class] = {}

                # Dados do Aluno
                if curr_class_hash and idx_aluno != -1 and idx_presenca != -1:
                    try:
                        raw_aluno = row.iloc[idx_aluno]
                        raw_status = row.iloc[idx_presenca]
                        
                        if pd.notna(raw_aluno) and pd.notna(raw_status):
                            display_aluno = clean_display_text(raw_aluno)
                            h_aluno = make_hash_key(display_aluno)
                            
                            if h_aluno in ["ALUNO", "NOME", ""] or "TURMA" in h_aluno: continue
                            
                            status_clean = make_hash_key(raw_status)
                            status_display = "-"
                            # CORREÇÃO: Mais flexibilidade no mapeamento do status
                            if status_clean in ["P", "PRESENTE", "PRESENCA"]: status_display = "P"
                            elif status_clean in ["F", "FALTA", "AUSENTE"]: status_display = "F"
                            elif status_clean in ["FJ", "JUSTIFICADA"]: status_display = "FJ"
                            
                            if status_display != "-":
                                if h_aluno not in student_map: student_map[h_aluno] = display_aluno
                                
                                if h_aluno not in consolidated_data[curr_school_hash][curr_class_hash]:
                                    consolidated_data[curr_school_hash][curr_class_hash][h_aluno] = {
                                        'dates': {}, 'obs': []
                                    }
                                
                                consolidated_data[curr_school_hash][curr_class_hash][h_aluno]['dates'][att_file['date']] = status_display
                                all_dates.add(att_file['date'])
                                
                                if idx_obs != -1:
                                    raw_obs = row.iloc[idx_obs]
                                    if pd.notna(raw_obs):
                                        obs_txt = str(raw_obs).strip()
                                        if obs_txt and obs_txt.lower() != "nan":
                                            short_date = att_file['date'][:5]
                                            consolidated_data[curr_school_hash][curr_class_hash][h_aluno]['obs'].append(f"[{short_date}] {obs_txt}")

                    except IndexError: pass

        except Exception as e:
            logger.error(f"Erro ao ler arquivo {att_file['filename']}: {e}")

    sorted_dates = sorted(list(all_dates), key=lambda x: datetime.strptime(x, "%d-%m-%Y"))
    
    # Determina o período global das visitas para fallback
    visit_period_fallback = ""
    if sorted_dates:
        visit_period_fallback = f"({sorted_dates[0][:5]} - {sorted_dates[-1][:5]})"

    # ==========================================================================
    # 3. LEITURA DE DADOS DE ANÁLISE
    # ==========================================================================
    analysis_data_map = {} 
    
    # Mapa auxiliar: (h_class, h_aluno) -> {nome_real_aluno, nome_real_turma}
    analysis_names_map = {} 

    if analysis_file_path:
        try:
            # Aqui carregamos o arquivo COM openpyxl para preservar dados e estrutura se possível
            xls = pd.ExcelFile(analysis_file_path, engine='openpyxl')
            sheet_name = next((s for s in xls.sheet_names if "ANÁLISE" in s.upper() or "ANALISE" in s.upper()), None)
            
            if sheet_name:
                df_ana = pd.read_excel(analysis_file_path, sheet_name=sheet_name, header=None, engine='openpyxl')
                curr_ana_class_hash = None
                curr_ana_class_display = None
                
                for idx, row in df_ana.iterrows():
                    row_txt = " ".join([str(x) for x in row.values if pd.notna(x)])
                    
                    if "TURMA:" in row_txt.upper():
                        parts = row_txt.upper().split("TURMA:")
                        if len(parts) > 1:
                            # Captura nome real e hash da turma
                            raw_turma_ana = parts[1].strip()
                            curr_ana_class_display = clean_display_text(raw_turma_ana)
                            curr_ana_class_hash = make_hash_key(curr_ana_class_display)
                        continue
                    
                    if "MÉDIA GERAL" in row_txt.upper() or "ALUNO" in row_txt.upper(): continue
                    
                    if curr_ana_class_hash and pd.notna(row[0]):
                        raw_aluno_ana = str(row[0]).strip()
                        h_aluno = make_hash_key(raw_aluno_ana)
                        
                        if h_aluno == "ALUNO": continue
                        
                        if len(row) > 2:
                            st = str(row[1]).strip()
                            pc = str(row[2]).strip()
                            if any(x in st for x in ['Faltoso', 'Monitorar', 'Regular', 'Abandono', 'Ativo']):
                                key = (curr_ana_class_hash, h_aluno)
                                analysis_data_map[key] = {'status': st, 'percent': pc}
                                
                                # Guarda os nomes reais caso precisemos criar a entrada no consolidated_data
                                analysis_names_map[key] = {
                                    'aluno_display': clean_display_text(raw_aluno_ana),
                                    'turma_display': curr_ana_class_display
                                }
        except Exception as e:
            logger.error(f"Erro lendo analise: {e}")

    # ==========================================================================
    # 3.1 MERGE DE DADOS DA ANÁLISE PARA DADOS PRINCIPAIS
    # ==========================================================================
    target_school_hash = primary_school_hash if primary_school_hash else "ESCOLA_GENERICA"
    if target_school_hash not in consolidated_data:
        consolidated_data[target_school_hash] = {}
        primary_school_name = "Unidade Escolar (Análise)"

    for (h_class_ana, h_aluno_ana), ana_info in analysis_data_map.items():
        if h_class_ana not in consolidated_data[target_school_hash]:
            consolidated_data[target_school_hash][h_class_ana] = {}
            if h_class_ana not in class_map:
                class_map[h_class_ana] = analysis_names_map[(h_class_ana, h_aluno_ana)]['turma_display']

        if h_aluno_ana not in consolidated_data[target_school_hash][h_class_ana]:
            consolidated_data[target_school_hash][h_class_ana][h_aluno_ana] = {
                'dates': {}, 
                'obs': []
            }
            if h_aluno_ana not in student_map:
                student_map[h_aluno_ana] = analysis_names_map[(h_class_ana, h_aluno_ana)]['aluno_display']


    # ==========================================================================
    # 4. GERAÇÃO DO ARQUIVO FINAL
    # ==========================================================================
    
    # DEFINIÇÃO DE ESTILOS GERAIS
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    font_bold_white = Font(bold=True, color="FFFFFF")
    font_bold_black = Font(bold=True, color="000000")
    
    # Cores de Cabeçalho
    fill_blue_header = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid") # Azul Escuro Profissional
    fill_purple_header = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid") # Roxo para Situação
    
    # Cores de Status
    fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_red = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid") # Vermelho mais vivo
    fill_yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    font_green = Font(color="006100")
    font_red = Font(color="FFFFFF", bold=True) # Texto Branco no Vermelho
    font_yellow = Font(color="9C6500")

    # 1. Carregar arquivo original para manter a formatação da Análise
    if analysis_file_path:
        try: wb = load_workbook(analysis_file_path)
        except: wb = Workbook()
    else: wb = Workbook()

    for sheet in wb.sheetnames:
        if "ANÁLISE DE FALTAS" in sheet.upper():
            wb[sheet].title = "Análise"
        elif "ANALISE DE FALTAS" in sheet.upper():
            wb[sheet].title = "Análise"

    # --- PROCESSAMENTO ESPECIAL DA ABA ANÁLISE (REGRA DOS 2 MESES) ---
    # Aplica regra: Coluna é data nos últimos 60 dias? Se sim, F = Vermelho.
    if "Análise" in wb.sheetnames:
        ws_ana = wb["Análise"]
        today = datetime.now()
        sixty_days_ago = today - timedelta(days=60)
        
        # Itera sobre a primeira linha (cabeçalhos)
        for cell in ws_ana[1]:
            if cell.value:
                # Tenta detectar se é data
                col_date = None
                val_str = str(cell.value).strip()
                
                # Se já for datetime
                if isinstance(cell.value, datetime):
                    col_date = cell.value
                else:
                    # Tenta converter string DD/MM/AAAA ou DD/MM/AA
                    try:
                        # Regex simples para datas
                        if re.match(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', val_str):
                            parts = re.split(r'[/-]', val_str)
                            if len(parts) == 3:
                                d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                                if y < 100: y += 2000 # Assume século 21
                                col_date = datetime(y, m, d)
                    except: pass
                
                # Se for data válida e estiver no intervalo
                if col_date and sixty_days_ago <= col_date <= today:
                    col_letter = get_column_letter(cell.column)
                    # Define intervalo da coluna inteira (ex: E2:E1000)
                    range_str = f"{col_letter}2:{col_letter}{ws_ana.max_row}"
                    
                    # Adiciona regra condicional
                    ws_ana.conditional_formatting.add(range_str,
                        CellIsRule(operator='equal', formula=['"F"'], stopIfTrue=True, fill=fill_red, font=font_bold_white)
                    )

    # --- ABA CHAMADAS ---
    if "CHAMADAS" in wb.sheetnames: del wb["CHAMADAS"]
    ws_chamadas = wb.create_sheet("CHAMADAS", 0)
    
    # ... (Lógica da aba Chamadas mantida mas simplificada visualmente) ...
    row_idx = 1
    ws_chamadas.cell(row=row_idx, column=1, value=f"RELATÓRIO CONSOLIDADO - {primary_school_name}").font = Font(bold=True, size=14)
    row_idx += 2
    
    for h_school in sorted(consolidated_data.keys()):
        if h_school in unit_annotations and unit_annotations[h_school]:
            ws_chamadas.cell(row=row_idx, column=1, value="ANOTAÇÕES DA UNIDADE:").font = font_bold_black
            row_idx += 1
            def date_sorter(d):
                try: return datetime.strptime(d, "%d-%m-%Y")
                except: return datetime.min
            seen_notes = set()
            for d_key in sorted(unit_annotations[h_school].keys(), key=date_sorter):
                notes = unit_annotations[h_school][d_key]
                for n in notes:
                    clean_n = str(n).replace("•", "").strip()
                    note_unique = f"{d_key}|{clean_n}"
                    if note_unique not in seen_notes and clean_n:
                        ws_chamadas.cell(row=row_idx, column=1, value=f"({d_key}) {clean_n}")
                        row_idx += 1
                        seen_notes.add(note_unique)
            row_idx += 1
        
        turmas_sorted = sorted(consolidated_data[h_school].keys())
        for h_class in turmas_sorted:
            display_class_name = class_map.get(h_class, h_class)
            cell_turma = ws_chamadas.cell(row=row_idx, column=1, value=f"TURMA: {display_class_name}")
            cell_turma.font = font_bold_white
            cell_turma.fill = PatternFill(start_color="4F81BD", fill_type="solid") # Azul padrão Excel
            row_idx += 1
            
            headers = ["ALUNO"] + sorted_dates + ["Observações"]
            for c, val in enumerate(headers, 1):
                cell = ws_chamadas.cell(row=row_idx, column=c, value=val)
                cell.font = font_bold_black
                cell.alignment = Alignment(horizontal='center', vertical='center') # ALINHAMENTO CORRIGIDO
                cell.border = thin_border
                letter = get_column_letter(c)
                if c == 1: ws_chamadas.column_dimensions[letter].width = 40
                elif c == len(headers): ws_chamadas.column_dimensions[letter].width = 50
                else: ws_chamadas.column_dimensions[letter].width = 12
            row_idx += 1
            
            alunos_sorted = sorted(consolidated_data[h_school][h_class].keys())
            for h_aluno in alunos_sorted:
                display_aluno_name = student_map.get(h_aluno, h_aluno)
                data_dict = consolidated_data[h_school][h_class][h_aluno]
                c_nome = ws_chamadas.cell(row=row_idx, column=1, value=display_aluno_name)
                c_nome.border = thin_border
                c_nome.alignment = Alignment(horizontal='left', vertical='center')
                
                for d_i, date_key in enumerate(sorted_dates, 2):
                    status = data_dict['dates'].get(date_key, "-")
                    c_st = ws_chamadas.cell(row=row_idx, column=d_i, value=status)
                    c_st.alignment = Alignment(horizontal='center', vertical='center')
                    c_st.border = thin_border
                    if status == 'P':
                        c_st.fill = fill_green; c_st.font = font_green
                    elif status == 'F':
                        c_st.fill = fill_red; c_st.font = font_bold_white
                    elif status == 'FJ':
                        c_st.fill = fill_yellow; c_st.font = font_yellow
                
                obs_final = " | ".join(data_dict['obs'])
                c_obs = ws_chamadas.cell(row=row_idx, column=len(headers), value=obs_final)
                c_obs.border = thin_border
                c_obs.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                row_idx += 1
            row_idx += 2

    # --- GERAÇÃO DE DADOS PARA ABAS SEGUINTES ---
    monitor_rows = []
    situation_rows = []

    for h_school, turmas in consolidated_data.items():
        for h_class, alunos in turmas.items():
            for h_aluno, data_dict in alunos.items():
                frequencias = data_dict['dates']
                dias_com_registro = sum(1 for d in sorted_dates if frequencias.get(d) in ['P', 'F', 'FJ'])
                total_faltas = sum(1 for d in sorted_dates if frequencias.get(d) == 'F')
                faltou_todas_visitas = (dias_com_registro >= 4 and total_faltas == dias_com_registro)
                ana_info = analysis_data_map.get((h_class, h_aluno))
                
                st_display = pc = ""
                if ana_info:
                    st_display = ana_info['status']
                    pc = ana_info['percent']
                    if faltou_todas_visitas: st_display = f"{st_display} - Faltou Visitas"
                elif faltou_todas_visitas:
                    st_display = "Faltou Visitas"; pc = "-"
                
                if not st_display: continue
                st_upper = st_display.upper()
                
                # CRITÉRIOS MONITORAR
                include_monitor = False; prio = 0
                if "ABANDONO" in st_upper: include_monitor = True; prio = 100
                elif "FALTOSO" in st_upper: include_monitor = True; prio = 80
                elif "MUITAS" in st_upper: include_monitor = True; prio = 80
                elif "FALTOU VISITAS" in st_upper: include_monitor = True; prio = 75
                
                final_pc = pc
                if pc != "-":
                    if "REGULAR" not in st_upper and "(" not in pc:
                            final_pc = f"{pc} {visit_period_fallback}"
                    elif "VISITAS" in st_upper and "(" not in pc:
                            final_pc = f"{pc} {visit_period_fallback}"

                if include_monitor:
                    monitor_rows.append({
                        'Turma': class_map.get(h_class, h_class),
                        'Aluno': student_map.get(h_aluno, h_aluno),
                        'Faltas': f"{total_faltas}/{dias_com_registro}",
                        'RawFaltas': total_faltas,
                        'Status': st_display,
                        'Percent': final_pc,
                        'p': prio
                    })

                # CRITÉRIOS SITUAÇÃO
                include_situation = False
                if "ABANDONO" in st_upper: include_situation = True
                elif "FALTOSO" in st_upper: include_situation = True
                elif "MUITAS" in st_upper: include_situation = True
                elif "FALTOU VISITAS" in st_upper: include_situation = True
                elif "MONITORAR" in st_upper: include_situation = True
                elif "EXCESSO" in st_upper: include_situation = True
                
                if include_situation:
                    situation_rows.append({
                        'Turma': class_map.get(h_class, h_class),
                        'Aluno': student_map.get(h_aluno, h_aluno)
                    })

    monitor_rows.sort(key=lambda x: (-x['p'], -x['RawFaltas'], x['Turma'], x['Aluno']))
    
    # --- ABA MONITORAR (Design Melhorado) ---
    if "MONITORAR" in wb.sheetnames: del wb["MONITORAR"]
    ws_mon = wb.create_sheet("MONITORAR")
    
    m_headers = ["Turma", "Aluno", "Faltas Visitas", "Status (Análise)", "% Presença"]
    
    # Cabeçalho
    for i, h in enumerate(m_headers, 1):
        cell = ws_mon.cell(row=1, column=i, value=h)
        cell.font = font_bold_white
        cell.fill = fill_blue_header # Azul Profissional
        cell.alignment = Alignment(horizontal='center', vertical='center') # ALINHAMENTO CORRIGIDO
        cell.border = thin_border
        
        # Ajuste inicial de largura
        letter = get_column_letter(i)
        if "Aluno" in h: ws_mon.column_dimensions[letter].width = 40
        else: ws_mon.column_dimensions[letter].width = 20

    # Dados
    for r, d in enumerate(monitor_rows, 2):
        vals = [d['Turma'], d['Aluno'], d['Faltas'], d['Status'], d['Percent']]
        for c, val in enumerate(vals, 1):
            cell = ws_mon.cell(row=r, column=c, value=val)
            cell.border = thin_border # Borda em TUDO
            
            # Alinhamento: Aluno e Status à esquerda, resto centralizado
            if c == 2: # Aluno
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- ABA SITUAÇÃO ALUNOS  ---
    if "Situação Alunos" in wb.sheetnames: del wb["Situação Alunos"]
    ws_sit = wb.create_sheet("Situação Alunos BuscaAtiva")
    
    sit_headers = ["Turma", "Aluno", "Situação", "Apoia"]
    
    # Cabeçalho
    for col_num, header in enumerate(sit_headers, 1):
        cell = ws_sit.cell(row=1, column=col_num, value=header)
        cell.font = font_bold_white
        cell.fill = fill_purple_header
        cell.alignment = Alignment(horizontal='center', vertical='center') # ALINHAMENTO CORRIGIDO
        cell.border = thin_border
        
        letter = get_column_letter(col_num)
        if "Aluno" in header: ws_sit.column_dimensions[letter].width = 40
        else: ws_sit.column_dimensions[letter].width = 20
    
    # Validação de Dados
    dv1 = DataValidation(type="list", formula1='"Ativo,Transferido,Desistente,Infrequente"', allow_blank=False)
    dv2 = DataValidation(type="list", formula1='"Não,Sim"', allow_blank=False)
    ws_sit.add_data_validation(dv1); ws_sit.add_data_validation(dv2)
    
    situation_rows.sort(key=lambda x: (x['Turma'], x['Aluno']))
    
    for item in situation_rows:
        ws_sit.append([item['Turma'], item['Aluno'], "Ativo", "Não"])
        rn = ws_sit.max_row
        
        for c in range(1, 5): 
            cell = ws_sit.cell(row=rn, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if c == 2: cell.alignment = Alignment(horizontal='left', vertical='center')
            
        dv1.add(ws_sit.cell(row=rn, column=3)); dv2.add(ws_sit.cell(row=rn, column=4))
        
    # Formatação Condicional na Situação
    lr = ws_sit.max_row if ws_sit.max_row > 1 else 2
    # Ativo = Cinza Claro (limpo)
    ws_sit.conditional_formatting.add(f'C2:C{lr}', CellIsRule(operator='equal', formula=['"Ativo"'], fill=PatternFill(start_color='D9D9D9', fill_type='solid')))
    # Sim (Apoia) = Azul Claro
    ws_sit.conditional_formatting.add(f'D2:D{lr}', CellIsRule(operator='equal', formula=['"Sim"'], fill=PatternFill(start_color='00B0F0', fill_type='solid')))
    
    # --- TABELA DE ESTATÍSTICAS  ---
    
    # Título Principal
    ws_sit["G2"] = "Alunos Busca Ativa"
    ws_sit["G2"].font = font_bold_white
    ws_sit["G2"].fill = fill_purple_header
    ws_sit["G2"].alignment = Alignment(horizontal='center', vertical='center')
    ws_sit.merge_cells("G2:H2")
    ws_sit["G2"].border = thin_border
    ws_sit["H2"].border = thin_border

    # Cabeçalhos da Tabela
    ws_sit["G3"] = "Categoria"
    ws_sit["H3"] = "Qtd."
    ws_sit["G3"].font = font_bold_black
    ws_sit["H3"].font = font_bold_black
    ws_sit["G3"].fill = PatternFill(start_color="DDDDDD", fill_type="solid")
    ws_sit["H3"].fill = PatternFill(start_color="DDDDDD", fill_type="solid")
    ws_sit["G3"].alignment = Alignment(horizontal='center', vertical='center')
    ws_sit["H3"].alignment = Alignment(horizontal='center', vertical='center')
    ws_sit["G3"].border = thin_border
    ws_sit["H3"].border = thin_border

    # Dados
    current_stat_row = 4
    
    # 1. Total Geral
    ws_sit.cell(row=current_stat_row, column=7, value="Total de Alunos").border = thin_border
    ws_sit.cell(row=current_stat_row, column=8, value=f'=COUNTA(B2:B{lr})').border = thin_border
    ws_sit.cell(row=current_stat_row, column=8).alignment = Alignment(horizontal='center', vertical='center')
    current_stat_row += 1

    # 2. Situações (C)
    situations = [
        ("Ativos", "Ativo"),
        ("Transferidos", "Transferido"),
        ("Desistentes", "Desistente"),
        ("Infrequentes", "Infrequente")
    ]
    
    for label, criteria in situations:
        cell_lbl = ws_sit.cell(row=current_stat_row, column=7, value=label)
        cell_val = ws_sit.cell(row=current_stat_row, column=8, value=f'=COUNTIF(C2:C{lr}, "{criteria}")')
        
        cell_lbl.border = thin_border
        cell_val.border = thin_border
        cell_val.alignment = Alignment(horizontal='center', vertical='center')
        
        # Destaque para não ativos
        if criteria != "Ativo":
            cell_lbl.font = Font(color="555555") 
            
        current_stat_row += 1
    
    # 3. APOIA (D)
    cell_apoia_lbl = ws_sit.cell(row=current_stat_row, column=7, value="APOIA (Sim)")
    cell_apoia_val = ws_sit.cell(row=current_stat_row, column=8, value=f'=COUNTIF(D2:D{lr}, "Sim")')
    
    cell_apoia_lbl.font = Font(bold=True, color="2C3E50")
    cell_apoia_lbl.border = thin_border
    cell_apoia_lbl.fill = PatternFill(start_color="EBF5FB", fill_type="solid") 
    
    cell_apoia_val.font = Font(bold=True)
    cell_apoia_val.border = thin_border
    cell_apoia_val.alignment = Alignment(horizontal='center', vertical='center')
    cell_apoia_val.fill = PatternFill(start_color="EBF5FB", fill_type="solid")

    ws_sit.column_dimensions['G'].width = 22
    ws_sit.column_dimensions['H'].width = 10

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out, primary_school_name, monitor_rows