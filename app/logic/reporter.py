import io
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color, Protection
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime

logger = logging.getLogger(__name__)

# Utilitário para nome do mês
MONTH_NAMES = {
    1: 'Jan', 2: 'Fev', 3: 'Mar', 4: 'Abr', 
    5: 'Mai', 6: 'Jun', 7: 'Jul', 8: 'Ago',
    9: 'Set', 10: 'Out', 11: 'Nov', 12: 'Dez'
}
def get_month_name(m):
    return MONTH_NAMES.get(int(m), str(m))

def generate_consolidated_report(chamada_files, analise_files):
    # Função legada para compatibilidade
    pass

def calculate_class_stats(all_data_rows):
    """
    Calcula as estatísticas GERAIS da turma (incluindo alunos regulares),
    para exibir no rodapé do bloco.
    """
    stats = {}
    
    for row in all_data_rows:
        turma = row.get('turma', 'Turma Desconhecida')
        if turma not in stats:
            stats[turma] = {'P': 0, 'F': 0, 'FJ': 0}
            
        try:
            stats[turma]['P'] += int(row.get('P', 0))
            stats[turma]['F'] += int(row.get('F', 0))
            stats[turma]['FJ'] += int(row.get('FJ', 0))
        except: pass
        
    final_stats = {}
    for turma, totals in stats.items():
        p = totals['P']
        f = totals['F']
        fj = totals['FJ']
        
        total_efetivo = p + f
        total_oportunidades = p + f + fj
        
        perc_presenca = (p / total_efetivo * 100) if total_efetivo > 0 else 100.0
        perc_fj = (fj / total_oportunidades * 100) if total_oportunidades > 0 else 0.0
        
        final_stats[turma] = {
            'perc_presenca': round(perc_presenca, 1),
            'perc_fj': round(perc_fj, 1)
        }
        
    return final_stats

def generate_analysis_excel(data_rows, show_monthly_details=True, include_situation_tab=False): 
    """
    Gera Excel com formatação condicional, RichText, totais por turma
    e largura de coluna auto-ajustável.
    """
    wb = openpyxl.Workbook()
    
    # === ABA 1: ANÁLISE DE FALTAS ===
    ws = wb.active
    ws.title = "Análise de Faltas"
    
    # 1. Calcula estatísticas gerais ANTES de filtrar
    class_stats = calculate_class_stats(data_rows)
    
    # --- FILTRAGEM ---
    filtered_rows = []
    for row in data_rows:
        raw_status = row.get('status', [])
        status_list = []
        
        if isinstance(raw_status, str):
            status_list = [s.strip() for s in raw_status.split(',')]
        elif isinstance(raw_status, list):
            status_list = [str(s).strip() for s in raw_status]
        
        if not status_list: continue
        if len(status_list) == 1 and status_list[0].upper() == "REGULAR": continue
            
        filtered_rows.append(row)
    
    if not filtered_rows:
        ws['A1'] = "Nenhum aluno com excesso de faltas encontrado."
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    # 2. Configuração do Cabeçalho Geral
    escolas = set(r.get('escola', 'N/A') for r in filtered_rows)
    escolas_str = ", ".join(sorted(list(escolas)))
    data_hora = datetime.now().strftime("%d/%m/%Y %H:%M")
    
    ws['A1'] = f"Escola(s): {escolas_str}"
    ws['A1'].font = Font(bold=True, size=12)
    ws['A2'] = f"Data/Hora: {data_hora}"
    
    current_row = 4
    
    # 3. Organização por Turma
    dados_por_turma = {}
    for row in filtered_rows:
        turma = row.get('turma', 'Turma Desconhecida')
        if turma not in dados_por_turma: dados_por_turma[turma] = []
        dados_por_turma[turma].append(row)
        
    sorted_turmas = sorted(dados_por_turma.keys())
    
    # Cabeçalhos das Colunas
    headers = [
        "Aluno", "Status", "% Presença", "% FJ", "Total P", "Total F", "Total FJ"
    ]
    if show_monthly_details:
        headers.append("F por mês")
    headers.extend(["Nºs de contato", "Data do contato", "Motivo das faltas"])
    
    # Estilos
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    turma_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    footer_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # CRUCIAL: Definição da fonte vermelha para RichText. 'FF0000' é Hex para Vermelho.
    red_font = InlineFont(color='FF0000', b=True)
    
    # Variável para rastrear a largura máxima necessária da coluna "F por mês"
    max_month_col_width = 20
    
    for turma in sorted_turmas:
        # Título da Turma
        ws.cell(row=current_row, column=1, value=f"Turma: {turma}")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=1).fill = turma_fill
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        current_row += 1
        
        # Cabeçalhos
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=h)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        current_row += 1
        
        # Alunos
        alunos = sorted(dados_por_turma[turma], key=lambda x: x.get('aluno', ''))
        for aluno in alunos:
            rich_monthly_text = "N/A"
            current_text_len = 3 
            
            if show_monthly_details and aluno.get('faltas_por_mes'):
                monthly_dict = aluno['faltas_por_mes']
                if isinstance(monthly_dict, dict):
                    try:
                        # Garante que as chaves sejam inteiros para ordenar corretamente
                        sorted_months = sorted([int(k) for k in monthly_dict.keys()])
                    except:
                        sorted_months = []
                        
                    # Pega os dois últimos meses ativos para análise crítica
                    last_two = sorted_months[-2:] if len(sorted_months) >= 2 else sorted_months
                    
                    is_compulsory = aluno.get('is_compulsory', False)
                    # Limite: 10 para obrigatório (Fund/GT4/GT5), 12 para não obrigatório (GT0-GT3)
                    limit = 10 if is_compulsory else 12
                    
                    rich_string = CellRichText()
                    calculated_len = 0
                    
                    for i, m in enumerate(sorted_months):
                        # Tenta pegar pelo inteiro ou pela string da chave original
                        count = monthly_dict.get(m)
                        if count is None: count = monthly_dict.get(str(m), 0)
                            
                        if count > 0:
                            seg_text = f"{get_month_name(m)}:{count}"
                            if i < len(sorted_months) - 1:
                                seg_text += " "
                            
                            calculated_len += len(seg_text)
                            
                            # Lógica da cor vermelha: Se for um dos 2 últimos meses E exceder o limite
                            if m in last_two and count >= limit:
                                rich_string.append(TextBlock(red_font, seg_text))
                            else:
                                rich_string.append(seg_text)
                    
                    if calculated_len > 0:
                        rich_monthly_text = rich_string
                        current_text_len = calculated_len

            if current_text_len > max_month_col_width:
                max_month_col_width = current_text_len

            status_list = aluno.get('status', [])
            status_str = ", ".join(status_list) if isinstance(status_list, list) else str(status_list)
            
            # Formatação do texto da porcentagem com o período individual
            percent_val = aluno.get('percentual_presenca', 0)
            percent_display = f"{percent_val}%"
            
            # Adiciona o período apenas se o aluno não for "Regular"
            if status_str.upper() != "REGULAR":
                periodo_aluno = aluno.get('periodo', '')
                if periodo_aluno:
                    percent_display = f"{percent_display} {periodo_aluno}"

            vals = [
                aluno.get('aluno', ''),
                status_str,
                percent_display,
                f"{aluno.get('percentual_justificado', 0)}%",
                aluno.get('P', 0),
                aluno.get('F', 0),
                aluno.get('FJ', 0)
            ]
            
            for col_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.border = border
                if col_idx > 2 and col_idx < 8:
                    cell.alignment = Alignment(horizontal='center')
                if col_idx == 2:
                    upper_status = str(val).upper()
                    if "FALTOSO" in upper_status or "MUITAS FJS" in upper_status:
                        cell.font = Font(bold=True)

            if show_monthly_details:
                col_idx_fmes = 8
                cell_fmes = ws.cell(row=current_row, column=col_idx_fmes)
                cell_fmes.value = rich_monthly_text
                cell_fmes.border = border
                for i in range(1, 4):
                    ws.cell(row=current_row, column=col_idx_fmes + i, value="").border = border

            current_row += 1
            
        # Rodapé da Turma
        stats = class_stats.get(turma, {'perc_presenca': 0, 'perc_fj': 0})
        summary_text = f"Média Geral da Turma: Presença {stats['perc_presenca']}% | FJ {stats['perc_fj']}%"
        ws.cell(row=current_row, column=1, value=summary_text)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        footer_cell = ws.cell(row=current_row, column=1)
        footer_cell.font = Font(bold=True, italic=True, size=10)
        footer_cell.alignment = Alignment(horizontal='left')
        footer_cell.fill = footer_fill
        for c in range(1, len(headers) + 1):
            ws.cell(row=current_row, column=c).border = border
            
        current_row += 2
        
    # Ajuste Larguras Aba 1
    widths = [40, 30, 25, 12, 8, 8, 8]
    if show_monthly_details:
        widths.append(max_month_col_width * 1.2)
    widths.extend([20, 20, 40])
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out