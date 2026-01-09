import os
import io
import re
import logging
import pandas as pd
from datetime import datetime
from collections import defaultdict
from bs4 import BeautifulSoup
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import openpyxl
from copy import copy

# Configuração de Logger
logger = logging.getLogger(__name__)

# ==============================================================================
# SEÇÃO 1: UTILITÁRIOS E CONSTANTES
# ==============================================================================

MONTH_NAMES = {
    1: 'Jan', 2: 'Fev', 3: 'Mar', 4: 'Abr', 
    5: 'Mai', 6: 'Jun', 7: 'Jul', 8: 'Ago',
    9: 'Set', 10: 'Out', 11: 'Nov', 12: 'Dez'
}

def get_month_name(month_number):
    """Converte número do mês para nome abreviado em português."""
    try:
        month_int = int(month_number)
        if 1 <= month_int <= 12:
            return MONTH_NAMES.get(month_int)
    except (ValueError, TypeError):
        pass
    return str(month_number)

def get_batch_id():
    """Gera um ID de lote baseado na data e hora atual."""
    return datetime.now().strftime('%Y%m%d_%H%M%S')

def allowed_file(filename):
    """Verifica se a extensão do arquivo é permitida (HTML)."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'html', 'htm'}

# ==============================================================================
# SEÇÃO 2: PARSER HTML (Núcleo da Análise)
# ==============================================================================

def determine_education_type(class_name):
    """Determina o tipo de educação: obrigatória (fundamental) ou não (infantil)."""
    if not class_name:
        return {"nivel": "infantil", "obrigatorio": False}
        
    class_name = class_name.upper().strip()
    result = {}
    
    # GT0-GT5
    gt_match = re.search(r'GT\s*([0-5])', class_name)
    if gt_match:
        gt_num = int(gt_match.group(1))
        result["nivel"] = "infantil"
        result["obrigatorio"] = gt_num >= 4 # GT4 e GT5 são obrigatórias
        return result
    
    # Fundamental (1º-9º ANO)
    ano_match = re.search(r'([1-9])(º|°|\s)?(\s)*(ANO|SERIE|SÉRIE)', class_name)
    if ano_match:
        result["nivel"] = "fundamental"
        result["obrigatorio"] = True
        return result
    
    result["nivel"] = "infantil"
    result["obrigatorio"] = False
    return result

def get_school_info(html_content, filename=None):
    """Extrai escola e turma do HTML ou nome do arquivo."""
    result = {'unit_name': 'Não identificada', 'class_name': 'Não identificada', 'school_days': 0}
    
    try:
        soup = BeautifulSoup(html_content, 'html.parser')
        
        # 1. Busca Escola no Cabeçalho
        prefixos_escola = ['CEI ', 'CENTRO ', 'CAIC ', 'EBM ', 'EM ', 'IR ', 'ER ', 'GE ', 'EB ']
        escola_encontrada = False
        
        for table in soup.find_all('table')[:3]:
            if escola_encontrada: break
            for span in table.find_all('span'):
                text = span.get_text().strip().upper()
                if any(text.startswith(p) for p in prefixos_escola) and len(text) > 10:
                    palavras_proibidas = ['REGISTRO', 'RELATÓRIO', 'TURMA']
                    if not any(p in text for p in palavras_proibidas):
                        result['unit_name'] = text
                        escola_encontrada = True
                        break
        
        if not escola_encontrada and filename:
            result['unit_name'] = os.path.splitext(os.path.basename(filename))[0].upper()

        # 2. Busca Turma (Prioridade: Nome do Arquivo -> HTML)
        if filename:
            clean_name = os.path.splitext(os.path.basename(filename))[0].upper()
            gt_match = re.search(r'GT(\d+)([A-Z]?)', clean_name)
            ano_match = re.search(r'(\d+)ANO(\d*)', clean_name)
            
            if gt_match:
                result['class_name'] = f"GT{gt_match.group(1)}{gt_match.group(2) or ''}"
            elif ano_match:
                turma = ano_match.group(2) or ''
                result['class_name'] = f"{ano_match.group(1)}º ANO{' - ' + turma if turma else ''}"

        # Se não achou no arquivo, busca no HTML "TURMA:"
        if result['class_name'] == 'Não identificada':
            text_content = soup.get_text()
            turma_match = re.search(r'TURMA:\s*(GT\s*\d+\s*[A-Z]?|(\d+)[ºª°]?\s*ANO\s*[-–—]?\s*(\d+)?)', text_content, re.IGNORECASE)
            if turma_match:
                raw_turma = turma_match.group(1).strip().upper()
                # Normaliza GT
                if 'GT' in raw_turma:
                    result['class_name'] = raw_turma.replace(' ', '')
                else:
                    result['class_name'] = raw_turma

        # 3. Dias Letivos
        days_match = re.search(r'DIAS\s+LETIVOS\s*[:=]?\s*(\d+)', soup.get_text().upper())
        if days_match:
            result['school_days'] = int(days_match.group(1))

    except Exception as e:
        logger.warning(f"Erro ao extrair info da escola: {e}")
        
    return result

def get_student_list(html_content):
    """Extrai lista de alunos buscando padrões visuais (negrito/uppercase) em tabelas."""
    students = []
    try:
        soup = BeautifulSoup(html_content, 'html.parser')
        
        # Estratégia: Buscar spans que parecem nomes (Maiúsculas, >2 palavras)
        palavras_excluidas = [
            'PROFESSOR', 'ALUNO', 'TOTAL', 'PALHOÇA', 'CEI', 'CONVIVER', 'DIAS', 'LETIVOS',
            'MATEMÁTICA', 'PORTUGUESA', 'GEOGRAFIA', 'HISTÓRIA', 'CIÊNCIAS', 'ARTE'
        ]
        
        candidates = []
        # Coleta de células de tabelas grandes
        for table in soup.find_all('table'):
            rows = table.find_all('tr')
            if len(rows) > 5: # Tabelas grandes sugerem listas
                for row in rows:
                    cells = row.find_all('td')
                    if cells:
                        # Geralmente nome é a 1ª ou 2ª coluna com texto
                        for i in range(min(3, len(cells))):
                            text = cells[i].get_text().strip()
                            if text and len(text) > 5:
                                candidates.append(text)

        # Filtragem
        for text in candidates:
            clean = re.sub(r'\d+', '', text).strip() # Remove matrículas
            if (clean.isupper() and len(clean.split()) >= 2 and 
                not any(p in clean for p in palavras_excluidas) and ":" not in clean):
                students.append(clean)
        
        # Remove duplicatas preservando ordem
        return list(dict.fromkeys(students))
    except Exception as e:
        logger.error(f"Erro ao listar alunos: {e}")
        return []

def find_totals_in_html(html_content, student_name):
    """Busca totais P, F, FJ procurando tabelas que tenham esses cabeçalhos."""
    result = {'P': 0, 'F': 0, 'FJ': 0}
    soup = BeautifulSoup(html_content, 'html.parser')
    
    try:
        # Encontra cabeçalhos P, F, FJ
        headers = soup.find_all('span', string=lambda t: t in ['P', 'F', 'FJ'])
        
        for header in headers:
            # Tenta achar a tabela pai
            table = header.find_parent('table')
            if not table: continue
            
            # Acha a linha do aluno nesta tabela
            student_row = table.find('tr', string=lambda t: t and student_name in t)
            if not student_row: continue
            
            # Mapeia colunas
            cols = {}
            header_row = header.find_parent('tr')
            if header_row:
                for idx, cell in enumerate(header_row.find_all('td')):
                    txt = cell.get_text().strip()
                    if txt in ['P', 'F', 'FJ']:
                        cols[txt] = idx
            
            # Extrai dados
            if cols:
                cells = student_row.find_all('td')
                for key, idx in cols.items():
                    if idx < len(cells):
                        val = cells[idx].get_text().strip()
                        if val.isdigit():
                            result[key] = int(val)
                
                if any(result.values()): return result # Retorna se achou algo
                
    except Exception as e:
        logger.error(f"Erro ao buscar totais para {student_name}: {e}")
        
    return result

def process_student_attendance(html_content, student_name):
    """Conta faltas por mês analisando colunas de data (DD/MM) e marcações 'F'."""
    result = {'faltas_por_mes': defaultdict(int), 'maior_falta_mensal': 0, 'faltas_por_mes_texto': ''}
    soup = BeautifulSoup(html_content, 'html.parser')
    
    try:
        # 1. Mapear colunas de data
        date_cols = {} # {index: month_int}
        for table in soup.find_all('table'):
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all(['td', 'th'])
                for idx, cell in enumerate(cells):
                    txt = cell.get_text().strip()
                    match = re.search(r'(\d{1,2})/(\d{1,2})', txt)
                    if match:
                        date_cols[idx] = int(match.group(2)) # Salva o mês
        
        # 2. Achar 'F' nas linhas do aluno
        student_cells = soup.find_all(string=lambda t: t and student_name in str(t))
        for s_cell in student_cells:
            row = s_cell.find_parent('tr')
            if not row: continue
            
            cells = row.find_all('td')
            for idx, cell in enumerate(cells):
                if cell.get_text().strip() == 'F':
                    # Tenta inferir o mês pela coluna
                    if idx in date_cols:
                        month = date_cols[idx]
                        if month not in [1, 7]: # Ignora Jan/Jul (Férias)
                            result['faltas_por_mes'][month] += 1

        # Formatação
        if result['faltas_por_mes']:
            result['maior_falta_mensal'] = max(result['faltas_por_mes'].values())
            texto = []
            for m in sorted(result['faltas_por_mes'].keys()):
                texto.append(f"{get_month_name(m)}:{result['faltas_por_mes'][m]}")
            result['faltas_por_mes_texto'] = " ".join(texto)
        else:
            result['faltas_por_mes_texto'] = "Sem faltas mensais registradas"

    except Exception as e:
        logger.error(f"Erro processando faltas detalhadas: {e}")
        
    return result

def analyze_attendance_html(html_content, filename=None):
    """Orquestrador da análise de um único arquivo HTML."""
    try:
        info = get_school_info(html_content, filename)
        students = get_student_list(html_content)
        edu_info = determine_education_type(info['class_name'])
        
        school_data = {
            'school_name': info['unit_name'],
            'class_name': info['class_name'],
            'education_type': edu_info['nivel'],
            'is_compulsory': edu_info['obrigatorio'],
            'school_days': info['school_days']
        }
        
        student_data_list = []
        for name in students:
            totals = find_totals_in_html(html_content, name)
            attendance = process_student_attendance(html_content, name)
            
            # Cálculo de percentuais
            total = totals['P'] + totals['F'] + totals['FJ']
            perc_presenca = round((totals['P'] / total * 100), 1) if total > 0 else 0
            
            total_faltas = totals['F'] + totals['FJ']
            perc_justificadas = round((totals['FJ'] / total_faltas * 100), 1) if total_faltas > 0 else 0

            data = {
                'aluno': name,
                'name': name,
                'turma': info['class_name'],
                'escola': info['unit_name'],
                'P': totals['P'],
                'F': totals['F'],
                'FJ': totals['FJ'],
                'percentual_presenca': perc_presenca,
                'percentual_justificado': perc_justificadas,
                'faltas_por_mes': attendance['faltas_por_mes'],
                'faltas_por_mes_texto': attendance['faltas_por_mes_texto'],
                'education_type': edu_info['nivel'],
                'is_compulsory': edu_info['obrigatorio']
            }
            student_data_list.append(data)
            
        return {'school_data': school_data, 'student_data': student_data_list}

    except Exception as e:
        logger.error(f"Erro fatal analisando HTML {filename}: {e}")
        return None

# ==============================================================================
# SEÇÃO 3: MOTOR DE REGRAS (Rules Engine)
# ==============================================================================

def apply_classification_rules(data_wrapper):
    """Aplica regras de Busca Ativa (Faltoso, Monitorar, Regular)."""
    students_list = data_wrapper.get('student_data', [])
    classified = []
    
    for student in students_list:
        status = []
        
        # Definição de limites baseada na obrigatoriedade
        # Obrigatório: Limite 10 faltas/mês. Não Obrigatório: Limite 12.
        is_compulsory = student.get('is_compulsory', True)
        limit_absentee = 10 if is_compulsory else 12
        limit_monitor = 7 if is_compulsory else 10
        
        monthly_absences = student.get('faltas_por_mes', {})
        max_monthly = max(monthly_absences.values()) if monthly_absences else 0
        perc_presenca = student.get('percentual_presenca', 100)
        
        # Regra 1: Faltoso (Crítico)
        if max_monthly >= limit_absentee or (perc_presenca < 60 and perc_presenca > 0):
            status.append("Faltoso")
            
        # Regra 2: Monitorar Faltas (Alerta)
        elif max_monthly >= limit_monitor or (perc_presenca < 75 and perc_presenca > 0):
            status.append("Monitorar Faltas")
            
        # Regra 3: Monitorar Justificadas (Excesso de atestado)
        perc_just = student.get('percentual_justificado', 0)
        if perc_just >= 60:
            status.append("Muitas FJs")
        elif perc_just >= 45:
            status.append("Monitorar FJs")
            
        if not status:
            status.append("Regular")
            
        student['status'] = status
        student['situacao'] = status # Compatibilidade
        classified.append(student)
        
    return classified

# ==============================================================================
# SEÇÃO 4: PROCESSAMENTO DE ARQUIVOS (API Interface)
# ==============================================================================

def process_analysis_files(files):
    """
    Função principal chamada pelo routes.py.
    Processa múltiplos arquivos HTML e retorna dados consolidados.
    """
    batch_id = get_batch_id()
    logger.info(f"Iniciando lote de análise: {batch_id}")
    
    all_results = []
    processed_count = 0
    errors = []
    
    for file in files:
        if not allowed_file(file.filename):
            continue
            
        try:
            content = file.read().decode('utf-8', errors='ignore')
            result = analyze_attendance_html(content, file.filename)
            
            if result and result.get('student_data'):
                all_results.extend(result['student_data'])
                processed_count += 1
            else:
                errors.append(f"{file.filename}: Sem dados extraídos")
                
        except Exception as e:
            errors.append(f"{file.filename}: {str(e)}")
            logger.error(f"Erro arquivo {file.filename}: {e}")

    # Aplica regras em todos os resultados coletados
    final_data = apply_classification_rules({'student_data': all_results})
    
    return {
        'data': final_data,
        'summary': {
            'processed': processed_count,
            'total_students': len(final_data),
            'errors': errors
        }
    }

# ==============================================================================
# SEÇÃO 5: HELPER EXCEL (Para Leitura de Relatórios Anteriores)
# ==============================================================================

def detect_analysis_file(file_name):
    """Detecta se é arquivo de análise (Excel) pelo nome."""
    # Lógica simplificada: Se não tem padrão de data específico de chamada
    return not re.search(r'\d{2}-\d{2}-\d{4}', file_name)

def process_excel_analysis_file(ws, file_info):
    """Lê um Excel de análise e joga no worksheet de relatório."""
    try:
        content = file_info.get('content')
        if not content: return
        
        df = pd.read_excel(io.BytesIO(content))
        
        # Cabeçalho
        ws.cell(1, 1, "Análise de Frequência Importada").font = Font(bold=True, size=14)
        
        # Dados
        rows_to_write = [df.columns.tolist()] + df.values.tolist()
        for r_idx, row in enumerate(rows_to_write, 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(r_idx, c_idx, val)
                if r_idx == 2: # Cabeçalho da tabela
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="E6E6E6")
                    
    except Exception as e:
        logger.error(f"Erro lendo Excel de análise: {e}")