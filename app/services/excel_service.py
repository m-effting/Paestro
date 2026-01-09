from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import io
from datetime import datetime
import pytz
import re
import unicodedata
import logging

# Configuração de Logger
logger = logging.getLogger(__name__)

def normalize_school_name(name):
    """
    Função auxiliar para normalizar nomes de escolas, removendo acentos e espaços extras.
    Útil para comparar chaves de dicionário.
    """
    if not name:
        return ""
    # Remove acentos
    nfkd_form = unicodedata.normalize('NFKD', name)
    name_ascii = "".join([c for c in nfkd_form if not unicodedata.combining(c)])
    # Remove espaços extras e converte para maiúsculas
    return " ".join(name_ascii.upper().split())

def export_to_excel(classes, attendance_status, observations, html_content=None, current_user=None, periodo=None, escola_nome=None):
    """
    Gera um arquivo Excel com a lista de presença formatada.
    
    Args:
        classes (dict): Dicionário {turma: [lista_alunos]}
        attendance_status (dict): Dicionário de presenças
        observations (dict): Dicionário de observações
        html_content (str): Conteúdo HTML (opcional, mantido para compatibilidade)
        current_user (str): Nome do usuário/dupla responsável
        periodo (str): Período da chamada (Matutino/Vespertino)
        escola_nome (str): Nome da unidade escolar
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "LISTA DE PRESENÇA"

    br_tz = pytz.timezone('America/Sao_Paulo')
    current_time = datetime.now(br_tz).strftime('%d/%m/%Y %H:%M')

    # Validação de valores padrão para evitar "None" no Excel
    safe_escola = escola_nome.upper() if escola_nome else "NÃO INFORMADO"
    safe_user = current_user.upper() if current_user else "NÃO INFORMADO"
    safe_periodo = periodo.upper() if periodo else "NÃO INFORMADO"

    header_rows = [
        ("UNIDADE:", safe_escola),
        ("RESPONSÁVEIS:", safe_user),
        ("PERÍODO:", safe_periodo),
        ("DATA E HORA:", current_time)
    ]

    # Renderiza o cabeçalho fixo
    for i, (label, value) in enumerate(header_rows, start=1):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'B{i}'] = value
        ws.merge_cells(f'B{i}:D{i}')

    current_row = len(header_rows) + 2  # Pula uma linha após cabeçalho

    # --- Seção de Anotações (Requer lógica de anotações no app_data) ---
    # Nota: Se passar 'unit_annotations' via argumento no futuro, descomentar e adaptar aqui.
    
    if not classes:
        ws.merge_cells(f"A{current_row}:D{current_row}")
        ws[f"A{current_row}"] = "NENHUMA CHAMADA SALVA ENCONTRADA"
        ws[f"A{current_row}"].font = Font(italic=True, color="FF0000")
    else:
        # Ordena as turmas alfabeticamente para o relatório ficar bonito
        sorted_classes = sorted(classes.items())

        for turma, alunos in sorted_classes:
            # Limpa o nome da turma para exibição
            turma_display = turma.split('(')[0].strip() if '(' in turma else turma
            
            # Cabeçalho da Turma
            ws.merge_cells(f"A{current_row}:D{current_row}")
            ws[f"A{current_row}"] = f"TURMA: {turma_display.upper()}"
            ws[f"A{current_row}"].font = Font(bold=True, size=12, color="FFFFFF")
            
            # Estilo simples de fundo azul para o nome da turma
            from openpyxl.styles import PatternFill
            blue_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            ws[f"A{current_row}"].fill = blue_fill
            
            current_row += 1

            # Cabeçalhos das Colunas
            headers = ["ALUNO", "PRESENÇA", "OBSERVAÇÃO"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                # Fundo cinza claro para cabeçalho da tabela
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
            current_row += 1

            # Lista de Alunos
            for aluno in sorted(alunos): # Alunos em ordem alfabética
                ws.cell(row=current_row, column=1).value = aluno
                
                # Status (P, F, J)
                status = attendance_status.get(turma, {}).get(aluno, "P") # Padrão Presente
                ws.cell(row=current_row, column=2).value = status
                ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
                
                # Observações
                obs = observations.get(turma, {}).get(aluno, "")
                ws.cell(row=current_row, column=3).value = obs
                
                current_row += 1

            current_row += 2  # Espaço maior entre turmas

    # Ajuste de Largura das Colunas
    ws.column_dimensions["A"].width = 40  # Nome do Aluno
    ws.column_dimensions["B"].width = 15  # Presença
    ws.column_dimensions["C"].width = 50  # Observação

    # Ajuste de Alinhamento Global
    for row in ws.iter_rows(min_row=len(header_rows)+2):
        for cell in row:
            if not cell.alignment.horizontal: # Preserva alinhamento 'center' se já existir
                cell.alignment = Alignment(vertical='center', horizontal='left')

    # Salva em memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    logger.info(f"Arquivo Excel gerado para {safe_escola}")
    return output

def get_excel_filename(escola_nome=None, periodo=None, current_user=None):
    """
    Gera o nome do arquivo no formato padronizado: 
    NOME_DA_UNIDADE_DIA-MÊS-ANO_PERIODO_NOME_DA_DUPLA.xlsx
    """

    def sanitize(text, remove_hyphens=False):
        if not text or not isinstance(text, str):
            return ""

        # Remove símbolos específicos que dão erro em nomes de arquivo
        symbols_to_remove = ['º', 'ª', '°', '¨', '´', '`', '^', '~']
        for symbol in symbols_to_remove:
            text = text.replace(symbol, '')

        # Normaliza caracteres unicode (remove acentos: João -> Joao)
        text = unicodedata.normalize('NFKD', text)
        text = ''.join([c for c in text if not unicodedata.combining(c)])

        # Padrão de caracteres permitidos (apenas letras, números e traços)
        pattern = r'[^\w\s]' if remove_hyphens else r'[^\w\s-]'
        text = re.sub(pattern, '', text)

        # Substitui espaços e pontos por underline
        text = re.sub(r'[\s\-\.]+', '_', text.strip())
        text = re.sub(r'_+', '_', text) # Remove underlines duplicados
        text = text.strip('_')

        return text.upper()

    # Monta os componentes do nome
    components = [
        sanitize(escola_nome, remove_hyphens=True) or "UNIDADE_NAO_INFORMADA",
        datetime.now().strftime('%d-%m-%Y'),
        sanitize(periodo) or "PERIODO_NAO_INFORMADO",
        sanitize(current_user) or "DUPLA_NAO_INFORMADA"
    ]

    # Junta tudo com _, limita a 100 caracteres e adiciona extensão
    filename = "_".join(filter(None, components)) + ".xlsx"
    return filename[:100]